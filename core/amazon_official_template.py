"""Utilities for reading Amazon Seller Central spreadsheet templates."""

from __future__ import annotations

import base64
import json
import os
import re
from io import BytesIO
from typing import BinaryIO, Dict, Iterable, List, Optional
from urllib.parse import parse_qs, unquote

from openpyxl import load_workbook


AMAZON_TEMPLATE_SHEET = "Template"

_MARKETPLACE_BY_ID = {
    "ATVPDKIKX0DER": "US",
    "A2EUQ1WTGCTBG2": "CA",
    "A1AM78C64UM0Y8": "MX",
    "A1F83G8C2ARO7P": "UK",
    "A1PA6795UKMFR9": "DE",
    "A13V1IB3VIYZZH": "FR",
    "APJ6JRA9NG5V4": "IT",
    "A1RKKUPIHCS9HS": "ES",
    "A1VC38T7YXB528": "JP",
}

_PRODUCT_TYPE_RE = re.compile(r"^[A-Z][A-Z0-9_]{2,}$")
_VARIATION_PRONE_TERMS = (
    "SHOE",
    "BOOT",
    "SANDAL",
    "APPAREL",
    "PANTS",
    "SHIRT",
    "DRESS",
    "SOCK",
    "HAT",
    "BRA",
    "SWIM",
    "SKIRT",
    "SHORTS",
    "COAT",
    "JACKET",
)


def parse_official_amazon_template(file_obj: BinaryIO | BytesIO, filename: str = "") -> Dict[str, object]:
    """Parse a Seller Central generated workbook and infer its product_type."""
    wb = load_workbook(file_obj, read_only=True, data_only=True, keep_vba=True)
    try:
        sheetnames = list(wb.sheetnames)
        template_sheet_name = AMAZON_TEMPLATE_SHEET if AMAZON_TEMPLATE_SHEET in sheetnames else wb.active.title
        ws = wb[template_sheet_name]

        settings_text = str(ws.cell(row=1, column=1).value or "").strip()
        settings = _parse_template_settings(settings_text)
        columns = _extract_template_columns(ws, settings)

        candidates = []
        candidates.extend(_extract_product_types_from_settings(settings))
        candidates.extend(_extract_product_types_from_attribute_ptd_map(wb))
        candidates.extend(_extract_product_types_from_valid_values(wb))
        candidates.extend(_extract_product_types_from_filename(filename))

        product_type = _first_unique_product_type(candidates)
        marketplace_id = _extract_marketplace_id(settings)
        marketplace = _marketplace_code_from_id(marketplace_id)
        label_row = _safe_int(settings.get("labelRow"), 4)
        attribute_row = _safe_int(settings.get("attributeRow"), 5)
        data_row = _safe_int(settings.get("dataRow"), 7)
        example_product_type = _find_product_type_example(columns)
        has_variation_columns = any(
            _text_contains(item.get("attribute"), ("variation_theme", "parentage_level", "child_parent_sku_relationship"))
            or _text_contains(item.get("label"), ("variation", "parentage", "parent sku"))
            for item in columns
        )

        return {
            "is_official_amazon_template": _looks_like_official_template(sheetnames, settings_text, columns),
            "filename": filename or "",
            "sheetnames": sheetnames,
            "template_sheet": template_sheet_name,
            "product_type": product_type,
            "marketplace": marketplace,
            "marketplace_id": marketplace_id,
            "label_row": label_row,
            "attribute_row": attribute_row,
            "data_row": data_row,
            "column_count": len(columns),
            "columns": columns[:80],
            "product_type_column_example": example_product_type,
            "has_variation_columns": has_variation_columns,
            "suggested_variation_mode": _suggest_variation_mode(product_type),
            "detection_sources": _detection_sources(candidates),
            "warning": _build_warning(product_type, example_product_type),
        }
    finally:
        wb.close()


def looks_like_product_type(value: str) -> bool:
    return bool(_PRODUCT_TYPE_RE.match(str(value or "").strip()))


def _parse_template_settings(settings_text: str) -> Dict[str, str]:
    if not settings_text:
        return {}
    text = settings_text
    if text.startswith("settings="):
        text = text[len("settings="):]
    parsed = parse_qs(text, keep_blank_values=True)
    return {key: values[-1] for key, values in parsed.items() if values}


def _extract_product_types_from_settings(settings: Dict[str, str]) -> List[Dict[str, str]]:
    candidates = []
    decoded_ptds = _decode_base64_text(settings.get("ptds", ""))
    for product_type in _split_product_types(decoded_ptds):
        candidates.append({"product_type": product_type, "source": "settings.ptds"})

    decoded_browse = _decode_base64_text(settings.get("browseClassifications", ""))
    for product_type in _extract_product_types_from_json_text(decoded_browse):
        candidates.append({"product_type": product_type, "source": "settings.browseClassifications"})
    return candidates


def _extract_product_types_from_attribute_ptd_map(wb) -> List[Dict[str, str]]:
    if "AttributePTDMAP" not in wb.sheetnames:
        return []
    ws = wb["AttributePTDMAP"]
    found = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 80), max_col=min(ws.max_column or 1, 4), values_only=True):
        for value in row:
            product_type = _clean_product_type(value)
            if product_type:
                found.append({"product_type": product_type, "source": "AttributePTDMAP"})
    return found


def _extract_product_types_from_valid_values(wb) -> List[Dict[str, str]]:
    if "Valid Values" not in wb.sheetnames:
        return []
    ws = wb["Valid Values"]
    found = []
    max_row = min(ws.max_row or 1, 40)
    max_col = min(ws.max_column or 1, 20)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        row_text = " ".join(str(value or "") for value in row)
        if "Product Type" not in row_text:
            continue
        for value in row:
            product_type = _clean_product_type(value)
            if product_type:
                found.append({"product_type": product_type, "source": "Valid Values"})
    return found


def _extract_product_types_from_filename(filename: str) -> List[Dict[str, str]]:
    if not filename:
        return []
    stem = os.path.splitext(os.path.basename(filename))[0]
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem).strip()
    product_type = _clean_product_type(stem)
    return [{"product_type": product_type, "source": "filename"}] if product_type else []


def _extract_template_columns(ws, settings: Dict[str, str]) -> List[Dict[str, str]]:
    label_row = _safe_int(settings.get("labelRow"), 4)
    attribute_row = _safe_int(settings.get("attributeRow"), 5)
    sample_row = max(attribute_row + 1, 6)
    group_row = max(label_row - 1, 1)
    columns = []
    max_col = ws.max_column or 0
    for col_idx in range(1, max_col + 1):
        label = _cell_text(ws.cell(row=label_row, column=col_idx).value)
        attribute = _cell_text(ws.cell(row=attribute_row, column=col_idx).value)
        example = _cell_text(ws.cell(row=sample_row, column=col_idx).value)
        if not label and not attribute:
            continue
        columns.append({
            "index": col_idx,
            "label": label,
            "attribute": attribute,
            "example": example,
            "group": _cell_text(ws.cell(row=group_row, column=col_idx).value),
        })
    return columns


def _first_unique_product_type(candidates: Iterable[Dict[str, str]]) -> str:
    seen = set()
    for item in candidates:
        product_type = _clean_product_type(item.get("product_type"))
        if not product_type or product_type in seen:
            continue
        seen.add(product_type)
        return product_type
    return ""


def _detection_sources(candidates: Iterable[Dict[str, str]]) -> List[Dict[str, str]]:
    result = []
    seen = set()
    for item in candidates:
        product_type = _clean_product_type(item.get("product_type"))
        source = str(item.get("source") or "").strip()
        key = (product_type, source)
        if not product_type or key in seen:
            continue
        seen.add(key)
        result.append({"product_type": product_type, "source": source})
    return result[:12]


def _find_product_type_example(columns: List[Dict[str, str]]) -> str:
    for item in columns:
        label = str(item.get("label") or "").strip().lower()
        attribute = str(item.get("attribute") or "").strip().lower()
        if label == "product type" or attribute.startswith("product_type#"):
            return str(item.get("example") or "").strip()
    return ""


def _build_warning(product_type: str, example_product_type: str) -> str:
    if product_type and example_product_type and product_type != example_product_type:
        return (
            f"官方表中 Product Type 列的示例值是 {example_product_type}，"
            f"但模板ID是 {product_type}；系统会使用模板ID生成 2.8 模板。"
        )
    return ""


def _suggest_variation_mode(product_type: str) -> str:
    upper = str(product_type or "").upper()
    if any(term in upper for term in _VARIATION_PRONE_TERMS):
        return "variation"
    return "single"


def _looks_like_official_template(sheetnames: List[str], settings_text: str, columns: List[Dict[str, str]]) -> bool:
    if AMAZON_TEMPLATE_SHEET in sheetnames and ("settings=" in settings_text or columns):
        return True
    official_sheets = {"Instructions", "Data Definitions", "Valid Values"}
    return len(official_sheets.intersection(sheetnames)) >= 2


def _extract_marketplace_id(settings: Dict[str, str]) -> str:
    raw = str(settings.get("primaryMarketplaceId") or "").strip()
    if raw.startswith("amzn1.mp.o."):
        return raw.rsplit(".", 1)[-1]
    return raw


def _marketplace_code_from_id(marketplace_id: str) -> str:
    marketplace_id = str(marketplace_id or "").strip()
    for known_id, code in _MARKETPLACE_BY_ID.items():
        if known_id in marketplace_id:
            return code
    return "US"


def _decode_base64_text(value: str) -> str:
    text = unquote(str(value or "").strip())
    if not text:
        return ""
    padding = "=" * (-len(text) % 4)
    try:
        return base64.b64decode((text + padding).encode("utf-8")).decode("utf-8", "replace")
    except Exception:
        return ""


def _extract_product_types_from_json_text(text: str) -> List[str]:
    if not text:
        return []
    try:
        payload = json.loads(text)
    except Exception:
        return []
    result = []

    def walk(value):
        if isinstance(value, dict):
            for key, child in value.items():
                if key in {"productType", "product_type"}:
                    product_type = _clean_product_type(child)
                    if product_type:
                        result.append(product_type)
                walk(child)
        elif isinstance(value, list):
            for child in value:
                walk(child)

    walk(payload)
    return result


def _split_product_types(text: str) -> List[str]:
    if not text:
        return []
    if text.strip().startswith("["):
        try:
            payload = json.loads(text)
            return [pt for pt in (_clean_product_type(item) for item in payload) if pt]
        except Exception:
            pass
    return [pt for pt in (_clean_product_type(item) for item in re.split(r"[,;\s]+", text)) if pt]


def _clean_product_type(value) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    lowered = raw.lower()
    if lowered.startswith("product type") or any(marker in raw for marker in ("[", "]", "#", ".")):
        return ""
    text = raw.upper()
    text = re.sub(r"[^A-Z0-9_]+", "_", text).strip("_")
    return text if _PRODUCT_TYPE_RE.match(text) else ""


def _safe_int(value, default: int) -> int:
    try:
        return int(str(value or "").strip())
    except Exception:
        return default


def _cell_text(value) -> str:
    return str(value or "").strip()


def _text_contains(value, needles) -> bool:
    text = str(value or "").lower()
    return any(needle in text for needle in needles)
