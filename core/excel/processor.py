"""
Excel 读写处理器 V2
支持前后对比输出、字段分组、SP-API字段验证
"""
import os
import json
import logging
from typing import List, Dict, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from core.runtime_paths import resource_path, runtime_path

try:
    import xlrd
except ImportError:  # pragma: no cover - 依赖由 requirements 声明
    xlrd = None

logger = logging.getLogger(__name__)

# SP-API字段分组及颜色
FIELD_GROUPS = {
    'identity': {
        'title': '产品身份',
        'color': 'FFDC3545',  # 红
        'fields': ['SKU', 'brand', 'UPC', 'EAN', 'GTIN', 'manufacturer', 'model_number',
                   'product_type', 'item_type_keyword', 'external_product_id_type']
    },
    'content': {
        'title': 'AI内容(原始→生成)',
        'color': 'FFFD7E14',  # 橙
        'fields': ['item_name_original', 'item_name_ai',
                   'bullet_point_1_original', 'bullet_point_1_ai',
                   'bullet_point_2_original', 'bullet_point_2_ai',
                   'bullet_point_3_original', 'bullet_point_3_ai',
                   'bullet_point_4_original', 'bullet_point_4_ai',
                   'bullet_point_5_original', 'bullet_point_5_ai',
                   'description_original', 'description_ai',
                   'search_terms_ai']
    },
    'images': {
        'title': '图片',
        'color': 'FF28A745',  # 绿
        'fields': ['main_image_original', 'main_image_ai',
                   'image_2', 'image_3', 'image_4', 'image_5',
                   'image_6', 'image_7', 'image_8', 'image_9']
    },
    'offer': {
        'title': '销售条款',
        'color': 'FF007BFF',  # 蓝
        'fields': ['price', 'currency', 'list_price', 'sale_price', 'sale_from_date', 'sale_end_date',
                   'quantity', 'condition_type', 'fulfillment_channel',
                   'max_order_quantity', 'handling_time', 'merchant_shipping_group_name', 'product_tax_code']
    },
    'shipping': {
        'title': '物流',
        'color': 'FF6F42C1',  # 紫
        'fields': ['item_weight', 'item_weight_unit',
                   'item_length', 'item_width', 'item_height', 'dimension_unit',
                   'package_weight', 'package_weight_unit',
                   'package_length', 'package_width', 'package_height']
    },
    'attributes': {
        'title': '属性',
        'color': 'FF6C757D',  # 灰
        'fields': ['color', 'size', 'material', 'department',
                   'target_gender', 'age_range', 'country_of_origin',
                   'style_name', 'pattern_type', 'closure_type',
                   'number_of_items', 'unit_count', 'unit_count_type']
    },
    'compliance': {
        'title': '合规',
        'color': 'FF795548',  # 棕
        'fields': ['batteries_required', 'batteries_included',
                   'battery_type', 'number_of_batteries', 'battery_cell_composition',
                   'lithium_battery_packaging', 'lithium_battery_energy_content', 'lithium_battery_weight',
                   'hazmat_declaration', 'cpsia_cautionary_statement',
                   'safety_warning', 'legal_disclaimer']
    },
    'variation': {
        'title': '变体',
        'color': 'FF17A2B8',  # 青
        'fields': ['parent_sku', 'parentage_level', 'variation_theme']
    },
    'category_fields': {
        'title': '品类字段',
        'color': 'FFE91E63',  # 粉
        'fields': []  # 动态填充
    },
    'status': {
        'title': '上架状态',
        'color': 'FF343A40',  # 深灰
        'fields': ['submit_status', 'submit_time', 'asin', 'issues']
    },
}

# 列名别名映射: 各种中英文列名变体 -> 标准化内部字段名
COLUMN_ALIASES = {
    # SKU
    'sku': 'SKU', 'seller_sku': 'SKU', '卖家SKU': 'SKU', '商品SKU': 'SKU', 'Seller SKU': 'SKU',

    # Product Name / Title
    'item_name': 'item_name', 'title': 'item_name', '商品名称': 'item_name', '标题': 'item_name',
    'product_name': 'item_name', 'Product Name': 'item_name', 'Title': 'item_name', '产品名称': 'item_name',

    # Brand
    'brand': 'brand', 'brand_name': 'brand', '品牌': 'brand', '品牌名称': 'brand', 'Brand': 'brand', 'Brand Name': 'brand',

    # Description
    'description': 'description', 'product_description': 'description', '描述': 'description',
    '商品描述': 'description', '产品描述': 'description', 'Description': 'description', 'Product Description': 'description',

    # Bullet Points
    'bullet_point_1': 'bullet_point_1', 'bullet_point1': 'bullet_point_1', '要点1': 'bullet_point_1', '卖点1': 'bullet_point_1',
    'bullet_point_2': 'bullet_point_2', 'bullet_point2': 'bullet_point_2', '要点2': 'bullet_point_2', '卖点2': 'bullet_point_2',
    'bullet_point_3': 'bullet_point_3', 'bullet_point3': 'bullet_point_3', '要点3': 'bullet_point_3', '卖点3': 'bullet_point_3',
    'bullet_point_4': 'bullet_point_4', 'bullet_point4': 'bullet_point_4', '要点4': 'bullet_point_4', '卖点4': 'bullet_point_4',
    'bullet_point_5': 'bullet_point_5', 'bullet_point5': 'bullet_point_5', '要点5': 'bullet_point_5', '卖点5': 'bullet_point_5',

    # Price
    'price': 'price', 'standard_price': 'price', '价格': 'price', '售价': 'price', 'Price': 'price',
    'currency': 'currency', '币种': 'currency',

    # Images
    'main_image': 'main_image', 'main_image_url': 'main_image', '主图': 'main_image', '主图链接': 'main_image', 'Main Image URL': 'main_image',
    'image_2': 'image_2', 'other_image_url1': 'image_2', '副图1': 'image_2', '图片2': 'image_2',
    'image_3': 'image_3', 'other_image_url2': 'image_3', '副图2': 'image_3', '图片3': 'image_3',
    'image_4': 'image_4', 'other_image_url3': 'image_4', '副图3': 'image_4', '图片4': 'image_4',
    'image_5': 'image_5', 'other_image_url4': 'image_5', '副图4': 'image_5', '图片5': 'image_5',
    'image_6': 'image_6', 'other_image_url5': 'image_6', '副图5': 'image_6', '图片6': 'image_6',
    'image_7': 'image_7', 'other_image_url6': 'image_7', '副图6': 'image_7', '图片7': 'image_7',
    'image_8': 'image_8', 'other_image_url7': 'image_8', '副图7': 'image_8', '图片8': 'image_8',
    'image_9': 'image_9', 'other_image_url8': 'image_9', '副图8': 'image_9', '图片9': 'image_9',

    # Quantity
    'quantity': 'quantity', 'stock': 'quantity', '库存': 'quantity', '数量': 'quantity', 'Quantity': 'quantity',

    # UPC/EAN/GTIN
    'upc': 'UPC', 'UPC': 'UPC', 'product_id': 'UPC', '条形码': 'UPC',
    'ean': 'EAN', 'EAN': 'EAN',
    'gtin': 'GTIN', 'GTIN': 'GTIN',

    # Product Type
    'product_type': 'product_type', '产品类型': 'product_type', 'Product Type': 'product_type',
    'feed_product_type': 'product_type', '商品类型': 'product_type',

    # Condition
    'condition_type': 'condition_type', '状况': 'condition_type', 'Condition': 'condition_type',

    # Fulfillment
    'fulfillment_channel': 'fulfillment_channel', '配送方式': 'fulfillment_channel', 'Fulfillment Channel': 'fulfillment_channel',

    # Variant fields
    'parent_sku': 'parent_sku', '父SKU': 'parent_sku', 'Parent SKU': 'parent_sku',
    'parentage_level': 'parentage_level', '父子关系': 'parentage_level', 'Parentage': 'parentage_level', 'parent_child': 'parentage_level',
    'variation_theme': 'variation_theme', '变体主题': 'variation_theme', 'Variation Theme': 'variation_theme',
    'color': 'color', '颜色': 'color', 'Color': 'color', 'color_name': 'color',
    'size': 'size', '尺码': 'size', '尺寸': 'size', 'Size': 'size', 'size_name': 'size',

    # Search Terms
    'search_terms': 'search_terms', '搜索词': 'search_terms', 'Search Terms': 'search_terms',
    'generic_keywords': 'search_terms', '关键词': 'search_terms',

    # Weight/Dimensions
    'item_weight': 'item_weight', '重量': 'item_weight', 'Weight': 'item_weight',
    'item_weight_unit': 'item_weight_unit', '重量单位': 'item_weight_unit', 'Weight Unit': 'item_weight_unit',
    'item_length': 'item_length', '长度': 'item_length',
    'item_width': 'item_width', '宽度': 'item_width',
    'item_height': 'item_height', '高度': 'item_height',
    'dimension_unit': 'dimension_unit', '尺寸单位': 'dimension_unit', 'Dimension Unit': 'dimension_unit',
    'package_weight': 'package_weight', '包装重量': 'package_weight', 'Package Weight': 'package_weight',
    'package_weight_unit': 'package_weight_unit', '包装重量单位': 'package_weight_unit',
    'package_length': 'package_length', '包装长度': 'package_length', 'Package Length': 'package_length',
    'package_width': 'package_width', '包装宽度': 'package_width', 'Package Width': 'package_width',
    'package_height': 'package_height', '包装高度': 'package_height', 'Package Height': 'package_height',

    # Attributes
    'target_gender': 'target_gender', '目标性别': 'target_gender', 'Target Gender': 'target_gender',
    'age_range': 'age_range', '年龄范围': 'age_range', 'Age Range': 'age_range',
    'department': 'department', '部门': 'department', 'Department': 'department',
    'material': 'material', '材质': 'material', 'Material': 'material', 'material_type': 'material',

    # Manufacturer
    'manufacturer': 'manufacturer', '制造商': 'manufacturer', 'Manufacturer': 'manufacturer',
    'model_number': 'model_number', '型号': 'model_number', 'Model Number': 'model_number',
    'part_number': 'model_number', 'Part Number': 'model_number',

    # Offer extras
    'sale_price': 'sale_price', '促销价': 'sale_price', 'Sale Price': 'sale_price',
    'sale_from_date': 'sale_from_date', '促销开始': 'sale_from_date', 'Sale Start Date': 'sale_from_date',
    'sale_end_date': 'sale_end_date', '促销结束': 'sale_end_date', 'Sale End Date': 'sale_end_date',
    'list_price': 'list_price', '建议零售价': 'list_price', 'List Price': 'list_price', 'msrp': 'list_price',
    'max_order_quantity': 'max_order_quantity', '最大订购量': 'max_order_quantity',
    'handling_time': 'handling_time', 'leadtime_to_ship': 'handling_time', '发货时间': 'handling_time',
    'merchant_shipping_group_name': 'merchant_shipping_group_name', '配送模板': 'merchant_shipping_group_name',
    'product_tax_code': 'product_tax_code', '税码': 'product_tax_code',

    # Identity extras
    'item_type_keyword': 'item_type_keyword', '商品类型关键词': 'item_type_keyword',
    'external_product_id_type': 'external_product_id_type', '产品ID类型': 'external_product_id_type',
    'product_identity_mode': 'product_identity_mode', '商品标识模式': 'product_identity_mode',
    'number_of_items': 'number_of_items', '包装件数': 'number_of_items',

    # Compliance extras
    'cpsia_cautionary_statement': 'cpsia_cautionary_statement', 'CPSIA': 'cpsia_cautionary_statement',
    'legal_disclaimer': 'legal_disclaimer', '法律声明': 'legal_disclaimer',
    'safety_warning': 'safety_warning', '安全警告': 'safety_warning',
    'number_of_batteries': 'number_of_batteries', '电池数量': 'number_of_batteries',
    'battery_cell_composition': 'battery_cell_composition', '电池成分': 'battery_cell_composition',
    'lithium_battery_packaging': 'lithium_battery_packaging', '锂电池包装': 'lithium_battery_packaging',
    'lithium_battery_energy_content': 'lithium_battery_energy_content',
    'lithium_battery_weight': 'lithium_battery_weight',

    # Attribute extras
    'style_name': 'style_name', '风格': 'style_name', 'Style': 'style_name',
    'pattern_type': 'pattern_type', '图案': 'pattern_type', 'Pattern': 'pattern_type', 'pattern_name': 'pattern_type',
    'closure_type': 'closure_type', '闭合类型': 'closure_type', 'Closure Type': 'closure_type',
    'unit_count': 'unit_count', '单位数量': 'unit_count',
    'unit_count_type': 'unit_count_type', '单位类型': 'unit_count_type',
}

# COLUMN_ALIASES标准名 -> detect_columns标准名 的桥接映射
_ALIAS_TARGET_TO_DETECT = {
    'SKU': 'sku', 'item_name': 'title', 'brand': 'brand', 'description': 'description',
    'bullet_point_1': 'bullet_point_1', 'bullet_point_2': 'bullet_point_2',
    'bullet_point_3': 'bullet_point_3', 'bullet_point_4': 'bullet_point_4',
    'bullet_point_5': 'bullet_point_5', 'price': 'price', 'currency': 'currency',
    'main_image': 'image_url', 'image_2': 'image_2', 'image_3': 'image_3',
    'image_4': 'image_4', 'image_5': 'image_5', 'quantity': 'quantity',
    'UPC': 'upc', 'EAN': 'ean', 'GTIN': 'gtin',
    'product_type': 'product_type', 'condition_type': 'condition_type',
    'fulfillment_channel': 'fulfillment_channel', 'parent_sku': 'parent_sku',
    'parentage_level': 'parentage_level', 'variation_theme': 'variation_theme',
    'color': 'color', 'size': 'size', 'search_terms': 'keywords',
    'item_weight': 'weight', 'item_weight_unit': 'item_weight_unit',
    'item_length': 'item_length', 'item_width': 'item_width', 'item_height': 'item_height',
    'dimension_unit': 'dimension_unit',
    'package_weight': 'package_weight', 'package_weight_unit': 'package_weight_unit',
    'package_length': 'package_length', 'package_width': 'package_width',
    'package_height': 'package_height',
    'target_gender': 'target_gender', 'age_range': 'age_range',
    'department': 'department', 'material': 'material',
    'manufacturer': 'manufacturer', 'model_number': 'model_number',
    # Offer extras
    'sale_price': 'sale_price', 'list_price': 'list_price',
    'sale_from_date': 'sale_from_date', 'sale_end_date': 'sale_end_date',
    'max_order_quantity': 'max_order_quantity', 'handling_time': 'handling_time',
    'merchant_shipping_group_name': 'merchant_shipping_group_name',
    'product_tax_code': 'product_tax_code',
    # Identity extras
    'item_type_keyword': 'item_type_keyword',
    'external_product_id_type': 'external_product_id_type',
    'number_of_items': 'number_of_items',
    # Compliance extras
    'cpsia_cautionary_statement': 'cpsia_cautionary_statement',
    'legal_disclaimer': 'legal_disclaimer', 'safety_warning': 'safety_warning',
    'number_of_batteries': 'number_of_batteries',
    'battery_cell_composition': 'battery_cell_composition',
    'lithium_battery_packaging': 'lithium_battery_packaging',
    'lithium_battery_energy_content': 'lithium_battery_energy_content',
    'lithium_battery_weight': 'lithium_battery_weight',
    # Attribute extras
    'style_name': 'style_name', 'pattern_type': 'pattern_type',
    'closure_type': 'closure_type', 'unit_count': 'unit_count',
    'unit_count_type': 'unit_count_type',
}


class ExcelProcessor:
    """Excel商品数据处理器 V2 — 支持前后对比"""

    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.headers = []
        self.data = []
        self.raw_headers = []  # 原始列名

    def read_input(self, filepath: str, header_row: int = None) -> List[Dict]:
        """
        读取Excel文件(兼容任意格式)

        自动检测表头行，支持中英文列名
        """
        logger.info(f"📖 读取Excel: {filepath}")
        ext = os.path.splitext(str(filepath))[1].lower()
        if ext == '.xls':
            return self._read_xls_input(filepath, header_row=header_row)

        self.workbook = load_workbook(filepath, data_only=True)
        self.sheet = self.workbook.active

        # 自动检测表头行
        if header_row is None:
            for row_idx in range(1, min(10, self.sheet.max_row + 1)):
                row_values = [str(cell.value or '').strip() for cell in self.sheet[row_idx]]
                field_indicators = ['SKU', 'sku', 'item_name', 'product_id', 'brand_name',
                                    'standard_price', '商品编号', '标题', '品牌', 'ASIN',
                                    'title', 'brand', 'price']
                if any(indicator in row_values for indicator in field_indicators):
                    header_row = row_idx
                    break
            if header_row is None:
                header_row = 1

        logger.info(f"  表头行: 第{header_row}行")

        # 读取表头
        self.raw_headers = []
        self.headers = []
        for cell in self.sheet[header_row]:
            raw = str(cell.value).strip() if cell.value else f"col_{cell.column}"
            self.raw_headers.append(raw)
            self.headers.append(raw)

        # 读取数据
        self.data = []
        for row_idx, row in enumerate(self.sheet.iter_rows(
                min_row=header_row + 1, values_only=True), start=header_row + 1):
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
                continue
            item = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(self.headers):
                    item[self.headers[col_idx]] = value
            item['_row_index'] = row_idx
            self.data.append(item)

        # Close workbook after data extraction (prevent file lock on Windows)
        self.workbook.close()

        logger.info(f"✅ 读取完成: {len(self.data)} 条商品, {len(self.headers)} 列")
        return self.data

    def _read_xls_input(self, filepath: str, header_row: int = None) -> List[Dict]:
        """使用 xlrd 读取旧版 .xls 文件。"""
        if xlrd is None:
            raise ImportError("读取 .xls 需要安装 xlrd")

        workbook = xlrd.open_workbook(filepath)
        sheet = workbook.sheet_by_index(0)
        self.workbook = None
        self.sheet = None

        if header_row is None:
            for row_idx in range(min(10, sheet.nrows)):
                row_values = [str(sheet.cell_value(row_idx, col_idx) or '').strip()
                              for col_idx in range(sheet.ncols)]
                field_indicators = ['SKU', 'sku', 'item_name', 'product_id', 'brand_name',
                                    'standard_price', '商品编号', '标题', '品牌', 'ASIN',
                                    'title', 'brand', 'price']
                if any(indicator in row_values for indicator in field_indicators):
                    header_row = row_idx + 1
                    break
            if header_row is None:
                header_row = 1

        logger.info(f"  表头行: 第{header_row}行")

        self.raw_headers = []
        self.headers = []
        header_idx = header_row - 1
        for col_idx in range(sheet.ncols):
            raw_value = sheet.cell_value(header_idx, col_idx) if header_idx < sheet.nrows else ''
            raw = str(raw_value).strip() if str(raw_value).strip() else f"col_{col_idx + 1}"
            self.raw_headers.append(raw)
            self.headers.append(raw)

        self.data = []
        for row_idx in range(header_idx + 1, sheet.nrows):
            row_values = [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
            if all(self._is_empty_cell(value) for value in row_values):
                continue

            item = {}
            for col_idx, value in enumerate(row_values):
                if col_idx >= len(self.headers):
                    continue
                item[self.headers[col_idx]] = self._normalize_xls_cell_value(value)
            item['_row_index'] = row_idx + 1
            self.data.append(item)

        logger.info(f"✅ 读取完成: {len(self.data)} 条商品, {len(self.headers)} 列")
        return self.data

    def _is_empty_cell(self, value) -> bool:
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        return False

    def _normalize_xls_cell_value(self, value):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    def detect_columns(self) -> Dict[str, Optional[str]]:
        """
        自动检测Excel中的关键列(模糊匹配)
        返回 {标准字段名: 实际列名}
        """
        mapping = {}

        # First pass: exact matching via COLUMN_ALIASES
        for header in self.headers:
            h = header.strip()
            if h in COLUMN_ALIASES:
                alias_target = COLUMN_ALIASES[h]
                std = _ALIAS_TARGET_TO_DETECT.get(alias_target)
                if std and std not in mapping:
                    mapping[std] = header

        # Second pass: fuzzy matching via keywords_map (fills gaps not covered above)
        keywords_map = {
            'sku': ['sku', 'SKU', '商品编号', '产品编号', '编号', 'product_id', 'seller_sku', 'item_sku'],
            'asin': ['asin', 'ASIN'],
            'title': ['title', '标题', '商品标题', '产品标题', 'product_title', '产品名称', 'name',
                       'item_name', '商品名称', 'listing_title'],
            'description': ['description', '描述', '商品描述', '产品描述', '详情', 'product_description'],
            'bullet_point_1': ['bullet_point_1', '卖点1', '要点1', 'bullet1', 'key_feature_1'],
            'bullet_point_2': ['bullet_point_2', '卖点2', '要点2', 'bullet2', 'key_feature_2'],
            'bullet_point_3': ['bullet_point_3', '卖点3', '要点3', 'bullet3', 'key_feature_3'],
            'bullet_point_4': ['bullet_point_4', '卖点4', '要点4', 'bullet4', 'key_feature_4'],
            'bullet_point_5': ['bullet_point_5', '卖点5', '要点5', 'bullet5', 'key_feature_5'],
            'bullet_points': ['bullet', '要点', '卖点', 'feature', 'bullet_point', '五点描述'],
            'price': ['price', '价格', '售价', 'list_price', 'standard_price', 'our_price'],
            'brand': ['brand', '品牌', '品牌名', 'brand_name', '品牌名称'],
            'product_type': ['product_type', '产品类型', '类目', 'category', '品类', 'item_type_keyword'],
            'image_url': ['image_url', '图片链接', '图片URL', 'main_image', '主图',
                          'main_image_url', 'image', 'main_product_image', '主图URL'],
            'image_2': ['other_image_url_1', '副图1', 'image_2'],
            'image_3': ['other_image_url_2', '副图2', 'image_3'],
            'image_4': ['other_image_url_3', '副图3', 'image_4'],
            'image_5': ['other_image_url_4', '副图4', 'image_5'],
            'keywords': ['keywords', '关键词', 'search_terms', '搜索词', 'generic_keywords', 'generic_keyword'],
            'quantity': ['quantity', '库存', '数量', 'inventory', 'stock'],
            'fulfillment_channel': ['fulfillment_channel', '配送方式', 'fulfillment', 'FBA_FBM'],
            'condition_type': ['condition_type', '商品状态', 'condition', '成色'],
            'color': ['color', 'colour', '颜色', 'color_name', 'colour_name'],
            'size': ['size', '尺寸', '尺码', 'size_name'],
            'material': ['material', '材质', 'material_type'],
            'weight': ['item_weight', '重量', 'weight'],
            'manufacturer': ['manufacturer', '制造商'],
            'model_number': ['model_number', '型号', 'part_number'],
            'country_of_origin': ['country_of_origin', '原产国'],
            'upc': ['upc', 'UPC', 'product_id', 'external_product_id', '条形码'],
            'ean': ['ean', 'EAN', 'ean13', 'ean_code'],
            'gtin': ['gtin', 'GTIN', 'gtin14', 'gtin_code'],
            'parent_sku': ['parent_sku', '父SKU', 'parent_child', '父商品SKU'],
            'variation_theme': ['variation_theme', '变体主题', 'theme'],
            'parentage_level': ['parentage_level', '关系类型', 'relationship_type', 'parent_child_type'],
            'target_gender': ['target_gender', '目标性别', 'gender'],
            'age_range': ['age_range', '年龄范围', 'age_range_description', '适用年龄'],
            'department': ['department', '部门', 'department_name'],
            'item_weight_unit': ['item_weight_unit', '重量单位', 'weight_unit', 'weight_unit_of_measure'],
            'dimension_unit': ['dimension_unit', '尺寸单位', 'length_unit', 'unit_of_dimension'],
            'package_weight': ['package_weight', '包装重量', 'shipping_weight'],
            'package_weight_unit': ['package_weight_unit', '包装重量单位'],
            'package_length': ['package_length', '包装长度'],
            'package_width': ['package_width', '包装宽度'],
            'package_height': ['package_height', '包装高度'],
            'currency': ['currency', '币种', 'currency_code'],
            # Offer extras
            'sale_price': ['sale_price', '促销价', 'discount_price'],
            'sale_from_date': ['sale_from_date', '促销开始', 'sale_start_date'],
            'sale_end_date': ['sale_end_date', '促销结束'],
            'list_price': ['list_price', '建议零售价', 'msrp', 'suggested_price'],
            'max_order_quantity': ['max_order_quantity', '最大订购量'],
            'handling_time': ['handling_time', '发货时间', 'leadtime_to_ship', 'lead_time_to_ship_max_days'],
            'merchant_shipping_group_name': ['merchant_shipping_group_name', '配送模板', 'shipping_template'],
            'product_tax_code': ['product_tax_code', '税码', 'tax_code'],
            # Identity extras
            'item_type_keyword': ['item_type_keyword', '商品类型关键词', 'item_type'],
            'external_product_id_type': ['external_product_id_type', '产品ID类型', 'product_id_type'],
            'product_identity_mode': ['product_identity_mode', '商品标识模式', 'identifier_mode'],
            'number_of_items': ['number_of_items', '包装件数', 'item_count'],
            # Compliance extras
            'cpsia_cautionary_statement': ['cpsia_cautionary_statement', 'cpsia', 'CPSIA'],
            'legal_disclaimer': ['legal_disclaimer', '法律声明', '免责声明'],
            'safety_warning': ['safety_warning', '安全警告'],
            'number_of_batteries': ['number_of_batteries', '电池数量', 'num_batteries'],
            'battery_cell_composition': ['battery_cell_composition', '电池成分', 'battery_composition'],
            'lithium_battery_packaging': ['lithium_battery_packaging', '锂电池包装'],
            'lithium_battery_energy_content': ['lithium_battery_energy_content', '锂电池能量'],
            'lithium_battery_weight': ['lithium_battery_weight', '锂电池重量'],
            # Attribute extras
            'style_name': ['style_name', '风格', 'style'],
            'pattern_type': ['pattern_type', '图案', 'pattern_name', 'pattern'],
            'closure_type': ['closure_type', '闭合类型', 'closure'],
            'unit_count': ['unit_count', '单位数量'],
            'unit_count_type': ['unit_count_type', '单位类型'],
        }

        for standard_name, search_terms in keywords_map.items():
            if standard_name in mapping:
                continue  # Skip fields already resolved by exact match in pass 1
            for header in self.headers:
                header_lower = header.lower().strip()
                for term in search_terms:
                    if term.lower() == header_lower:
                        mapping[standard_name] = header
                        break
                if standard_name in mapping:
                    break

        logger.info(f"📊 列检测结果: {len(mapping)} 个匹配")
        return mapping

    def detect_variants(self) -> List[Dict]:
        """
        检测变体关系，返回分组结构

        逻辑:
        1. 如果有 parentage_level 列，按 parent/child 分组
        2. 如果有 parent_sku 列，按 parent_sku 分组
        3. 如果都没有，检查是否有重复的标题/品牌来推断

        Returns:
            [{"type": "standalone"|"parent", "parent_sku": str, "title": str, "skus": [row_dicts]}]
        """
        if not self.data:
            return []

        col_map = self.detect_columns()
        groups = []

        # 获取实际列名
        parentage_col = col_map.get('parentage_level')
        parent_sku_col = col_map.get('parent_sku')
        sku_col = col_map.get('sku')
        title_col = col_map.get('title')
        brand_col = col_map.get('brand')

        # === 方式1: 按 parentage_level 列分组 ===
        if parentage_col:
            logger.info("📦 使用 parentage_level 列检测变体关系")
            parent_map = {}  # parent_sku -> {"parent_row": ..., "children": [...]}

            for row in self.data:
                level = str(row.get(parentage_col, '')).strip().lower()
                p_sku = str(row.get(parent_sku_col, '')).strip() if parent_sku_col else ''
                sku = str(row.get(sku_col, '')).strip() if sku_col else ''

                if level == 'parent':
                    if sku not in parent_map:
                        parent_map[sku] = {'parent_row': row, 'children': []}
                    else:
                        parent_map[sku]['parent_row'] = row
                elif level == 'child' and p_sku:
                    if p_sku not in parent_map:
                        parent_map[p_sku] = {'parent_row': None, 'children': []}
                    parent_map[p_sku]['children'].append(row)
                else:
                    # 独立商品
                    groups.append({
                        'type': 'standalone',
                        'parent_sku': sku,
                        'title': str(row.get(title_col, '')).strip() if title_col else '',
                        'skus': [row],
                    })

            for p_sku, info in parent_map.items():
                all_rows = []
                if info['parent_row']:
                    all_rows.append(info['parent_row'])
                all_rows.extend(info['children'])
                title = ''
                if info['parent_row'] and title_col:
                    title = str(info['parent_row'].get(title_col, '')).strip()
                elif info['children'] and title_col:
                    title = str(info['children'][0].get(title_col, '')).strip()
                groups.append({
                    'type': 'parent',
                    'parent_sku': p_sku,
                    'title': title,
                    'skus': all_rows,
                })

            logger.info(f"📊 变体检测: {len(groups)} 个分组")
            return groups

        # === 方式2: 按 parent_sku 列分组 ===
        if parent_sku_col:
            logger.info("📦 使用 parent_sku 列检测变体关系")
            parent_map = {}

            for row in self.data:
                p_sku = str(row.get(parent_sku_col, '')).strip()
                if p_sku:
                    if p_sku not in parent_map:
                        parent_map[p_sku] = []
                    parent_map[p_sku].append(row)
                else:
                    sku = str(row.get(sku_col, '')).strip() if sku_col else ''
                    groups.append({
                        'type': 'standalone',
                        'parent_sku': sku,
                        'title': str(row.get(title_col, '')).strip() if title_col else '',
                        'skus': [row],
                    })

            for p_sku, rows in parent_map.items():
                title = str(rows[0].get(title_col, '')).strip() if title_col else ''
                groups.append({
                    'type': 'parent',
                    'parent_sku': p_sku,
                    'title': title,
                    'skus': rows,
                })

            logger.info(f"📊 变体检测: {len(groups)} 个分组")
            return groups

        # === 方式3: 推断 — 按标题+品牌组合分组 ===
        logger.info("📦 无明确变体列，尝试按标题+品牌推断变体关系")
        title_brand_map = {}

        for row in self.data:
            title = str(row.get(title_col, '')).strip() if title_col else ''
            brand = str(row.get(brand_col, '')).strip() if brand_col else ''
            key = f"{title}||{brand}" if title else None

            if key and key in title_brand_map:
                title_brand_map[key].append(row)
            elif key:
                title_brand_map[key] = [row]
            else:
                sku = str(row.get(sku_col, '')).strip() if sku_col else ''
                groups.append({
                    'type': 'standalone',
                    'parent_sku': sku,
                    'title': title,
                    'skus': [row],
                })

        for key, rows in title_brand_map.items():
            title = key.split('||')[0]
            first_sku = str(rows[0].get(sku_col, '')).strip() if sku_col else ''
            if len(rows) > 1:
                groups.append({
                    'type': 'parent',
                    'parent_sku': first_sku,
                    'title': title,
                    'skus': rows,
                })
            else:
                groups.append({
                    'type': 'standalone',
                    'parent_sku': first_sku,
                    'title': title,
                    'skus': rows,
                })

        logger.info(f"📊 变体检测(推断): {len(groups)} 个分组")
        return groups

    def validate_row(self, row: dict, product_type: str = None) -> dict:
        """
        校验单行数据

        Args:
            row: 已通过detect_columns映射后的行数据(或原始行数据)
            product_type: 产品类型(用于特定规则校验)

        Returns:
            {'sku': str, 'valid': bool, 'issues': [str]}
        """
        issues = []

        # --- 必填检查 ---
        sku = row.get('sku') or row.get('SKU') or row.get('seller_sku') or ''
        if not sku:
            issues.append('缺少SKU')

        title = row.get('title') or row.get('item_name') or row.get('AI标题') or ''
        if not title:
            issues.append('缺少标题')

        # --- 长度检查 ---
        title = str(title) if title else ''
        if len(title) > 200:
            issues.append(f'标题超长({len(title)}/200)')

        # Bullet Points长度
        for i in range(1, 6):
            bp = row.get(f'bullet_point_{i}') or row.get(f'AI卖点{i}') or ''
            bp = str(bp) if bp else ''
            if len(bp) > 500:
                issues.append(f'卖点{i}超长({len(bp)}/500)')

        # 描述长度
        desc = row.get('description') or row.get('AI商品描述') or ''
        desc = str(desc) if desc else ''
        if len(desc) > 2000:
            issues.append(f'描述超长({len(desc)}/2000)')

        # 搜索词字节数
        keywords = row.get('keywords') or row.get('AI搜索关键词') or ''
        keywords = str(keywords) if keywords else ''
        if keywords:
            kw_bytes = len(keywords.encode('utf-8'))
            if kw_bytes > 250:
                issues.append(f'搜索词超过250字节({kw_bytes}字节)')

        # --- 格式检查 ---
        price = row.get('price')
        if price is None:
            price = row.get('standard_price')
        if price is None:
            price = ''
        if price == '' or price is None:
            issues.append('缺少价格')
        else:
            try:
                price_val = float(str(price).replace(',', '').replace('$', '').strip())
                if price_val <= 0:
                    issues.append(f'价格无效({price})')
                elif price_val > 10000:
                    issues.append(f'价格疑似填错单位或小数点({price})')
            except (ValueError, TypeError):
                issues.append(f'价格格式错误({price})')

        quantity = row.get('quantity')
        if quantity is None:
            quantity = ''
        if quantity != '' and quantity is not None:
            try:
                qty_val = int(quantity)
                if qty_val < 0:
                    issues.append(f'库存不能为负数({quantity})')
            except (ValueError, TypeError):
                issues.append(f'库存格式错误({quantity})')

        # 图片URL格式
        img_url = row.get('main_image_url') or row.get('image_url') or row.get('main_image') or ''
        img_url = str(img_url).strip() if img_url else ''
        if not img_url:
            issues.append('缺少主图URL')
        elif not img_url.startswith(('http://', 'https://', 's3://')):
            issues.append('主图URL格式无效(需以http(s)或s3://开头)')

        # --- 推荐字段缺失(警告级别，标记但不影响valid) ---
        # 这里只检查硬性校验，推荐字段由validate_required_fields处理

        return {
            'sku': str(sku),
            'valid': len(issues) == 0,
            'issues': issues,
        }

    def write_comparison_output(self, data: List[Dict], output_path: str,
                                col_map: Dict, ai_results: Dict = None):
        """
        写出前后对比Excel

        每个商品显示:
        - 原始值(中文)
        - AI生成值(英文)
        - 其他不变的字段
        - 提交状态
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "前后对比"

        # 定义输出列结构
        output_columns = self._build_output_columns(col_map)

        # 样式定义
        thin_border = Border(
            left=Side(style='thin', color='FFCCCCCC'),
            right=Side(style='thin', color='FFCCCCCC'),
            top=Side(style='thin', color='FFCCCCCC'),
            bottom=Side(style='thin', color='FFCCCCCC'),
        )

        # 写分组标题行 (第1行)
        current_col = 1
        for group_key, group_info in self._get_active_groups(output_columns):
            cols_in_group = [c for c in output_columns if c.get('group') == group_key]
            if not cols_in_group:
                continue
            cell = ws.cell(row=1, column=current_col,
                          value=group_info['title'])
            cell.fill = PatternFill(start_color=group_info['color'],
                                     end_color=group_info['color'], fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.alignment = Alignment(horizontal='center')
            if len(cols_in_group) > 1:
                ws.merge_cells(start_row=1, start_column=current_col,
                              end_row=1, end_column=current_col + len(cols_in_group) - 1)
            current_col += len(cols_in_group)

        # 写列名行 (第2行)
        for col_idx, col_def in enumerate(output_columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_def['label'])
            # 原始列浅色背景，AI列深色背景
            if col_def.get('is_ai'):
                cell.fill = PatternFill(start_color='FFFFF3CD', end_color='FFFFF3CD',
                                       fill_type="solid")  # 淡黄色=AI生成
                cell.font = Font(bold=True, size=10, color='FF856404')
            else:
                cell.fill = PatternFill(start_color='FFE2E3E5', end_color='FFE2E3E5',
                                       fill_type="solid")  # 浅灰色=原始
                cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        # 写数据 (第3行开始)
        for row_idx, item in enumerate(data, start=3):
            for col_idx, col_def in enumerate(output_columns, start=1):
                value = self._get_cell_value(item, col_def, col_map)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.border = thin_border

                # AI生成的值加淡黄背景
                if col_def.get('is_ai') and value:
                    cell.fill = PatternFill(start_color='FFFFFDE7', end_color='FFFFFDE7',
                                           fill_type="solid")

                # 状态列着色
                if col_def.get('key') == 'submit_status':
                    if value == 'ACCEPTED':
                        cell.font = Font(color='FF28A745', bold=True)
                    elif value == 'INVALID':
                        cell.font = Font(color='FFDC3545', bold=True)
                    elif value == 'PENDING':
                        cell.font = Font(color='FFFD7E14')

        # 设置列宽
        for col_idx, col_def in enumerate(output_columns, start=1):
            width = col_def.get('width', 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 冻结前2行
        ws.freeze_panes = 'A3'

        # 添加字段说明Sheet
        self._add_field_doc_sheet(wb)

        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        wb.save(output_path)
        logger.info(f"✅ 对比Excel输出: {output_path} ({len(data)} 条)")

    def _build_output_columns(self, col_map: Dict) -> List[Dict]:
        """构建输出列定义"""
        columns = []

        # --- 产品身份 ---
        columns.append({'key': 'sku', 'label': 'SKU', 'group': 'identity',
                        'source': 'sku', 'width': 15})
        columns.append({'key': 'brand', 'label': '品牌 Brand', 'group': 'identity',
                        'source': 'brand', 'width': 15})
        columns.append({'key': 'upc', 'label': 'UPC/EAN', 'group': 'identity',
                        'source': 'upc', 'width': 15})
        columns.append({'key': 'product_type', 'label': '产品类型', 'group': 'identity',
                        'source': 'product_type', 'width': 15})

        # --- AI内容(前后对比) ---
        columns.append({'key': 'title_original', 'label': '原始标题(中文)', 'group': 'content',
                        'source': 'title', 'width': 35})
        columns.append({'key': 'title_ai', 'label': '→ AI标题(英文)', 'group': 'content',
                        'source': 'AI标题', 'is_ai': True, 'width': 40})

        for i in range(1, 6):
            columns.append({
                'key': f'bullet_{i}_orig', 'label': f'原始卖点{i}',
                'group': 'content', 'source': f'bullet_point_{i}', 'width': 25
            })
            columns.append({
                'key': f'bullet_{i}_ai', 'label': f'→ AI卖点{i}',
                'group': 'content', 'source': f'AI卖点{i}', 'is_ai': True, 'width': 30
            })

        columns.append({'key': 'desc_original', 'label': '原始描述', 'group': 'content',
                        'source': 'description', 'width': 35})
        columns.append({'key': 'desc_ai', 'label': '→ AI描述(英文)', 'group': 'content',
                        'source': 'AI商品描述', 'is_ai': True, 'width': 40})

        columns.append({'key': 'search_terms_ai', 'label': 'AI搜索关键词', 'group': 'content',
                        'source': 'AI搜索关键词', 'is_ai': True, 'width': 30})
        columns.append({'key': 'ai_text_model', 'label': 'AI文案模型', 'group': 'content',
                        'source': 'AI文案模型', 'is_ai': True, 'width': 18})
        columns.append({'key': 'ai_text_protocol', 'label': 'AI文案协议', 'group': 'content',
                        'source': 'AI文案协议', 'is_ai': True, 'width': 18})
        columns.append({'key': 'ai_text_attempts', 'label': 'AI文案尝试次数', 'group': 'content',
                        'source': 'AI文案尝试次数', 'is_ai': True, 'width': 12})
        columns.append({'key': 'ai_text_generated_at', 'label': 'AI文案生成时间', 'group': 'content',
                        'source': 'AI文案生成时间', 'is_ai': True, 'width': 20})
        columns.append({'key': 'ai_text_error', 'label': 'AI文案最后错误', 'group': 'content',
                        'source': 'AI文案最后错误', 'is_ai': True, 'width': 28})

        # --- 图片 ---
        columns.append({'key': 'main_img_orig', 'label': '原始主图', 'group': 'images',
                        'source': 'image_url', 'width': 30})
        columns.append({'key': 'main_img_ai', 'label': '→ AI主图(白底)', 'group': 'images',
                        'source': 'AI主图路径', 'is_ai': True, 'width': 30})
        columns.append({'key': 'main_img_ai_locator', 'label': '→ AI主图提交地址', 'group': 'images',
                        'source': 'AI主图URL', 'is_ai': True, 'width': 36})
        columns.append({'key': 'main_img_ai_upload_status', 'label': '→ 主图上传状态', 'group': 'images',
                        'source': 'AI主图上传状态', 'is_ai': True, 'width': 14})
        columns.append({'key': 'ai_image_model', 'label': 'AI图片模型', 'group': 'images',
                        'source': 'AI图片模型', 'is_ai': True, 'width': 18})
        columns.append({'key': 'ai_image_protocol', 'label': 'AI图片协议', 'group': 'images',
                        'source': 'AI图片协议', 'is_ai': True, 'width': 18})
        columns.append({'key': 'ai_image_attempts', 'label': 'AI图片尝试次数', 'group': 'images',
                        'source': 'AI图片尝试次数', 'is_ai': True, 'width': 12})
        columns.append({'key': 'ai_image_generated_at', 'label': 'AI图片生成时间', 'group': 'images',
                        'source': 'AI图片生成时间', 'is_ai': True, 'width': 20})
        columns.append({'key': 'ai_image_error', 'label': 'AI图片最后错误', 'group': 'images',
                        'source': 'AI图片最后错误', 'is_ai': True, 'width': 28})
        for i in range(2, 10):
            columns.append({
                'key': f'img_{i}', 'label': f'副图{i-1}',
                'group': 'images', 'source': f'image_{i}', 'width': 25
            })
            columns.append({
                'key': f'ai_img_{i}', 'label': f'→ AI副图{i-1}',
                'group': 'images', 'source': f'AI副图{i}路径', 'is_ai': True, 'width': 30
            })
            columns.append({
                'key': f'ai_img_{i}_locator', 'label': f'→ AI副图{i-1}提交地址',
                'group': 'images', 'source': f'AI副图{i}URL', 'is_ai': True, 'width': 36
            })

        # --- 销售条款 ---
        columns.append({'key': 'price', 'label': '价格', 'group': 'offer',
                        'source': 'price', 'width': 12})
        columns.append({'key': 'currency', 'label': '币种', 'group': 'offer',
                        'source': 'currency', 'width': 8})
        columns.append({'key': 'list_price', 'label': '建议零售价', 'group': 'offer',
                        'source': 'list_price', 'width': 12})
        columns.append({'key': 'sale_price', 'label': '促销价', 'group': 'offer',
                        'source': 'sale_price', 'width': 12})
        columns.append({'key': 'sale_from_date', 'label': '促销开始', 'group': 'offer',
                        'source': 'sale_from_date', 'width': 14})
        columns.append({'key': 'sale_end_date', 'label': '促销结束', 'group': 'offer',
                        'source': 'sale_end_date', 'width': 14})
        columns.append({'key': 'quantity', 'label': '库存', 'group': 'offer',
                        'source': 'quantity', 'width': 10})
        columns.append({'key': 'max_order_quantity', 'label': '最大订购量', 'group': 'offer',
                        'source': 'max_order_quantity', 'width': 10})
        columns.append({'key': 'condition', 'label': '状���', 'group': 'offer',
                        'source': 'condition_type', 'width': 10})
        columns.append({'key': 'fulfillment', 'label': '配送方式', 'group': 'offer',
                        'source': 'fulfillment_channel', 'width': 12})
        columns.append({'key': 'handling_time', 'label': '发货时间(天)', 'group': 'offer',
                        'source': 'handling_time', 'width': 10})
        columns.append({'key': 'merchant_shipping_group_name', 'label': '配送模板', 'group': 'offer',
                        'source': 'merchant_shipping_group_name', 'width': 15})
        columns.append({'key': 'product_tax_code', 'label': '税码', 'group': 'offer',
                        'source': 'product_tax_code', 'width': 10})
        columns.append({'key': 'external_product_id_type', 'label': 'ID类型', 'group': 'offer',
                        'source': 'external_product_id_type', 'width': 10})

        # --- 物流 ---
        columns.append({'key': 'weight', 'label': '重量(g)', 'group': 'shipping',
                        'source': 'weight', 'width': 10})
        columns.append({'key': 'length', 'label': '长(cm)', 'group': 'shipping',
                        'source': 'item_length', 'width': 8})
        columns.append({'key': 'width_dim', 'label': '宽(cm)', 'group': 'shipping',
                        'source': 'item_width', 'width': 8})
        columns.append({'key': 'height', 'label': '高(cm)', 'group': 'shipping',
                        'source': 'item_height', 'width': 8})

        # --- 属性 ---
        columns.append({'key': 'color', 'label': '颜色', 'group': 'attributes',
                        'source': 'color', 'width': 12})
        columns.append({'key': 'size', 'label': '尺寸', 'group': 'attributes',
                        'source': 'size', 'width': 12})
        columns.append({'key': 'material', 'label': '材质', 'group': 'attributes',
                        'source': 'material', 'width': 12})
        columns.append({'key': 'style_name', 'label': '风格', 'group': 'attributes',
                        'source': 'style_name', 'width': 12})
        columns.append({'key': 'pattern_type', 'label': '图案', 'group': 'attributes',
                        'source': 'pattern_type', 'width': 12})
        columns.append({'key': 'target_gender', 'label': '目标性别', 'group': 'attributes',
                        'source': 'target_gender', 'width': 10})
        columns.append({'key': 'age_range', 'label': '年龄范围', 'group': 'attributes',
                        'source': 'age_range', 'width': 12})
        columns.append({'key': 'country', 'label': '原��国', 'group': 'attributes',
                        'source': 'country_of_origin', 'width': 10})
        columns.append({'key': 'number_of_items', 'label': '包装件数', 'group': 'attributes',
                        'source': 'number_of_items', 'width': 10})

        # --- 合规 ---
        columns.append({'key': 'batteries_required', 'label': '需要电池', 'group': 'compliance',
                        'source': 'batteries_required', 'width': 10})
        columns.append({'key': 'batteries_included', 'label': '含电池', 'group': 'compliance',
                        'source': 'batteries_included', 'width': 10})
        columns.append({'key': 'battery_type', 'label': '电池类型', 'group': 'compliance',
                        'source': 'battery_type', 'width': 12})
        columns.append({'key': 'number_of_batteries', 'label': '电池数量', 'group': 'compliance',
                        'source': 'number_of_batteries', 'width': 10})
        columns.append({'key': 'lithium_battery_packaging', 'label': '锂电池包装', 'group': 'compliance',
                        'source': 'lithium_battery_packaging', 'width': 15})
        columns.append({'key': 'hazmat_declaration', 'label': '危化品声明', 'group': 'compliance',
                        'source': 'hazmat_declaration', 'width': 15})
        columns.append({'key': 'cpsia_cautionary_statement', 'label': 'CPSIA', 'group': 'compliance',
                        'source': 'cpsia_cautionary_statement', 'width': 15})
        columns.append({'key': 'safety_warning', 'label': '安全警告', 'group': 'compliance',
                        'source': 'safety_warning', 'width': 15})
        columns.append({'key': 'legal_disclaimer', 'label': '法律声明', 'group': 'compliance',
                        'source': 'legal_disclaimer', 'width': 15})

        # --- 品类字段(动态) ---
        cat_fields = self._load_selected_category_fields()
        for f in cat_fields:
            columns.append({
                'key': f['key'], 'label': f.get('label_zh', f['key']),
                'group': 'category_fields', 'source': f['key'], 'width': 15
            })

        # --- 上架状态 ---
        columns.append({'key': 'submit_status', 'label': '提交状态', 'group': 'status',
                        'source': 'submit_status', 'width': 12})
        columns.append({'key': 'submit_time', 'label': '提交时间', 'group': 'status',
                        'source': 'submit_time', 'width': 18})
        columns.append({'key': 'asin', 'label': 'ASIN', 'group': 'status',
                        'source': 'asin', 'width': 15})
        columns.append({'key': 'issues', 'label': '问题详情', 'group': 'status',
                        'source': 'issues', 'width': 40})

        return columns

    def _load_selected_category_fields(self) -> List[Dict]:
        """加载用户选择的品类字段配置"""
        config_path = runtime_path('config', 'selected_fields.json')
        if not os.path.exists(config_path):
            return []
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            selected_keys = config.get('fields', [])
            if not selected_keys:
                return []
            # 从注册表获取字段定义
            registry_path = resource_path('config', 'sp_api_fields.json')
            if not os.path.exists(registry_path):
                return [{'key': k} for k in selected_keys]
            with open(registry_path, 'r', encoding='utf-8') as f:
                registry = json.load(f)
            # 构建 key->定义 的映射
            all_fields = {}
            for cat in registry.get('categories', {}).values():
                for field in cat.get('fields', []):
                    all_fields[field['key']] = field
            for field in registry.get('cross_category_fields', []):
                all_fields[field['key']] = field
            return [all_fields.get(k, {'key': k, 'label_zh': k, 'label_en': k, 'type': 'text'})
                    for k in selected_keys]
        except Exception as e:
            logger.warning(f"加载品类字段配置失败: {e}")
            return []

    def _get_active_groups(self, columns):
        """获取有列的分组"""
        seen = []
        for col in columns:
            group_key = col.get('group')
            if group_key and group_key not in [s[0] for s in seen]:
                group_info = FIELD_GROUPS.get(group_key, {
                    'title': group_key, 'color': 'FF4472C4'
                })
                seen.append((group_key, group_info))
        return seen

    def _get_cell_value(self, item: Dict, col_def: Dict, col_map: Dict):
        """获取单元格值"""
        source = col_def.get('source', '')

        # 直接从item取(AI生成的字段)
        if source in item and item[source] is not None:
            return item[source]

        # 通过col_map映射取
        if source in col_map and col_map[source] in item:
            return item[col_map[source]]

        # 尝试直接匹配
        for header in self.headers:
            if header.lower() == source.lower():
                return item.get(header)

        return None

    def _add_field_doc_sheet(self, wb: Workbook):
        """添加字段说明Sheet"""
        ws = wb.create_sheet(title="字段说明")

        headers = ['分组', '字段名', '说明', '是否必填', 'AI可处理']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4',
                                   fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        docs = [
            ('产品身份', 'SKU', '卖家自定义商品编号', '✅必填', '❌'),
            ('产品身份', 'Brand', '品牌名称', '✅必填', '❌'),
            ('产品身份', 'UPC/EAN', '商品条形码', '⚠️强烈推荐', '❌'),
            ('产品身份', 'Product Type', '亚马逊产品类型', '✅必填', '❌'),
            ('AI内容', 'Title', '商品标题(≤200字符)', '✅必填', '✅翻译+SEO'),
            ('AI内容', 'Bullet Points', '5条卖点(每条≤500字符)', '⚠️推荐', '✅生成'),
            ('AI内容', 'Description', '商品描述(≤2000字符)', '⚠️推荐', '✅生成'),
            ('AI内容', 'Search Terms', '搜索关键词(≤250字节)', '⚠️推荐', '✅生成'),
            ('图片', 'Main Image', '主图(纯白底,≥1000×1000)', '✅必填', '✅换背景'),
            ('图片', 'Other Images', '副图(最多8张)', '⚠️推荐', '✅换背景'),
            ('销售', 'Price', '售价(USD)', '✅必填', '❌'),
            ('销售', 'Quantity', '库存数量', '✅必填', '❌'),
            ('销售', 'Condition', '商品状态(新/二手)', '✅必填', '❌'),
            ('销售', 'Fulfillment', '配送方式(FBA/FBM)', '✅必填', '❌'),
            ('物流', 'Dimensions', '商品尺寸(长宽高)', '⚠️推荐', '❌'),
            ('物流', 'Weight', '商品重量', '⚠️推荐', '❌'),
            ('属性', 'Color/Size/Material', '颜色/尺寸/材质', '按类型', '❌'),
            ('合规', 'Battery Info', '电池信息', '按类型', '❌'),
            ('变体', 'Variation', '父子SKU/变体主题', '有变体时', '❌'),
        ]

        for row_idx, doc in enumerate(docs, start=2):
            for col_idx, val in enumerate(doc, start=1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # 列宽
        for i, w in enumerate([12, 20, 30, 12, 12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ===== 兼容旧接口 =====
    def write_output(self, data: List[Dict], output_path: str,
                     extra_columns: Optional[List[str]] = None):
        """兼容旧版写出接口"""
        wb = Workbook()
        ws = wb.active
        ws.title = "商品数据"

        all_columns = list(self.headers)
        if extra_columns:
            for col in extra_columns:
                if col not in all_columns:
                    all_columns.append(col)

        header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4",
                                   fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for col_idx, header in enumerate(all_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, item in enumerate(data, start=2):
            for col_idx, header in enumerate(all_columns, start=1):
                value = item.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, ValueError, AttributeError):
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
        wb.save(output_path)
        logger.info(f"✅ Excel输出: {output_path} ({len(data)} 条)")
