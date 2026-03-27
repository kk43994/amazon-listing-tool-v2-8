"""
临时匿名图床工具。

当前仅支持 x0.at，用于把 Stage1 生成的本地 AI 图片批量上传后，
将返回的公网 URL 写回 Excel 的 AI主图URL 列，方便后续 Stage2 提交。
"""
from __future__ import annotations

import mimetypes
import os
from typing import Callable, Dict, Optional

import requests
from openpyxl import load_workbook

X0AT_UPLOAD_URL = "https://x0.at/"
HEADER_ROW_INDICATORS = (
    "SKU", "sku", "item_name", "product_id", "brand_name",
    "standard_price", "商品编号", "标题", "品牌", "ASIN",
    "title", "brand", "price",
)


def upload_file_to_x0at(file_path: str, timeout: int = 120) -> str:
    """上传单个文件到 x0.at，返回公网 URL。"""
    mime_type = mimetypes.guess_type(file_path)[0] or "application/octet-stream"
    filename = os.path.basename(file_path)

    with open(file_path, "rb") as fh:
        response = requests.post(
            X0AT_UPLOAD_URL,
            files={"file": (filename, fh, mime_type)},
            timeout=timeout,
        )

    response.raise_for_status()
    url = response.text.strip()
    if not url.startswith(("http://", "https://")):
        raise ValueError(f"x0.at 返回了无效URL: {url}")
    return url


def populate_ai_image_urls(
    excel_path: str,
    output_path: Optional[str] = None,
    *,
    only_missing: bool = True,
    uploader: Optional[Callable[[str], str]] = None,
) -> Dict[str, int]:
    """
    批量上传 Excel 中 AI主图路径 对应的本地图片，并写回 AI主图URL。

    Returns:
        {
            'uploaded': int,
            'copied': int,
            'skipped': int,
            'missing': int,
            'failed': int,
            'total': int,
        }
    """
    uploader = uploader or upload_file_to_x0at
    workbook = load_workbook(excel_path)
    worksheet = workbook.active
    header_row = _detect_header_row(worksheet)
    headers = _build_header_map(worksheet, header_row)

    path_col = _find_first_column(headers, ("AI主图路径", "→ AI主图(白底)"))
    if path_col is None:
        workbook.close()
        raise ValueError("找不到 AI主图路径 列")

    url_col = headers.get("AI主图URL")
    if url_col is None:
        url_col = worksheet.max_column + 1
        worksheet.cell(row=header_row, column=url_col, value="AI主图URL")

    stats = {
        "uploaded": 0,
        "copied": 0,
        "skipped": 0,
        "missing": 0,
        "failed": 0,
        "total": max(worksheet.max_row - header_row, 0),
    }

    try:
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            path_value = worksheet.cell(row=row_idx, column=path_col).value
            url_cell = worksheet.cell(row=row_idx, column=url_col)
            existing_url = str(url_cell.value or "").strip()

            if only_missing and existing_url:
                stats["skipped"] += 1
                continue

            resolved_path = _resolve_local_path(excel_path, path_value)
            if not resolved_path:
                stats["skipped"] += 1
                continue

            if resolved_path.startswith(("http://", "https://")):
                url_cell.value = resolved_path
                stats["copied"] += 1
                continue

            if not os.path.exists(resolved_path):
                stats["missing"] += 1
                continue

            try:
                url_cell.value = uploader(resolved_path)
                stats["uploaded"] += 1
            except Exception:
                stats["failed"] += 1

        target_path = output_path or excel_path
        workbook.save(target_path)
        return stats
    finally:
        workbook.close()


def _detect_header_row(worksheet) -> int:
    for row_idx in range(1, min(10, worksheet.max_row + 1)):
        row_values = [str(cell.value or "").strip() for cell in worksheet[row_idx]]
        if any(indicator in row_values for indicator in HEADER_ROW_INDICATORS):
            return row_idx
    return 1


def _build_header_map(worksheet, header_row: int) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col_idx in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=header_row, column=col_idx).value
        if value is None:
            continue
        headers[str(value).strip()] = col_idx
    return headers


def _find_first_column(headers: Dict[str, int], candidates) -> Optional[int]:
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
    return None


def _resolve_local_path(excel_path: str, cell_value) -> str:
    raw = str(cell_value or "").strip()
    if not raw:
        return ""

    if raw.startswith(("http://", "https://")):
        return raw

    if os.path.isabs(raw):
        return raw

    excel_dir = os.path.dirname(os.path.abspath(excel_path))
    candidate = os.path.abspath(os.path.join(excel_dir, raw))
    if os.path.exists(candidate):
        return candidate

    return os.path.abspath(raw)
