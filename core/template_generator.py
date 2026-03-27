"""
Amazon 官方字段模板生成器。

能力：
- 将 Product Type schema 解析结果转换成更适合 Excel / 采集插件使用的模板定义
- 生成“首行说明 + 第二行稳定字段名”的 Excel 模板
- 在隐藏工作表中写入模板元数据，便于后续上传识别与诊断
"""
from __future__ import annotations

import os
import re
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


GROUP_ORDER = [
    "meta",
    "identity",
    "variation",
    "offer",
    "content",
    "images",
    "shipping",
    "compliance",
    "other",
]

LEVEL_ORDER = {
    "required": 0,
    "recommended": 1,
    "optional": 2,
}

REQUIRED_FILL = PatternFill(start_color="FFF1B8", end_color="FFF1B8", fill_type="solid")
RECOMMENDED_FILL = PatternFill(start_color="E6F7FF", end_color="E6F7FF", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
META_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
HEADER_FONT = Font(color="262626", bold=True, size=10)
DESC_FONT = Font(color="595959", size=9)
META_FONT = Font(color="262626", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

LABEL_ZH = {
    "sku": "SKU",
    "product_type": "产品类型",
    "item_name": "商品标题",
    "brand": "品牌",
    "manufacturer": "制造商",
    "model_number": "型号",
    "model_name": "型号名称",
    "item_type_keyword": "商品类型关键词",
    "product_description": "商品描述",
    "bullet_point_1": "卖点1",
    "bullet_point_2": "卖点2",
    "bullet_point_3": "卖点3",
    "bullet_point_4": "卖点4",
    "bullet_point_5": "卖点5",
    "generic_keywords": "搜索关键词",
    "main_image_url": "主图URL",
    "other_image_url_1": "副图1 URL",
    "other_image_url_2": "副图2 URL",
    "other_image_url_3": "副图3 URL",
    "other_image_url_4": "副图4 URL",
    "standard_price": "售价",
    "currency": "币种",
    "quantity": "库存",
    "fulfillment_channel": "配送方式",
    "condition_type": "商品状况",
    "upc": "UPC/EAN/GTIN",
    "external_product_id_type": "产品ID类型",
    "asin": "ASIN",
    "product_identity_mode": "商品标识模式",
    "parent_sku": "父SKU",
    "parentage_level": "父子关系",
    "variation_theme": "变体主题",
    "color": "颜色",
    "size": "尺寸",
    "material": "材质",
    "department": "部门",
    "target_gender": "目标性别",
    "age_range_description": "年龄范围",
    "special_feature": "特色功能",
    "care_instructions": "护理说明",
    "included_components": "包装内容",
    "number_of_items": "包装件数",
    "unit_count": "单位数量",
    "item_weight": "商品重量",
    "item_width": "商品宽度",
    "item_length": "商品长度",
    "item_height": "商品高度",
    "dimension_unit": "尺寸单位",
    "package_weight": "包装重量",
    "package_weight_unit": "包装重量单位",
    "package_length": "包装长度",
    "package_width": "包装宽度",
    "package_height": "包装高度",
    "country_of_origin": "原产国",
    "batteries_required": "需要电池",
    "batteries_included": "含电池",
    "battery_type": "电池类型",
    "number_of_batteries": "电池数量",
    "battery_cell_composition": "电池成分",
    "lithium_battery_packaging": "锂电池包装",
    "lithium_battery_energy_content": "锂电池能量",
    "lithium_battery_weight": "锂电池重量",
    "hazmat_declaration": "危化品声明",
    "supplier_declared_dg_hz_regulation": "危化品监管声明",
    "cpsia_cautionary_statement": "CPSIA 提示",
    "safety_warning": "安全警告",
    "legal_disclaimer": "法律免责声明",
}

FIELD_EXAMPLES = {
    "sku": "MOUSE-BLK-001",
    "product_type": "COMPUTER_INPUT_DEVICE",
    "item_name": "Wireless Ergonomic Mouse for Laptop",
    "brand": "YourBrand",
    "manufacturer": "Your Factory",
    "main_image_url": "https://example.com/main.jpg",
    "standard_price": "19.99",
    "currency": "USD",
    "quantity": "100",
    "upc": "123456789012",
    "external_product_id_type": "UPC",
    "generic_keywords": "wireless mouse ergonomic usb silent",
    "parentage_level": "parent / child",
    "variation_theme": "COLOR_NAME / SIZE_NAME / COLOR_SIZE",
}

GROUP_MAP = {
    "product_identity": "identity",
    "product_description": "content",
    "offer": "offer",
    "image": "images",
    "variation": "variation",
    "dimensions": "shipping",
    "compliance": "compliance",
    "other": "other",
}

SIMPLE_ATTRIBUTE_EXCLUDE = {
    "item_name",
    "brand",
    "product_description",
    "bullet_point",
    "generic_keyword",
    "main_product_image_locator",
    "purchasable_offer",
    "fulfillment_availability",
    "condition_type",
    "externally_assigned_product_identifier",
    "merchant_suggested_asin",
    "item_dimensions",
    "item_package_dimensions",
    "item_width_height",
    "item_weight",
    "item_package_weight",
    "supplier_declared_dg_hz_regulation",
}

PRESET_FIELD_SPECS = {
    "item_name": [{"key": "item_name", "group": "identity"}],
    "brand": [{"key": "brand", "group": "identity"}],
    "manufacturer": [{"key": "manufacturer", "group": "identity"}],
    "model_number": [{"key": "model_number", "group": "identity"}],
    "model_name": [{"key": "model_name", "group": "identity"}],
    "item_type_keyword": [{"key": "item_type_keyword", "group": "identity"}],
    "merchant_suggested_asin": [{"key": "asin", "group": "identity"}],
    "product_description": [{"key": "product_description", "group": "content"}],
    "bullet_point": [
        {"key": f"bullet_point_{idx}", "group": "content"}
        for idx in range(1, 6)
    ],
    "generic_keyword": [{"key": "generic_keywords", "group": "content"}],
    "main_product_image_locator": [{"key": "main_image_url", "group": "images"}],
    **{
        f"other_product_image_locator_{idx}": [{"key": f"other_image_url_{idx}", "group": "images"}]
        for idx in range(1, 9)
    },
    "purchasable_offer": [
        {"key": "standard_price", "group": "offer"},
        {"key": "currency", "group": "offer", "level": "recommended", "default": "USD"},
    ],
    "fulfillment_availability": [
        {"key": "quantity", "group": "offer"},
        {"key": "fulfillment_channel", "group": "offer", "level": "recommended", "default": "DEFAULT"},
    ],
    "condition_type": [{"key": "condition_type", "group": "offer", "default": "new_new"}],
    "externally_assigned_product_identifier": [
        {"key": "upc", "group": "identity"},
        {"key": "external_product_id_type", "group": "identity", "level": "recommended", "default": "UPC"},
        {"key": "product_identity_mode", "group": "identity", "level": "recommended", "default": "real_gtin"},
    ],
    "item_weight": [{"key": "item_weight", "group": "shipping"}],
    "item_package_weight": [
        {"key": "package_weight", "group": "shipping"},
        {"key": "package_weight_unit", "group": "shipping", "level": "recommended"},
    ],
    "item_dimensions": [
        {"key": "item_length", "group": "shipping"},
        {"key": "item_width", "group": "shipping"},
        {"key": "item_height", "group": "shipping"},
        {"key": "dimension_unit", "group": "shipping", "level": "recommended"},
    ],
    "item_width_height": [
        {"key": "item_width", "group": "shipping"},
        {"key": "item_height", "group": "shipping"},
        {"key": "dimension_unit", "group": "shipping", "level": "recommended"},
    ],
    "item_package_dimensions": [
        {"key": "package_length", "group": "shipping"},
        {"key": "package_width", "group": "shipping"},
        {"key": "package_height", "group": "shipping"},
        {"key": "dimension_unit", "group": "shipping", "level": "recommended"},
    ],
    "supplier_declared_dg_hz_regulation": [{"key": "hazmat_declaration", "group": "compliance"}],
}


def build_template_definition(
    schema_fields: Dict,
    product_type: str,
    marketplace: str = "US",
    variation_mode: str = "single",
    template_id: Optional[str] = None,
    template_version: str = "2.0",
) -> Dict:
    """将 schema 字段清单转换为 Excel 模板定义。"""
    variation_mode = "variation" if str(variation_mode or "").strip().lower() == "variation" else "single"
    required = schema_fields.get("required_fields", []) or []
    optional = schema_fields.get("optional_fields", []) or []
    enum_fields = schema_fields.get("enum_fields", {}) or {}
    field_groups = schema_fields.get("field_groups", {}) or {}
    required_names = {str(field.get("name", "")).strip() for field in required if str(field.get("name", "")).strip()}

    columns: List[Dict] = []
    seen = {}

    def add_column(spec: Dict, field_info: Optional[Dict] = None):
        key = str(spec.get("key", "")).strip()
        if not key:
            return
        column = {
            "key": key,
            "label_zh": spec.get("label_zh") or LABEL_ZH.get(key) or _humanize_key(key),
            "label_en": spec.get("label_en") or (field_info or {}).get("title") or key,
            "group": spec.get("group") or GROUP_MAP.get((field_info or {}).get("group", ""), "other"),
            "level": spec.get("level") or ("required" if spec.get("required") else "optional"),
            "description": spec.get("description") or (field_info or {}).get("description", "") or "",
            "example": spec.get("example") or FIELD_EXAMPLES.get(key, ""),
            "enum_values": list(spec.get("enum_values") or []),
            "source_attribute": spec.get("source_attribute") or (field_info or {}).get("name", key),
            "default": spec.get("default", ""),
            "requirement_note": spec.get("requirement_note", ""),
        }

        if field_info:
            field_name = str(field_info.get("name", "")).strip()
            if not spec.get("level"):
                column["level"] = "required" if field_name in required_names else "optional"
            if field_name in enum_fields and not column["enum_values"]:
                column["enum_values"] = list(enum_fields.get(field_name) or [])
            if field_info.get("enum") and not column["enum_values"]:
                column["enum_values"] = list(field_info.get("enum_values") or [])

        existing = seen.get(key)
        if existing:
            merged = columns[existing]
            if LEVEL_ORDER.get(column["level"], 9) < LEVEL_ORDER.get(merged["level"], 9):
                merged["level"] = column["level"]
            if column["description"] and not merged["description"]:
                merged["description"] = column["description"]
            if column["example"] and not merged["example"]:
                merged["example"] = column["example"]
            if column["enum_values"] and not merged["enum_values"]:
                merged["enum_values"] = column["enum_values"]
            if column["default"] and not merged["default"]:
                merged["default"] = column["default"]
            if column["requirement_note"] and not merged["requirement_note"]:
                merged["requirement_note"] = column["requirement_note"]
            return

        seen[key] = len(columns)
        columns.append(column)

    # 固定控制字段
    base_columns = [
        {"key": "sku", "label_en": "Seller SKU", "group": "meta", "level": "required", "example": FIELD_EXAMPLES["sku"]},
        {
            "key": "product_type",
            "label_en": "Product Type",
            "group": "meta",
            "level": "required",
            "example": product_type,
            "default": product_type,
            "description": "Amazon 官方产品类型，建议整列保持一致",
        },
        {
            "key": "product_identity_mode",
            "label_en": "Product Identity Mode",
            "group": "identity",
            "level": "recommended",
            "default": "real_gtin",
            "example": "real_gtin / gtin_exemption",
            "description": "美国站第一版建议默认真实 GTIN，免码商品再改为 gtin_exemption",
        },
        {"key": "upc", "label_en": "UPC / EAN / GTIN", "group": "identity", "level": "recommended", "example": FIELD_EXAMPLES["upc"]},
        {"key": "external_product_id_type", "label_en": "Product ID Type", "group": "identity", "level": "recommended", "default": "UPC"},
    ]
    if variation_mode == "variation":
        base_columns.extend([
            {
                "key": "parentage_level",
                "label_en": "Parentage Level",
                "group": "variation",
                "level": "required",
                "enum_values": ["parent", "child"],
                "default": "child",
                "requirement_note": "父体填 parent，子体填 child",
            },
            {
                "key": "parent_sku",
                "label_en": "Parent SKU",
                "group": "variation",
                "level": "required",
                "requirement_note": "父体可留空，子体必须填写所属父SKU",
            },
            {
                "key": "variation_theme",
                "label_en": "Variation Theme",
                "group": "variation",
                "level": "required",
                "example": FIELD_EXAMPLES["variation_theme"],
            },
            {"key": "color", "label_en": "Color", "group": "variation", "level": "recommended"},
            {"key": "size", "label_en": "Size", "group": "variation", "level": "recommended"},
        ])

    for spec in base_columns:
        add_column(spec)

    for field in required + optional:
        field_name = str(field.get("name", "")).strip()
        if not field_name:
            continue
        preset = PRESET_FIELD_SPECS.get(field_name)
        if preset:
            for spec in preset:
                add_column(dict(spec, source_attribute=field_name), field_info=field)
            continue
        if field_name in SIMPLE_ATTRIBUTE_EXCLUDE:
            continue
        if not re.match(r"^[a-z][a-z0-9_]*$", field_name):
            continue
        add_column({"key": field_name, "source_attribute": field_name}, field_info=field)

    columns.sort(key=_column_sort_key)

    required_total = sum(1 for column in columns if column["level"] == "required")
    recommended_total = sum(1 for column in columns if column["level"] == "recommended")

    return {
        "template_id": template_id or _build_template_id(product_type, marketplace, variation_mode),
        "template_version": template_version,
        "product_type": product_type,
        "marketplace": marketplace,
        "variation_mode": variation_mode,
        "generated_at": "",
        "required_total": required_total,
        "recommended_total": recommended_total,
        "columns": columns,
        "field_groups": field_groups,
    }


def generate_template(
    template_or_schema_fields: Dict,
    output_path: str,
    *,
    product_type: str = "",
    marketplace: str = "US",
    variation_mode: str = "single",
) -> str:
    """
    生成 Excel 模板。

    兼容两种入参：
    - 直接传完整 template_definition
    - 传 parse_schema 的结果，并额外提供 product_type
    """
    if "columns" in template_or_schema_fields and "product_type" in template_or_schema_fields:
        template_def = dict(template_or_schema_fields)
    else:
        template_def = build_template_definition(
            template_or_schema_fields,
            product_type=product_type,
            marketplace=marketplace,
            variation_mode=variation_mode,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Amazon Template"
    columns = list(template_def.get("columns", []))

    for col_idx, column in enumerate(columns, start=1):
        note = _build_note_text(column)
        note_cell = ws.cell(row=1, column=col_idx, value=note)
        note_cell.font = DESC_FONT
        note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        note_cell.border = THIN_BORDER
        note_cell.fill = _fill_for_level(column["level"])

        header_cell = ws.cell(row=2, column=col_idx, value=column["key"])
        header_cell.font = HEADER_FONT
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_cell.border = THIN_BORDER
        header_cell.fill = _fill_for_level(column["level"])

        width = max(
            16,
            min(
                34,
                max(
                    len(str(column.get("key", ""))),
                    len(str(column.get("label_zh", ""))),
                    len(str(column.get("label_en", ""))) // 2,
                ) + 4,
            ),
        )
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 42
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    enum_sheet = wb.create_sheet("Enum Values")
    enum_sheet.sheet_state = "hidden"
    enum_col = 1
    for col_idx, column in enumerate(columns, start=1):
        values = list(column.get("enum_values") or [])
        if not values:
            continue
        enum_sheet.cell(row=1, column=enum_col, value=column["key"]).font = META_FONT
        for row_idx, value in enumerate(values[:500], start=2):
            enum_sheet.cell(row=row_idx, column=enum_col, value=str(value))
        if len(values) <= 50 and len(",".join(str(v) for v in values)) <= 240:
            formula = '"' + ",".join(str(v) for v in values) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=column["level"] != "required")
            dv.error = f"{column['key']} 必须从下拉选项中选择"
            dv.errorTitle = "无效值"
            col_letter = ws.cell(row=2, column=col_idx).column_letter
            dv.add(f"{col_letter}3:{col_letter}2000")
            ws.add_data_validation(dv)
        enum_col += 1

    meta_sheet = wb.create_sheet("__template_meta__")
    meta_sheet.sheet_state = "hidden"
    meta_rows = [
        ("template_id", template_def.get("template_id", "")),
        ("template_version", template_def.get("template_version", "")),
        ("product_type", template_def.get("product_type", "")),
        ("marketplace", template_def.get("marketplace", "")),
        ("variation_mode", template_def.get("variation_mode", "")),
        ("required_total", int(template_def.get("required_total", 0) or 0)),
        ("recommended_total", int(template_def.get("recommended_total", 0) or 0)),
    ]
    for row_idx, (key, value) in enumerate(meta_rows, start=1):
        meta_sheet.cell(row=row_idx, column=1, value=key)
        meta_sheet.cell(row=row_idx, column=2, value=value)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def _build_template_id(product_type: str, marketplace: str, variation_mode: str) -> str:
    safe_product_type = re.sub(r"[^0-9A-Za-z_-]+", "_", str(product_type or "").strip()).strip("_") or "product"
    safe_marketplace = re.sub(r"[^0-9A-Za-z_-]+", "_", str(marketplace or "US").strip()).strip("_") or "US"
    safe_variation = re.sub(r"[^0-9A-Za-z_-]+", "_", str(variation_mode or "single").strip()).strip("_") or "single"
    return f"{safe_marketplace.lower()}_{safe_product_type.lower()}_{safe_variation.lower()}"


def _column_sort_key(column: Dict):
    group = str(column.get("group", "other") or "other").strip()
    group_idx = GROUP_ORDER.index(group) if group in GROUP_ORDER else len(GROUP_ORDER)
    return (
        group_idx,
        LEVEL_ORDER.get(str(column.get("level", "optional") or "optional").strip(), 9),
        str(column.get("key", "")),
    )


def _fill_for_level(level: str):
    normalized = str(level or "").strip().lower()
    if normalized == "required":
        return REQUIRED_FILL
    if normalized == "recommended":
        return RECOMMENDED_FILL
    if normalized == "meta":
        return META_FILL
    return OPTIONAL_FILL


def _humanize_key(key: str) -> str:
    words = [segment for segment in str(key or "").replace("_", " ").split() if segment]
    if not words:
        return key
    return " ".join(word.capitalize() for word in words)


def _build_note_text(column: Dict) -> str:
    prefix_map = {
        "required": "[必填]",
        "recommended": "[建议]",
        "optional": "[选填]",
        "meta": "[系统]",
    }
    level = str(column.get("level", "optional") or "optional").strip().lower()
    parts = [prefix_map.get(level, "[选填]")]
    label = str(column.get("label_zh", "") or "").strip()
    if label:
        parts.append(label)
    label_en = str(column.get("label_en", "") or "").strip()
    if label_en and label_en.lower() != label.lower():
        parts.append(f"/ {label_en}")
    description = str(column.get("description", "") or "").strip()
    if description:
        parts.append(f"· {description}")
    requirement_note = str(column.get("requirement_note", "") or "").strip()
    if requirement_note:
        parts.append(f"· {requirement_note}")
    example = str(column.get("example", "") or "").strip()
    if example:
        parts.append(f"· 示例: {example}")
    default = str(column.get("default", "") or "").strip()
    if default:
        parts.append(f"· 默认: {default}")
    return " ".join(part for part in parts if part)
