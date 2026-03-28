"""
模板推荐、模板缓存、模板识别与诊断辅助。
"""
from __future__ import annotations

import json
import os
import re
import time
from html import unescape
from typing import Dict, List, Optional, Tuple
from urllib.parse import parse_qs, unquote, urlparse

import requests
from openpyxl import load_workbook

from config import get_config
from core.template_generator import CURRENT_TEMPLATE_VERSION, build_template_definition, generate_template


TEMPLATE_CACHE_TTL = 7 * 24 * 3600
DEFAULT_MARKETPLACE = "US"
DEFAULT_VARIATION_MODE = "single"
TEMPLATE_META_SHEET = "__template_meta__"

FALLBACK_TYPE_HINTS = [
    (("mouse", "gaming mouse", "wireless mouse"), "COMPUTER_INPUT_DEVICE", "Computer Input Device"),
    (("mug", "cup", "tumbler", "bottle", "thermos"), "DRINKING_CUP", "Drinking Cup"),
    (("phone case", "iphone case", "case", "cover"), "CELLULAR_PHONE_CASE", "Cellular Phone Case"),
    (("headphone", "earbud", "earphone"), "HEADPHONES", "Headphones"),
]


def _base_output_dir() -> str:
    return str(get_config().OUTPUT_DIR)


def _template_definition_dir() -> str:
    path = os.path.join(_base_output_dir(), "template_definitions")
    os.makedirs(path, exist_ok=True)
    return path


def _generated_template_dir() -> str:
    path = os.path.join(_base_output_dir(), "generated_templates")
    os.makedirs(path, exist_ok=True)
    return path


def _template_overlay_dir() -> str:
    path = os.path.join(_base_output_dir(), "template_overlays")
    os.makedirs(path, exist_ok=True)
    return path


def _definition_path(template_id: str) -> str:
    return os.path.join(_template_definition_dir(), f"{template_id}.json")


def _workbook_path(template_id: str) -> str:
    return os.path.join(_generated_template_dir(), f"{template_id}.xlsx")


def _overlay_path(product_type: str, marketplace: str) -> str:
    safe_product_type = re.sub(r"[^0-9A-Za-z_-]+", "_", str(product_type or "").strip()).strip("_") or "product"
    safe_marketplace = re.sub(r"[^0-9A-Za-z_-]+", "_", str(marketplace or DEFAULT_MARKETPLACE).strip()).strip("_") or DEFAULT_MARKETPLACE
    return os.path.join(_template_overlay_dir(), f"{safe_marketplace.lower()}_{safe_product_type.lower()}.json")


def extract_source_keyword(source_url: str = "", title: str = "", keyword: str = "") -> Dict[str, str]:
    """尽量从输入链接/标题/关键词中提取一个适合官方产品类型搜索的 query。"""
    explicit_title = str(title or "").strip()
    explicit_keyword = str(keyword or "").strip()
    source_url = str(source_url or "").strip()

    if explicit_keyword:
        return {"query": explicit_keyword, "title": explicit_title or explicit_keyword, "source": "keyword"}
    if explicit_title:
        return {"query": explicit_title, "title": explicit_title, "source": "title"}

    fetched_title = _extract_title_from_url(source_url)
    if fetched_title:
        return {"query": fetched_title, "title": fetched_title, "source": "url_title"}

    url_title = _guess_title_from_url(source_url)
    if url_title:
        return {"query": url_title, "title": url_title, "source": "url_slug"}

    return {"query": "", "title": "", "source": ""}


def recommend_product_types(
    *,
    source_url: str = "",
    title: str = "",
    keyword: str = "",
    marketplace: str = DEFAULT_MARKETPLACE,
) -> Dict[str, object]:
    info = extract_source_keyword(source_url=source_url, title=title, keyword=keyword)
    query = str(info.get("query", "") or "").strip()
    candidates: List[Dict] = []
    source = str(info.get("source", "") or "").strip()

    if query:
        try:
            from amazon.sp_client import SPClient
            from amazon.schema_manager import search_product_types

            client = SPClient(marketplace_id="ATVPDKIKX0DER")
            raw_candidates = search_product_types(query, sp_client=client) or []
            candidates.extend(_normalize_product_type_candidates(raw_candidates))
            source = source or "amazon_api"
        except Exception:
            pass

    if not candidates and query:
        candidates.extend(_fallback_product_type_candidates(query))
        source = source or "heuristic"

    return {
        "query": query,
        "title": str(info.get("title", "") or "").strip(),
        "source": source or "empty",
        "marketplace": marketplace,
        "candidates": candidates[:8],
    }


def ensure_template_definition(
    product_type: str,
    marketplace: str = DEFAULT_MARKETPLACE,
    variation_mode: str = DEFAULT_VARIATION_MODE,
    refresh: bool = False,
) -> Dict:
    product_type = str(product_type or "").strip()
    marketplace = str(marketplace or DEFAULT_MARKETPLACE).strip().upper() or DEFAULT_MARKETPLACE
    variation_mode = "variation" if str(variation_mode or "").strip().lower() == "variation" else "single"
    if not product_type:
        raise ValueError("缺少 product_type")

    template_id = _template_id(product_type, marketplace, variation_mode)
    definition_path = _definition_path(template_id)
    if not refresh and os.path.exists(definition_path):
        mtime = os.path.getmtime(definition_path)
        if time.time() - mtime < TEMPLATE_CACHE_TTL:
            with open(definition_path, "r", encoding="utf-8") as fh:
                definition = json.load(fh)
            if _is_template_definition_current(definition):
                _apply_template_overlay(definition)
                with open(definition_path, "w", encoding="utf-8") as fh:
                    json.dump(definition, fh, ensure_ascii=False, indent=2)
                return definition

    from amazon.schema_manager import fetch_schema, parse_schema

    raw = fetch_schema(product_type, marketplace)
    schema_fields = parse_schema(raw)
    definition = build_template_definition(
        schema_fields=schema_fields,
        product_type=product_type,
        marketplace=marketplace,
        variation_mode=variation_mode,
        template_id=template_id,
    )
    definition["generated_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
    _apply_template_overlay(definition)

    with open(definition_path, "w", encoding="utf-8") as fh:
        json.dump(definition, fh, ensure_ascii=False, indent=2)
    return definition


def ensure_template_workbook(template_id: str) -> Tuple[str, Dict]:
    definition = load_template_definition(template_id)
    output_path = _workbook_path(template_id)
    generate_template(definition, output_path)
    return output_path, definition


def load_template_definition(template_id: str) -> Dict:
    template_id = str(template_id or "").strip()
    path = _definition_path(template_id)
    if not template_id or not os.path.exists(path):
        raise FileNotFoundError(f"模板定义不存在: {template_id}")
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def read_template_metadata(filepath: str) -> Dict[str, str]:
    """从模板工作簿隐藏页提取元数据。"""
    if not filepath or not os.path.exists(filepath):
        return {}

    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        if TEMPLATE_META_SHEET not in wb.sheetnames:
            return {}
        ws = wb[TEMPLATE_META_SHEET]
        meta = {}
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            key = str(row[0] or "").strip()
            if not key:
                continue
            meta[key] = str(row[1] or "").strip()
        return meta
    finally:
        wb.close()


def _is_template_definition_current(definition: Dict) -> bool:
    version = str(definition.get("template_version", "") or "").strip()
    if version != CURRENT_TEMPLATE_VERSION:
        return False
    columns = {str(item.get("key", "") or "").strip() for item in (definition.get("columns", []) or [])}
    if "hazmat_declaration" in columns and "supplier_declared_dg_hz_regulation" not in columns:
        return False
    return True


def evaluate_template_row(
    row: Dict,
    template_definition: Optional[Dict],
    *,
    col_map: Optional[Dict] = None,
) -> Dict:
    """按模板定义检查当前行的已填/缺失情况。"""
    if not template_definition:
        return {
            "template_id": "",
            "required_total": 0,
            "required_filled": 0,
            "required_missing": [],
            "recommended_missing": [],
            "optional_missing": [],
            "blocking_issues": [],
            "variation_issues": [],
        }

    col_map = col_map or {}
    columns = list(template_definition.get("columns", []))
    required_missing = []
    recommended_missing = []
    optional_missing = []

    for column in columns:
        key = str(column.get("key", "") or "").strip()
        value = _pick_row_value(row, key, col_map)
        if value:
            continue
        info = {
            "key": key,
            "label_zh": str(column.get("label_zh", "") or key).strip(),
            "label_en": str(column.get("label_en", "") or key).strip(),
            "level": str(column.get("level", "optional") or "optional").strip(),
            "group": str(column.get("group", "") or "other").strip(),
        }
        if info["level"] == "required":
            required_missing.append(info)
        elif info["level"] == "recommended":
            recommended_missing.append(info)
        else:
            optional_missing.append(info)

    variation_issues = _evaluate_variation_row(row, template_definition, col_map)
    blocking = [f"{item['label_zh']} ({item['key']})" for item in required_missing]
    blocking.extend(variation_issues)

    required_total = sum(1 for column in columns if str(column.get("level", "")).strip() == "required")
    return {
        "template_id": str(template_definition.get("template_id", "") or "").strip(),
        "required_total": required_total,
        "required_filled": max(required_total - len(required_missing), 0),
        "required_missing": required_missing,
        "recommended_missing": recommended_missing,
        "optional_missing": optional_missing,
        "blocking_issues": blocking,
        "variation_issues": variation_issues,
    }


def summarize_template_issues(result: Dict) -> str:
    template_data = result.get("template", {}) or {}
    parts = []
    required_missing = template_data.get("required_missing", []) or []
    recommended_missing = template_data.get("recommended_missing", []) or []
    variation_issues = template_data.get("variation_issues", []) or []
    if required_missing:
        parts.append(f"模板必填缺失 {len(required_missing)} 项")
    if recommended_missing:
        parts.append(f"建议补充 {len(recommended_missing)} 项")
    if variation_issues:
        parts.append(f"变体问题 {len(variation_issues)} 项")
    return "；".join(parts)


def update_template_overlay(product_type: str, marketplace: str, missing_fields: List[Dict]):
    product_type = str(product_type or "").strip()
    marketplace = str(marketplace or DEFAULT_MARKETPLACE).strip().upper() or DEFAULT_MARKETPLACE
    if not product_type:
        return

    overlay_path = _overlay_path(product_type, marketplace)
    overlay = {}
    if os.path.exists(overlay_path):
        with open(overlay_path, "r", encoding="utf-8") as fh:
            overlay = json.load(fh) or {}

    fields = {str(item.get('key', '') or item.get('name', '') or '').strip(): dict(item) for item in (overlay.get('fields') or []) if str(item.get('key', '') or item.get('name', '') or '').strip()}
    for field in missing_fields or []:
        key = str(field.get('name', '') or field.get('key', '') or '').strip()
        if not key:
            continue
        title = str(field.get('title', '') or key).strip() or key
        fields[key] = {
            'key': key,
            'label_zh': title,
            'label_en': str(field.get('label_en', '') or title).strip() or title,
            'level': 'required',
            'group': str(field.get('group', '') or 'overlay').strip() or 'overlay',
            'description': (
                str(field.get('description', '') or '').strip()
                or '历史 Amazon 预览中出现的真实阻断缺字段，建议在模板中提前填写。'
            ),
            'source_attribute': key,
        }

    overlay = {
        'product_type': product_type,
        'marketplace': marketplace,
        'updated_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        'fields': list(fields.values()),
    }
    with open(overlay_path, 'w', encoding='utf-8') as fh:
        json.dump(overlay, fh, ensure_ascii=False, indent=2)


def _apply_template_overlay(definition: Dict):
    overlay_path = _overlay_path(definition.get('product_type', ''), definition.get('marketplace', DEFAULT_MARKETPLACE))
    if not os.path.exists(overlay_path):
        return definition

    with open(overlay_path, 'r', encoding='utf-8') as fh:
        overlay = json.load(fh) or {}

    columns = list(definition.get('columns', []) or [])
    by_key = {str(item.get('key', '') or '').strip(): item for item in columns if str(item.get('key', '') or '').strip()}
    changed = False
    for field in overlay.get('fields') or []:
        key = str(field.get('key', '') or '').strip()
        if not key:
            continue
        if key in by_key:
            if by_key[key].get('level') != 'required':
                by_key[key]['level'] = 'required'
                changed = True
            if field.get('description') and field.get('description') not in str(by_key[key].get('description', '') or ''):
                by_key[key]['description'] = str(by_key[key].get('description', '') or '').strip() + ' ' + str(field.get('description', '') or '').strip()
                changed = True
            continue
        columns.append(dict(field))
        changed = True

    if changed:
        def level_rank(value):
            return {'required': 0, 'recommended': 1, 'optional': 2}.get(str(value or '').strip(), 9)

        columns.sort(key=lambda item: (level_rank(item.get('level')), str(item.get('group', 'overlay')), str(item.get('key', ''))))
        definition['columns'] = columns
        definition['required_total'] = sum(1 for item in columns if str(item.get('level', '')).strip() == 'required')
        definition['recommended_total'] = sum(1 for item in columns if str(item.get('level', '')).strip() == 'recommended')
    return definition


def template_definition_summary(template_definition: Optional[Dict]) -> Dict[str, object]:
    template_definition = template_definition or {}
    columns = list(template_definition.get("columns", []))
    return {
        "template_id": str(template_definition.get("template_id", "") or "").strip(),
        "product_type": str(template_definition.get("product_type", "") or "").strip(),
        "variation_mode": str(template_definition.get("variation_mode", "single") or "single").strip(),
        "required_total": int(template_definition.get("required_total", 0) or 0),
        "recommended_total": int(template_definition.get("recommended_total", 0) or 0),
        "column_count": len(columns),
    }


def _template_id(product_type: str, marketplace: str, variation_mode: str) -> str:
    safe_product_type = re.sub(r"[^0-9A-Za-z_-]+", "_", str(product_type or "").strip()).strip("_") or "product"
    return f"{marketplace.lower()}_{safe_product_type.lower()}_{variation_mode.lower()}"


def _normalize_product_type_candidates(raw_candidates: List[Dict]) -> List[Dict]:
    normalized = []
    seen = set()
    for item in raw_candidates:
        product_type = str(
            item.get("name")
            or item.get("productType")
            or item.get("product_type")
            or ""
        ).strip()
        if not product_type or product_type in seen:
            continue
        seen.add(product_type)
        normalized.append({
            "product_type": product_type,
            "display_name": str(
                item.get("displayName")
                or item.get("display_name")
                or item.get("name")
                or product_type
            ).strip(),
            "score": item.get("score") or item.get("relevance") or 0,
        })
    return normalized


def _fallback_product_type_candidates(query: str) -> List[Dict]:
    text = str(query or "").strip().lower()
    if not text:
        return []

    results = []
    for hints, product_type, display_name in FALLBACK_TYPE_HINTS:
        if any(hint in text for hint in hints):
            results.append({
                "product_type": product_type,
                "display_name": display_name,
                "score": 0.8,
            })
    if results:
        return results

    normalized = re.sub(r"[^0-9A-Za-z]+", "_", text).strip("_").upper() or "PRODUCT"
    return [{
        "product_type": normalized,
        "display_name": normalized.replace("_", " ").title(),
        "score": 0.3,
    }]


def _extract_title_from_url(source_url: str) -> str:
    text = str(source_url or "").strip()
    if not text or not re.match(r"^https?://", text, re.IGNORECASE):
        return ""

    try:
        response = requests.get(
            text,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                              "(KHTML, like Gecko) Chrome/123.0 Safari/537.36",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            },
            timeout=8,
        )
        response.raise_for_status()
    except Exception:
        return ""

    html = response.text or ""
    patterns = [
        r'<meta[^>]+property=["\']og:title["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+name=["\']twitter:title["\'][^>]+content=["\']([^"\']+)["\']',
        r'<title>(.*?)</title>',
    ]
    for pattern in patterns:
        match = re.search(pattern, html, flags=re.IGNORECASE | re.DOTALL)
        if match:
            return _clean_title(match.group(1))
    return ""


def _guess_title_from_url(source_url: str) -> str:
    text = str(source_url or "").strip()
    if not text:
        return ""
    try:
        parsed = urlparse(text)
    except Exception:
        return ""

    query = parse_qs(parsed.query or "")
    for key in ("title", "name", "keyword", "itemName"):
        if query.get(key):
            value = _clean_title(query[key][0])
            if value:
                return value

    path = unquote(parsed.path or "")
    segments = [segment for segment in path.split("/") if segment]
    if not segments:
        return ""
    candidate = segments[-1]
    candidate = re.sub(r"\.[A-Za-z0-9]{1,6}$", "", candidate)
    candidate = re.sub(r"[-_]+", " ", candidate)
    candidate = re.sub(r"\s+", " ", candidate)
    return _clean_title(candidate)


def _clean_title(value: str) -> str:
    text = unescape(str(value or "").strip())
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"\s*[|｜-]\s*.+$", "", text)
    return text.strip()


def _pick_row_value(row: Dict, key: str, col_map: Dict) -> str:
    direct = row.get(key)
    if str(direct or "").strip():
        return str(direct or "").strip()
    mapped_col = col_map.get(key)
    if mapped_col and str(row.get(mapped_col, "") or "").strip():
        return str(row.get(mapped_col, "") or "").strip()
    return ""


def _evaluate_variation_row(row: Dict, template_definition: Dict, col_map: Dict) -> List[str]:
    if str(template_definition.get("variation_mode", "single") or "single").strip() != "variation":
        return []

    parentage = _pick_row_value(row, "parentage_level", col_map).lower()
    parent_sku = _pick_row_value(row, "parent_sku", col_map)
    variation_theme = _pick_row_value(row, "variation_theme", col_map)
    color = _pick_row_value(row, "color", col_map)
    size = _pick_row_value(row, "size", col_map)
    issues = []

    if parentage not in ("parent", "child"):
        issues.append("变体模板要求填写 parentage_level，且值必须是 parent 或 child")
    if not variation_theme:
        issues.append("变体模板要求填写 variation_theme")
    if parentage == "child" and not parent_sku:
        issues.append("子体缺少 parent_sku")
    if parentage == "child" and not (color or size):
        issues.append("子体至少需要填写颜色或尺寸中的一个变体维度")
    return issues
