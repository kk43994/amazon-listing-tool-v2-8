"""
生成亚马逊商品采集模板Excel
基于SP-API putListingsItem支持的全部上架字段
给甲方使用：按此模板采集竞品数据 → 我们AI仿写 → SP-API上架
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# Sheet 1: 采集模板（甲方填写）
# ============================================================
ws = wb.active
ws.title = "商品采集模板"

# 颜色定义
COLORS = {
    'red':    'FFDC3545',  # 必填
    'orange': 'FFFD7E14',  # AI仿写
    'green':  'FF28A745',  # 图片
    'blue':   'FF007BFF',  # 销售
    'purple': 'FF6F42C1',  # 物流
    'gray':   'FF6C757D',  # 属性
    'brown':  'FF795548',  # 合规
    'teal':   'FF17A2B8',  # 变体
}

thin_border = Border(
    left=Side(style='thin', color='FFCCCCCC'),
    right=Side(style='thin', color='FFCCCCCC'),
    top=Side(style='thin', color='FFCCCCCC'),
    bottom=Side(style='thin', color='FFCCCCCC'),
)

# 列定义: (列名, 分组, 颜色, 宽度, 必填?, 说明, 示例值)
COLUMNS = [
    # === 产品身份 (必填) ===
    ('SKU', '产品身份', 'red', 15, True,
     '卖家自定义商品编号，唯一标识',
     'MY-EARBUDS-001'),
    ('brand_name', '产品身份', 'red', 18, True,
     '你的品牌名（不是竞品品牌！填你自己的）',
     'MyBrand'),
    ('product_type', '产品身份', 'red', 20, True,
     '亚马逊产品类型，见Sheet3"产品类型参考"',
     'WIRELESS_ACCESSORY'),
    ('external_product_id', '产品身份', 'red', 18, False,
     'UPC/EAN条码（如有）',
     '012345678901'),
    ('external_product_id_type', '产品身份', 'red', 12, False,
     '条码类型: UPC / EAN / GTIN',
     'UPC'),
    ('manufacturer', '产品身份', 'red', 18, False,
     '制造商名称',
     'MyBrand Inc.'),
    ('model_number', '产品身份', 'red', 15, False,
     '型号',
     'MB-TWS-2026'),

    # === AI仿写内容（竞品原文，我们会AI改写） ===
    ('item_name', 'AI仿写', 'orange', 50, True,
     '【竞品原标题】采集竞品的完整标题，我们会AI改写',
     'Wireless Earbuds, Bluetooth 5.3 Headphones with ENC Noise Cancelling Mic, 48H Playtime, IPX7 Waterproof'),
    ('bullet_point_1', 'AI仿写', 'orange', 45, True,
     '【竞品原卖点1】第1条Bullet Point',
     'SUPERIOR SOUND QUALITY — Custom 13mm drivers deliver deep bass and crystal-clear treble...'),
    ('bullet_point_2', 'AI仿写', 'orange', 45, True,
     '【竞品原卖点2】第2条Bullet Point',
     'LONG BATTERY LIFE — 8 hours per charge with 48 hours total with charging case...'),
    ('bullet_point_3', 'AI仿写', 'orange', 45, True,
     '【竞品原卖点3】第3条Bullet Point',
     'COMFORTABLE FIT — Ergonomic design with 3 sizes of ear tips for all-day comfort...'),
    ('bullet_point_4', 'AI仿写', 'orange', 45, False,
     '【竞品原卖点4】第4条',
     'WATERPROOF DESIGN — IPX7 rated for sweat, rain, and splashes...'),
    ('bullet_point_5', 'AI仿写', 'orange', 45, False,
     '【竞品原卖点5】第5条',
     'EASY CONNECTIVITY — Bluetooth 5.3 provides stable connection within 50ft range...'),
    ('product_description', 'AI仿写', 'orange', 60, True,
     '【竞品原描述】完整商品描述（可含HTML标签）',
     '<b>Premium Wireless Earbuds</b><br>Experience immersive sound with our latest...'),
    ('generic_keywords', 'AI仿写', 'orange', 40, False,
     '【竞品搜索词】参考用，我们会重新生成',
     'wireless earbuds bluetooth headphones noise cancelling'),

    # === 图片（竞品图片URL，我们会AI换背景） ===
    ('main_image_url', '图片', 'green', 40, True,
     '【竞品主图URL】主图链接，我们会AI换背景',
     'https://m.media-amazon.com/images/I/xxxxx.jpg'),
    ('other_image_url_1', '图片', 'green', 40, False,
     '【竞品副图1】',
     'https://m.media-amazon.com/images/I/yyyyy.jpg'),
    ('other_image_url_2', '图片', 'green', 40, False,
     '【竞品副图2】', ''),
    ('other_image_url_3', '图片', 'green', 40, False,
     '【竞品副图3】', ''),
    ('other_image_url_4', '图片', 'green', 40, False,
     '【竞品副图4】', ''),
    ('other_image_url_5', '图片', 'green', 40, False,
     '【竞品副图5】', ''),
    ('other_image_url_6', '图片', 'green', 40, False,
     '【竞品副图6】', ''),
    ('other_image_url_7', '图片', 'green', 40, False,
     '【竞品副图7】', ''),
    ('other_image_url_8', '图片', 'green', 40, False,
     '【竞品副图8】', ''),

    # === 销售条款 ===
    ('standard_price', '销售', 'blue', 12, True,
     '你的售价���美元），不是竞品价格',
     '29.99'),
    ('currency', '销售', 'blue', 8, False,
     '货币（默认USD）',
     'USD'),
    ('quantity', '销售', 'blue', 10, True,
     '库存数量',
     '100'),
    ('condition_type', '销售', 'blue', 12, True,
     '商品状态: New / Refurbished / Used',
     'New'),
    ('fulfillment_channel', '销售', 'blue', 15, True,
     '配送方式: DEFAULT(自发) / AMAZON_NA(FBA)',
     'DEFAULT'),

    # === 物流尺寸 ===
    ('item_weight', '物流', 'purple', 12, False,
     '商品净重（克）',
     '45'),
    ('item_weight_unit', '物流', 'purple', 10, False,
     '重量单位: grams / kilograms / pounds / ounces',
     'grams'),
    ('item_length', '物流', 'purple', 10, False,
     '商品长度',
     '6.5'),
    ('item_width', '物流', 'purple', 10, False,
     '商品宽度',
     '5.0'),
    ('item_height', '物流', 'purple', 10, False,
     '商品高度',
     '3.2'),
    ('item_dimension_unit', '物流', 'purple', 10, False,
     '尺寸单位: centimeters / inches',
     'centimeters'),
    ('package_weight', '物流', 'purple', 12, False,
     '包装重量（克）',
     '180'),
    ('package_length', '物流', 'purple', 10, False,
     '包装长度', '15'),
    ('package_width', '物流', 'purple', 10, False,
     '包装宽度', '10'),
    ('package_height', '物流', 'purple', 10, False,
     '包装高度', '5'),

    # === 产品属性 ===
    ('color_name', '属性', 'gray', 12, False,
     '颜色',
     'Black'),
    ('size_name', '属性', 'gray', 12, False,
     '尺寸',
     'One Size'),
    ('material_type', '属性', 'gray', 15, False,
     '主要材质',
     'Silicone'),
    ('department', '属性', 'gray', 12, False,
     '部门: Mens / Womens / Unisex / Kids',
     'Unisex'),
    ('target_gender', '属性', 'gray', 12, False,
     '目标性别: Male / Female / Unisex',
     'Unisex'),
    ('age_range_description', '属性', 'gray', 15, False,
     '适用年龄: Adult / Kid / Baby',
     'Adult'),
    ('country_of_origin', '属性', 'gray', 12, False,
     '原产国',
     'China'),
    ('item_type_keyword', '属性', 'gray', 20, False,
     '商品类型关键词',
     'in-ear-headphones'),

    # === 合规安全 ===
    ('batteries_required', '合规', 'brown', 12, False,
     '是否需要电池: TRUE / FALSE',
     'TRUE'),
    ('batteries_included', '合规', 'brown', 12, False,
     '是否包含电池: TRUE / FALSE',
     'TRUE'),
    ('battery_cell_composition', '合规', 'brown', 15, False,
     '电池类型: lithium_ion / lithium_metal / alkaline 等',
     'lithium_ion'),
    ('lithium_battery_packaging', '合规', 'brown', 18, False,
     '锂电池包装: batteries_contained_in_equipment / batteries_packed_with_equipment',
     'batteries_contained_in_equipment'),
    ('supplier_declared_dg_hz_regulation', '合规', 'brown', 15, False,
     '危险品声明: not_applicable / ghs / ...',
     'not_applicable'),

    # === 变体 ===
    ('parent_child', '变体', 'teal', 10, False,
     '父/子: Parent / Child（无变体留空）',
     ''),
    ('parent_sku', '变体', 'teal', 15, False,
     '父SKU（子商品填写）',
     ''),
    ('relationship_type', '变体', 'teal', 12, False,
     '关系类型: Variation',
     ''),
    ('variation_theme', '变体', 'teal', 15, False,
     '变体主题: Color / Size / ColorSize 等',
     ''),

    # === 竞品参考（不上架，仅参考） ===
    ('competitor_asin', '参考', 'gray', 15, False,
     '【仅参考】竞品ASIN，不会提交到亚马逊',
     'B0XXXXXXXX'),
    ('competitor_price', '参考', 'gray', 12, False,
     '【仅参考】竞品价格，用于定价参考',
     '34.99'),
    ('competitor_rating', '参考', 'gray', 10, False,
     '【仅参考】竞品评分',
     '4.3'),
    ('competitor_reviews', '参考', 'gray', 10, False,
     '【仅参考】竞品评论数',
     '1253'),
    ('notes', '参考', 'gray', 30, False,
     '备注',
     ''),
]

# ===== 写第1行: 分组标题 =====
current_group = None
group_start_col = 1
for col_idx, (name, group, color, width, req, desc, example) in enumerate(COLUMNS, 1):
    if group != current_group:
        if current_group is not None:
            # 合并前一组
            if col_idx - 1 > group_start_col:
                ws.merge_cells(start_row=1, start_column=group_start_col,
                              end_row=1, end_column=col_idx - 1)
        current_group = group
        group_start_col = col_idx

    cell = ws.cell(row=1, column=col_idx, value=group)
    cell.fill = PatternFill(start_color=COLORS.get(color, 'FF6C757D'),
                            end_color=COLORS.get(color, 'FF6C757D'),
                            fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True, size=10)
    cell.alignment = Alignment(horizontal='center')

# 最后一组合并
if len(COLUMNS) > group_start_col:
    last_group_cols = [i for i, c in enumerate(COLUMNS, 1) if c[1] == current_group]
    if len(last_group_cols) > 1:
        ws.merge_cells(start_row=1, start_column=last_group_cols[0],
                      end_row=1, end_column=last_group_cols[-1])

# ===== 写第2行: 列名 =====
for col_idx, (name, group, color, width, req, desc, example) in enumerate(COLUMNS, 1):
    label = f"{'★ ' if req else ''}{name}"
    cell = ws.cell(row=2, column=col_idx, value=label)
    if req:
        cell.fill = PatternFill(start_color='FFFFF3CD', end_color='FFFFF3CD',
                               fill_type="solid")
        cell.font = Font(bold=True, size=10, color='FF856404')
    else:
        cell.fill = PatternFill(start_color='FFE9ECEF', end_color='FFE9ECEF',
                               fill_type="solid")
        cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# ===== 写第3行: 字段说明 =====
for col_idx, (name, group, color, width, req, desc, example) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=3, column=col_idx, value=desc)
    cell.font = Font(size=9, color='FF666666', italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = thin_border

# ===== 写第4行: 示例数据 =====
for col_idx, (name, group, color, width, req, desc, example) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=4, column=col_idx, value=example)
    cell.font = Font(size=10, color='FF333333')
    cell.border = thin_border
    cell.fill = PatternFill(start_color='FFF8F9FA', end_color='FFF8F9FA',
                           fill_type="solid")

# 列宽
for col_idx, (name, group, color, width, req, desc, example) in enumerate(COLUMNS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# 冻结前3行
ws.freeze_panes = 'A5'

# 行高
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 25
ws.row_dimensions[3].height = 50

# ============================================================
# Sheet 2: 填写说明
# ============================================================
ws2 = wb.create_sheet(title="填写说明")

instructions = [
    ("亚马逊商品采集模板 — 填写说明", "", "", ""),
    ("", "", "", ""),
    ("一、模板使用流程", "", "", ""),
    ("步骤", "操作", "说明", ""),
    ("1", "确定竞品", "在亚马逊上找到要仿的竞品商品", ""),
    ("2", "按模板采集", "复制竞品的标题、卖点、描述、图片URL等到对应列", ""),
    ("3", "填写自己的信息", "SKU、品牌、价格、库存等填你自己的", ""),
    ("4", "提交给我们", "我们会用AI改写标题/卖点/描述，换图片背景", ""),
    ("5", "确认后上架", "你确认AI改写的内容后，我们SP-API自动上架", ""),
    ("", "", "", ""),
    ("二、列颜色说明", "", "", ""),
    ("颜色", "含义", "详细说明", ""),
    ("🔴 红色 — 产品身份", "必填/自填", "SKU、品牌、产品类型等，填你自己的信息", ""),
    ("🟠 橙色 — AI仿写", "采集竞品", "从竞品listing复制，我们AI改写后替换", ""),
    ("🟢 绿色 — 图片", "采集竞品", "竞品图片URL，我们AI换背景后使用", ""),
    ("🔵 蓝色 — 销售", "自填", "你的定价、库存、配送方式", ""),
    ("🟣 紫色 — 物流", "采集/自填", "商品尺寸重量，可从竞品复制", ""),
    ("⚪ 灰色 — 属性", "采集/自填", "颜色、尺寸、材质等", ""),
    ("🟤 棕色 — 合规", "自填", "电池、危险品等安全信息", ""),
    ("🔵 青色 — 变体", "自填", "有多个颜色/尺寸时填写", ""),
    ("", "", "", ""),
    ("三、★标记字段为必填", "", "", ""),
    ("", "★ SKU — 唯一商品编号，你自己定义", "", ""),
    ("", "★ brand_name — 你的品牌名（不是竞品品牌！）", "", ""),
    ("", "★ product_type — 亚马逊产品类型，见Sheet3", "", ""),
    ("", "★ item_name — 竞品标题（我们AI改写）", "", ""),
    ("", "★ bullet_point_1~3 — 竞品前3条卖点（至少3条）", "", ""),
    ("", "★ product_description — 竞品描述（我们AI改写）", "", ""),
    ("", "★ main_image_url — 竞品主图URL（我们AI换背景）", "", ""),
    ("", "★ standard_price — 你的售价（美元）", "", ""),
    ("", "★ quantity — 库存数量", "", ""),
    ("", "★ condition_type — 商品状态（一般填New）", "", ""),
    ("", "★ fulfillment_channel — 配送方式", "", ""),
    ("", "", "", ""),
    ("四、注意事项", "", "", ""),
    ("", "1. 品牌名必须填你自己注册的品牌，不能用竞品品牌", "", ""),
    ("", "2. 图片URL要完整(https://开头)，从竞品listing右键复制图片地址", "", ""),
    ("", "3. 价格填美元，不要加$符号", "", ""),
    ("", "4. 同一Excel可以填多个商品，每行一个", "", ""),
    ("", "5. 有变体商品(多色/多码)需要填变体列", "", ""),
    ("", "6. UPC/EAN如果没有可以留空，但建议购买条码", "", ""),
]

for row_idx, (a, b, c, d) in enumerate(instructions, 1):
    ws2.cell(row=row_idx, column=1, value=a)
    ws2.cell(row=row_idx, column=2, value=b)
    ws2.cell(row=row_idx, column=3, value=c)

# 标题样式
ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
for r in [3, 11, 22, 35]:
    if r <= len(instructions):
        ws2.cell(row=r, column=1).font = Font(bold=True, size=12)

ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 50

# ============================================================
# Sheet 3: 产品类型参考
# ============================================================
ws3 = wb.create_sheet(title="产品类型参考")

ws3.cell(row=1, column=1, value="Product Type").font = Font(bold=True, size=11)
ws3.cell(row=1, column=2, value="中文名").font = Font(bold=True, size=11)
ws3.cell(row=1, column=3, value="适用商品举例").font = Font(bold=True, size=11)

product_types = [
    ("WIRELESS_ACCESSORY", "无线配件", "蓝牙耳机、无线充电器、手机壳"),
    ("HEADPHONES", "耳机", "头戴式耳机、入耳式耳机"),
    ("PHONE_ACCESSORY", "手机配件", "手机壳、钢化膜、支架"),
    ("PORTABLE_ELECTRONIC", "便携电子", "充电宝、蓝牙音箱"),
    ("LUGGAGE", "箱包", "行李箱、旅行袋"),
    ("BACKPACK", "背包", "双肩包、书包、登山包"),
    ("HANDBAG", "手提包", "女士手提包、钱包"),
    ("SHIRT", "上衣", "T恤、衬衫、polo衫"),
    ("PANTS", "裤子", "长裤、短裤、运动裤"),
    ("DRESS", "连衣裙", "各类连衣裙"),
    ("SHOES", "鞋", "运动鞋、休闲鞋、凉鞋"),
    ("WATCH", "手表", "智能手表、机械表、石英表"),
    ("JEWELRY", "珠宝", "项链、手链、戒指、耳环"),
    ("HOME_BED_AND_BATH", "家居床浴", "床单、毛巾、浴帘"),
    ("HOME_FURNITURE", "家具", "桌子、椅子、书架"),
    ("KITCHEN", "厨房用品", "锅具、刀具、收��"),
    ("DRINKING_CUP", "水杯", "保温杯、咖啡杯、马克杯"),
    ("BOTTLE", "水壶", "运动水壶、保温壶"),
    ("HOME_LIGHTING", "灯具", "台灯、落地灯、LED灯"),
    ("SPORTING_GOODS", "运动用品", "瑜伽垫、跳绳、哑铃"),
    ("FITNESS", "健身", "健身器材、阻力带"),
    ("OUTDOOR_RECREATION", "户外", "帐篷、睡袋、野营"),
    ("TOY", "玩具", "积木、模型、益智玩具"),
    ("GAME", "游戏", "桌游、卡牌、电子游戏配件"),
    ("BEAUTY", "美妆", "化妆品、美容工具"),
    ("SKINCARE", "护肤", "面膜、精华、防晒"),
    ("PET_SUPPLIES", "宠物用品", "狗粮、猫玩具、宠物床"),
    ("BABY_PRODUCT", "母婴", "奶瓶、婴儿车、尿布"),
    ("AUTO_ACCESSORY", "汽车配件", "车载充电器、手机支架"),
    ("OFFICE_PRODUCTS", "办公用品", "文具、收纳、打印耗材"),
    ("GIFT_SET", "礼品套装", "礼盒、组合礼品"),
    ("TOOLS", "工具", "电动工具、手动工具"),
    ("PRODUCT", "通用", "无法分类时使用PRODUCT"),
]

for row_idx, (pt, cn, example) in enumerate(product_types, 2):
    ws3.cell(row=row_idx, column=1, value=pt)
    ws3.cell(row=row_idx, column=2, value=cn)
    ws3.cell(row=row_idx, column=3, value=example)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 40

# ============================================================
# Sheet 4: Marketplace站点参考
# ============================================================
ws4 = wb.create_sheet(title="站点参考")

ws4.cell(row=1, column=1, value="站点").font = Font(bold=True)
ws4.cell(row=1, column=2, value="Marketplace ID").font = Font(bold=True)
ws4.cell(row=1, column=3, value="API Endpoint").font = Font(bold=True)

sites = [
    ("🇺🇸 美国 (默认)", "ATVPDKIKX0DER", "sellingpartnerapi-na.amazon.com"),
    ("🇨🇦 加拿大", "A2EUQ1WTGCTBG2", "sellingpartnerapi-na.amazon.com"),
    ("🇲🇽 墨西哥", "A1AM78C64UM0Y8", "sellingpartnerapi-na.amazon.com"),
    ("🇬🇧 英国", "A1F83G8C2ARO7P", "sellingpartnerapi-eu.amazon.com"),
    ("🇩🇪 德国", "A1PA6795UKMFR9", "sellingpartnerapi-eu.amazon.com"),
    ("🇫🇷 法国", "A13V1IB3VIYZZH", "sellingpartnerapi-eu.amazon.com"),
    ("🇮🇹 意大利", "APJ6JRA9NG5V4", "sellingpartnerapi-eu.amazon.com"),
    ("🇪🇸 西班牙", "A1RKKUPIHCS9HS", "sellingpartnerapi-eu.amazon.com"),
    ("🇯🇵 日本", "A1VC38T7YXB528", "sellingpartnerapi-fe.amazon.com"),
    ("🇦🇺 澳大利亚", "A39IBJ37TRP1C6", "sellingpartnerapi-fe.amazon.com"),
]

for row_idx, (site, mp_id, endpoint) in enumerate(sites, 2):
    ws4.cell(row=row_idx, column=1, value=site)
    ws4.cell(row=row_idx, column=2, value=mp_id)
    ws4.cell(row=row_idx, column=3, value=endpoint)

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 40

# ============================================================
# 保存
# ============================================================
output_dir = r'C:\Users\zhouk\Desktop\亚马逊2.8'
output_path = os.path.join(output_dir, '亚马逊商品采集模板_v1.0.xlsx')
wb.save(output_path)
print(f"✅ 模板已生成: {output_path}")
print(f"   Sheet1: 商品采集模板 ({len(COLUMNS)} 列)")
print("   Sheet2: 填写说明")
print(f"   Sheet3: 产品类型参考 ({len(product_types)} 个类型)")
print(f"   Sheet4: 站点参考 ({len(sites)} 个站点)")
