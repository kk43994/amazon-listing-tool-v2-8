"""
补齐 4 个零覆盖模块的测试 + ai_text system prompt 测试。
"""
import base64
import io
import json
import os
import tempfile
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from PIL import Image


# ===== 1. core/template_generator.py =====

from core.template_generator import build_template_definition, generate_template, CURRENT_TEMPLATE_VERSION


def _make_schema_fields():
    return {
        'required_fields': [
            {'name': 'item_name', 'type': 'string', 'title': 'Product Title', 'description': 'Title', 'group': 'product_identity'},
            {'name': 'brand', 'type': 'string', 'title': 'Brand', 'description': '', 'group': 'product_identity'},
        ],
        'optional_fields': [
            {'name': 'color', 'type': 'string', 'title': 'Color', 'description': '', 'group': 'product_description'},
            {'name': 'material', 'type': 'string', 'title': 'Material', 'description': '', 'group': 'product_description'},
        ],
        'enum_fields': {
            'color': ['Red', 'Blue', 'Green'],
        },
        'field_groups': {
            'product_identity': {'title': 'Product Identity', 'description': '', 'propertyNames': ['item_name', 'brand']},
            'product_description': {'title': 'Details', 'description': '', 'propertyNames': ['color', 'material']},
        },
    }


def test_build_template_definition_basic():
    defn = build_template_definition(_make_schema_fields(), 'WIRELESS_ACCESSORY')
    assert defn['product_type'] == 'WIRELESS_ACCESSORY'
    assert defn['marketplace'] == 'US'
    assert defn['variation_mode'] == 'single'
    assert defn['template_version'] == CURRENT_TEMPLATE_VERSION
    cols = defn['columns']
    assert isinstance(cols, list) and len(cols) > 0
    keys = [c['key'] for c in cols]
    assert 'sku' in keys
    assert 'product_type' in keys


def test_build_template_definition_includes_required_fields():
    defn = build_template_definition(_make_schema_fields(), 'TEST_TYPE')
    cols = {c['key']: c for c in defn['columns']}
    assert 'item_name' in cols or any('item_name' in c['key'] for c in defn['columns'])


def test_build_template_definition_variation_mode():
    defn = build_template_definition(_make_schema_fields(), 'TEST', variation_mode='variation')
    assert defn['variation_mode'] == 'variation'
    keys = [c['key'] for c in defn['columns']]
    assert 'parentage_level' in keys


def test_build_template_definition_enum_values():
    defn = build_template_definition(_make_schema_fields(), 'TEST')
    cols = {c['key']: c for c in defn['columns']}
    if 'color' in cols:
        assert 'Red' in cols['color'].get('enum_values', [])


def test_generate_template_creates_xlsx():
    defn = build_template_definition(_make_schema_fields(), 'TEST_TYPE')
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        output_path = f.name
    try:
        generate_template(defn, output_path)
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 1000
    finally:
        os.unlink(output_path)


def test_generate_template_has_header_rows():
    from openpyxl import load_workbook
    defn = build_template_definition(_make_schema_fields(), 'TEST_TYPE')
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        output_path = f.name
    try:
        generate_template(defn, output_path)
        wb = load_workbook(output_path)
        ws = wb.active
        # Row 1 = 中文说明, Row 2 = 英文字段名
        assert ws.cell(row=2, column=1).value is not None
        wb.close()
    finally:
        os.unlink(output_path)


# ===== 2. core/image_resolver.py =====

from core.image_resolver import resolve_image_from_url, ImageResolveResult


def test_resolve_empty_url():
    result = resolve_image_from_url('', '/tmp')
    assert not result.success
    assert '空' in result.error


def test_resolve_base64_url():
    img = Image.new('RGB', (10, 10), (255, 0, 0))
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    b64 = base64.b64encode(buf.getvalue()).decode()
    data_url = f'data:image/png;base64,{b64}'

    result = resolve_image_from_url(data_url, '/tmp')
    assert result.success
    assert result.source_type == 'base64'
    assert result.image.size == (10, 10)


def test_resolve_local_file():
    with tempfile.TemporaryDirectory() as tmpdir:
        materials_dir = os.path.join(tmpdir, 'materials')
        os.makedirs(materials_dir)
        img_path = os.path.join(materials_dir, 'test.png')
        Image.new('RGB', (50, 50), (0, 0, 255)).save(img_path)

        result = resolve_image_from_url('/files/materials/test.png', tmpdir)
        assert result.success
        assert result.source_type == 'local'


def test_resolve_local_file_path_traversal():
    with tempfile.TemporaryDirectory() as tmpdir:
        result = resolve_image_from_url('/files/materials/../../etc/passwd', tmpdir)
        assert not result.success
        assert '非法' in (result.error or '') or '不存在' in (result.error or '')


def test_resolve_unrecognized_url():
    result = resolve_image_from_url('ftp://example.com/img.png', '/tmp')
    assert not result.success
    assert '无法识别' in result.error


# ===== 3. amazon/schema_manager.py =====

from amazon.schema_manager import parse_schema, _extract_type, _extract_enum


def test_parse_schema_basic():
    schema_data = {
        'schema': {
            'properties': {
                'item_name': {'type': 'array', 'items': {'type': 'object', 'properties': {'value': {'type': 'string'}}}},
                'brand': {'type': 'array', 'items': {'type': 'object', 'properties': {'value': {'type': 'string'}}}},
                'color': {'type': 'array', 'items': {'type': 'object', 'properties': {'value': {'type': 'string', 'enum': ['Red', 'Blue']}}}},
            },
            'required': ['item_name', 'brand'],
        },
        'property_groups': {
            'identity': {'title': 'Identity', 'description': '', 'propertyNames': ['item_name', 'brand']},
            'details': {'title': 'Details', 'description': '', 'propertyNames': ['color']},
        },
    }
    result = parse_schema(schema_data)
    assert len(result['required_fields']) == 2
    assert len(result['optional_fields']) == 1
    assert 'color' in result['enum_fields']
    assert result['enum_fields']['color'] == ['Red', 'Blue']
    assert 'identity' in result['field_groups']


def test_parse_schema_empty():
    result = parse_schema({})
    assert result['required_fields'] == []
    assert result['optional_fields'] == []


def test_extract_type_direct():
    assert _extract_type({'type': 'integer'}) == 'integer'


def test_extract_type_nested():
    assert _extract_type({'properties': {'value': {'type': 'boolean'}}}) == 'boolean'


def test_extract_type_default():
    assert _extract_type({}) == 'string'


def test_extract_enum_direct():
    assert _extract_enum({'enum': ['a', 'b']}) == ['a', 'b']


def test_extract_enum_nested():
    assert _extract_enum({'properties': {'value': {'enum': ['x', 'y']}}}) == ['x', 'y']


def test_extract_enum_none():
    assert _extract_enum({}) is None


# ===== 4. core/ai_client.py — 主函数 mock 测试 =====

from core.ai_client import ai_text, ai_image_edit, ai_image_generate, AMAZON_SYSTEM_PROMPT


def test_ai_text_openai_sends_system_prompt(monkeypatch):
    """验证 ai_text 通过 OpenAI 协议时发送 system prompt。"""
    monkeypatch.setenv('AI_API_KEY', 'test-key')
    monkeypatch.setenv('AI_TEXT_PROTOCOL', 'openai_chat_completions')
    from config import reload_config
    reload_config()

    captured_messages = []

    class FakeChoice:
        def __init__(self):
            self.message = SimpleNamespace(content='Generated Title')

    class FakeResponse:
        choices = [FakeChoice()]

    class FakeCompletions:
        def create(self, **kwargs):
            captured_messages.extend(kwargs.get('messages', []))
            return FakeResponse()

    class FakeChat:
        completions = FakeCompletions()

    class FakeClient:
        chat = FakeChat()

    monkeypatch.setattr('core.ai_client._build_openai_client', lambda *a, **kw: FakeClient())

    result = ai_text('Write a title', temperature=0.5)
    assert result == 'Generated Title'
    assert len(captured_messages) == 2
    assert captured_messages[0]['role'] == 'system'
    assert 'Amazon' in captured_messages[0]['content']
    assert captured_messages[1]['role'] == 'user'
    assert captured_messages[1]['content'] == 'Write a title'


def test_ai_text_custom_system_prompt(monkeypatch):
    """验证自定义 system_prompt 覆盖默认值。"""
    monkeypatch.setenv('AI_API_KEY', 'test-key')
    monkeypatch.setenv('AI_TEXT_PROTOCOL', 'openai_chat_completions')
    from config import reload_config
    reload_config()

    captured = []

    class FakeChoice:
        def __init__(self):
            self.message = SimpleNamespace(content='OK')

    class FakeResponse:
        choices = [FakeChoice()]

    class FakeCompletions:
        def create(self, **kwargs):
            captured.extend(kwargs.get('messages', []))
            return FakeResponse()

    class FakeChat:
        completions = FakeCompletions()

    class FakeClient:
        chat = FakeChat()

    monkeypatch.setattr('core.ai_client._build_openai_client', lambda *a, **kw: FakeClient())

    ai_text('hello', system_prompt='You are a helpful assistant.')
    assert captured[0]['content'] == 'You are a helpful assistant.'


def test_ai_text_gemini_includes_system(monkeypatch):
    """验证 Gemini 协议时 system prompt 被拼入 contents。"""
    monkeypatch.setenv('AI_API_KEY', 'test-key')
    monkeypatch.setenv('AI_TEXT_PROTOCOL', 'gemini_generate_content')
    monkeypatch.setenv('AI_TEXT_ENDPOINT_TEMPLATE', '/v1/models/{model}:generateContent')
    from config import reload_config
    reload_config()

    captured_payload = {}

    def fake_gemini(**kwargs):
        captured_payload.update(kwargs)
        return {'candidates': [{'content': {'parts': [{'text': 'Result'}]}}]}

    monkeypatch.setattr('core.ai_client._gemini_generate_content', fake_gemini)

    result = ai_text('Write bullets')
    assert result == 'Result'
    prompt_text = captured_payload['contents'][0]['parts'][0]['text']
    assert 'Amazon' in prompt_text
    assert 'Write bullets' in prompt_text


def test_ai_text_returns_empty_on_error(monkeypatch):
    monkeypatch.setenv('AI_API_KEY', 'test-key')
    monkeypatch.setenv('AI_TEXT_PROTOCOL', 'openai_chat_completions')
    from config import reload_config
    reload_config()

    monkeypatch.setattr('core.ai_client._build_openai_client', lambda *a, **kw: (_ for _ in ()).throw(Exception('network error')))

    result = ai_text('test', raise_on_error=False)
    assert result == ''


def test_ai_text_raises_on_error(monkeypatch):
    monkeypatch.setenv('AI_API_KEY', 'test-key')
    monkeypatch.setenv('AI_TEXT_PROTOCOL', 'openai_chat_completions')
    from config import reload_config
    reload_config()

    def bad_client(*a, **kw):
        raise ConnectionError('timeout')

    monkeypatch.setattr('core.ai_client._build_openai_client', bad_client)

    with pytest.raises(ConnectionError):
        ai_text('test', raise_on_error=True)


def test_ai_image_edit_returns_none_on_error(monkeypatch):
    monkeypatch.setenv('AI_API_KEY', 'test-key')
    monkeypatch.setenv('AI_IMAGE_PROTOCOL', 'openai_images')
    from config import reload_config
    reload_config()

    import httpx
    def fake_post(*a, **kw):
        raise httpx.ConnectError('fail')

    monkeypatch.setattr('httpx.Client.post', fake_post)

    result = ai_image_edit(b'\x89PNG fake', 'white bg')
    assert result is None


def test_ai_image_generate_returns_none_on_error(monkeypatch):
    monkeypatch.setenv('AI_API_KEY', 'test-key')
    monkeypatch.setenv('AI_IMAGE_PROTOCOL', 'openai_images')
    from config import reload_config
    reload_config()

    import httpx
    def fake_post(*a, **kw):
        raise httpx.ConnectError('fail')

    monkeypatch.setattr('httpx.Client.post', fake_post)

    result = ai_image_generate('a white square')
    assert result is None
