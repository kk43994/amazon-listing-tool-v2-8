"""
生成亚马逊美国站上架完整测试Excel
包含所有必填字段和重要选填字段
基于 Amazon Seller Central Flat File / SP-API Listings 要求
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_full_sample():
    wb = Workbook()
    
    # ===== Sheet 1: 商品数据 (主表) =====
    ws = wb.active
    ws.title = "商品数据"

    # Amazon US Marketplace 上架字段
    # 分组: 必填 | 重要选填 | 图片 | 变体
    headers = [
        # === 基础信息 (必填) ===
        'SKU',                          # 卖家自定义SKU
        'product_id',                   # ASIN/UPC/EAN/ISBN
        'product_id_type',              # 产品ID类型: ASIN/UPC/EAN
        'item_name',                    # 商品标题 (200字符以内)
        'brand_name',                   # 品牌名
        'manufacturer',                 # 制造商
        'part_number',                  # 型号编号
        'product_type',                 # 产品类型 (如 HOME, ELECTRONICS)
        
        # === 分类 ===
        'recommended_browse_nodes',     # 推荐分类节点ID
        'item_type_keyword',            # 物品类型关键词
        
        # === 描述 (重要) ===
        'product_description',          # 商品描述 (2000字符)
        'bullet_point_1',               # 五点描述1
        'bullet_point_2',               # 五点描述2
        'bullet_point_3',               # 五点描述3
        'bullet_point_4',               # 五点描述4
        'bullet_point_5',               # 五点描述5
        'generic_keywords',             # 后台搜索关键词 (250字节)
        
        # === 价格与库存 ===
        'standard_price',               # 标准售价 (USD)
        'quantity',                     # 库存数量
        'fulfillment_channel',          # 配送渠道: DEFAULT(自配)/AMAZON_NA(FBA)
        'condition_type',               # 商品状况: New/Refurbished/Used
        
        # === 图片 ===
        'main_image_url',               # 主图URL (必须白底, 1000x1000px+)
        'other_image_url_1',            # 副图1
        'other_image_url_2',            # 副图2
        'other_image_url_3',            # 副图3
        'other_image_url_4',            # 副图4
        'other_image_url_5',            # 副图5
        'swatch_image_url',             # 色板图
        
        # === 物流尺寸 ===
        'item_weight',                  # 商品重量 (磅)
        'item_weight_unit_of_measure',  # 重量单位
        'item_length',                  # 长度
        'item_width',                   # 宽度
        'item_height',                  # 高度
        'item_dimensions_unit_of_measure',  # 尺寸单位
        
        # === 包装信息 ===
        'package_weight',               # 包装重量
        'package_length',               # 包装长度
        'package_width',                # 包装宽度
        'package_height',               # 包装高度
        'number_of_items',              # 件数
        
        # === 变体 ===
        'parent_child',                 # 父/子: Parent/Child
        'parent_sku',                   # 父SKU (子商品填写)
        'relationship_type',            # 关系类型: Variation
        'variation_theme',              # 变体主题: Size/Color/SizeColor
        'color_name',                   # 颜色名称
        'size_name',                    # 尺寸名称
        'material_type',                # 材质
        
        # === 合规 ===
        'country_of_origin',            # 原产国
        'batteries_required',           # 是否需要电池
        'are_batteries_included',       # 是否含电池
        'cpsia_cautionary_statement',   # CPSIA安全声明
        
        # === 其他 ===
        'target_audience_keywords',     # 目标受众
        'style_name',                   # 款式名称
        'department_name',              # 部门 (男/女/男女通用)
        'update_delete',                # 操作: Update/PartialUpdate/Delete
    ]

    # 颜色分组
    group_colors = {
        '基础必填': 'FF4472C4',     # 蓝色
        '分类': 'FF7030A0',         # 紫色
        '描述文案': 'FF00B050',     # 绿色
        '价格库存': 'FFF4B084',     # 橙色
        '图片': 'FFFF6B6B',         # 红色
        '物流尺寸': 'FF92D050',     # 浅绿
        '包装': 'FF00B0F0',         # 天蓝
        '变体': 'FFFFC000',         # 金色
        '合规': 'FFBFBFBF',         # 灰色
        '其他': 'FFD9E2F3',         # 浅蓝
    }

    field_groups = {
        'SKU': '基础必填', 'product_id': '基础必填', 'product_id_type': '基础必填',
        'item_name': '基础必填', 'brand_name': '基础必填', 'manufacturer': '基础必填',
        'part_number': '基础必填', 'product_type': '基础必填',
        'recommended_browse_nodes': '分类', 'item_type_keyword': '分类',
        'product_description': '描述文案', 'bullet_point_1': '描述文案',
        'bullet_point_2': '描述文案', 'bullet_point_3': '描述文案',
        'bullet_point_4': '描述文案', 'bullet_point_5': '描述文案',
        'generic_keywords': '描述文案',
        'standard_price': '价格库存', 'quantity': '价格库存',
        'fulfillment_channel': '价格库存', 'condition_type': '价格库存',
        'main_image_url': '图片', 'other_image_url_1': '图片',
        'other_image_url_2': '图片', 'other_image_url_3': '图片',
        'other_image_url_4': '图片', 'other_image_url_5': '图片',
        'swatch_image_url': '图片',
        'item_weight': '物流尺寸', 'item_weight_unit_of_measure': '物流尺寸',
        'item_length': '物流尺寸', 'item_width': '物流尺寸',
        'item_height': '物流尺寸', 'item_dimensions_unit_of_measure': '物流尺寸',
        'package_weight': '包装', 'package_length': '包装',
        'package_width': '包装', 'package_height': '包装',
        'number_of_items': '包装',
        'parent_child': '变体', 'parent_sku': '变体',
        'relationship_type': '变体', 'variation_theme': '变体',
        'color_name': '变体', 'size_name': '变体', 'material_type': '变体',
        'country_of_origin': '合规', 'batteries_required': '合规',
        'are_batteries_included': '合规', 'cpsia_cautionary_statement': '合规',
        'target_audience_keywords': '其他', 'style_name': '其他',
        'department_name': '其他', 'update_delete': '其他',
    }

    # 写表头（第1行：分组名，第2行：字段名）
    # Row 1: Group headers
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        group = field_groups.get(header, '其他')
        color = group_colors.get(group, 'FFD9E2F3')

        # Row 1: Group
        cell1 = ws.cell(row=1, column=col_idx, value=group)
        cell1.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell1.font = Font(bold=True, size=10, color='FFFFFF' if color not in ['FFFFC000', 'FFF4B084', 'FF92D050', 'FFD9E2F3'] else '000000')
        cell1.alignment = Alignment(horizontal='center', vertical='center')
        cell1.border = thin_border

        # Row 2: Field name
        cell2 = ws.cell(row=2, column=col_idx, value=header)
        cell2.fill = PatternFill(start_color='FF2D3748', end_color='FF2D3748', fill_type='solid')
        cell2.font = Font(bold=True, size=9, color='FFFFFF')
        cell2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell2.border = thin_border

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 35

    # 测试数据 - 5个真实风格的美国站商品
    test_products = [
        {
            'SKU': 'WBE-001-BLK',
            'product_id': '',
            'product_id_type': 'UPC',
            'item_name': 'ProSound X5 Wireless Bluetooth 5.3 Earbuds - Active Noise Cancelling, 40H Playtime, IPX7 Waterproof, Touch Control, USB-C Fast Charging, Deep Bass for iPhone Android (Black)',
            'brand_name': 'ProSound',
            'manufacturer': 'ProSound Technology Co., Ltd.',
            'part_number': 'PS-X5-BLK',
            'product_type': 'HEADPHONES',
            'recommended_browse_nodes': '172541',
            'item_type_keyword': 'in-ear-headphones',
            'product_description': '<b>ProSound X5 — Your Perfect Audio Companion</b><br><br>Experience crystal-clear sound with our latest Bluetooth 5.3 wireless earbuds. Featuring advanced Active Noise Cancelling technology, the X5 blocks out 95% of ambient noise so you can focus on your music, calls, or podcasts.<br><br><b>Key Features:</b><ul><li>40 hours total playtime (8h earbuds + 32h charging case)</li><li>IPX7 waterproof rating for gym & outdoor use</li><li>Ergonomic design with 3 sizes of ear tips for secure fit</li><li>Touch control for music, calls, and voice assistant</li><li>USB-C fast charging — 10 min charge = 2 hours playtime</li></ul>',
            'bullet_point_1': 'CRYSTAL CLEAR SOUND — Advanced 13mm dynamic drivers deliver rich bass, clear mids, and crisp highs for an immersive listening experience',
            'bullet_point_2': 'ACTIVE NOISE CANCELLING — Block out 95% of background noise with our ANC technology. Switch to Transparency mode to stay aware of your surroundings',
            'bullet_point_3': '40 HOURS TOTAL PLAYTIME — 8 hours on a single charge plus 32 hours from the compact charging case. USB-C fast charging gives 2 hours of playback from just 10 minutes',
            'bullet_point_4': 'IPX7 WATERPROOF — Sweat-proof and rain-proof design perfect for workouts, running, and outdoor activities',
            'bullet_point_5': 'SEAMLESS CONNECTIVITY — Bluetooth 5.3 ensures stable connection up to 50ft. One-step pairing with iPhone, Samsung, and all Android devices',
            'generic_keywords': 'wireless earbuds noise cancelling bluetooth headphones waterproof gym workout earphones usb-c charging deep bass',
            'standard_price': '29.99',
            'quantity': '500',
            'fulfillment_channel': 'AMAZON_NA',
            'condition_type': 'New',
            'main_image_url': 'https://m.media-amazon.com/images/I/7159iuhZwPL.jpg',
            'other_image_url_1': 'https://m.media-amazon.com/images/I/717FY9S8wFL.jpg',
            'other_image_url_2': 'https://m.media-amazon.com/images/I/71AvOPRUOCL.jpg',
            'other_image_url_3': 'https://m.media-amazon.com/images/I/71sWRy5QxIL.jpg',
            'other_image_url_4': 'https://m.media-amazon.com/images/I/71vE9217imL.jpg',
            'other_image_url_5': '',
            'swatch_image_url': 'https://m.media-amazon.com/images/I/51IhSrucIyL._SS200_.png',
            'item_weight': '0.12',
            'item_weight_unit_of_measure': 'pounds',
            'item_length': '2.5',
            'item_width': '2.0',
            'item_height': '1.2',
            'item_dimensions_unit_of_measure': 'inches',
            'package_weight': '0.35',
            'package_length': '6.5',
            'package_width': '4.5',
            'package_height': '2.8',
            'number_of_items': '1',
            'parent_child': 'Child',
            'parent_sku': 'WBE-001',
            'relationship_type': 'Variation',
            'variation_theme': 'Color',
            'color_name': 'Black',
            'size_name': '',
            'material_type': 'Plastic, Silicone',
            'country_of_origin': 'CN',
            'batteries_required': 'No',
            'are_batteries_included': 'Yes',
            'cpsia_cautionary_statement': '',
            'target_audience_keywords': 'Unisex Adults',
            'style_name': 'X5 ANC',
            'department_name': 'Unisex',
            'update_delete': 'Update',
        },
        {
            'SKU': 'SSB-002-32OZ',
            'product_id': '',
            'product_id_type': 'UPC',
            'item_name': 'HydroPeak 32oz Insulated Stainless Steel Water Bottle with Straw Lid - Vacuum Double Wall, Keeps Cold 24H Hot 12H, BPA Free, Leak Proof Sports Bottle (Ocean Blue)',
            'brand_name': 'HydroPeak',
            'manufacturer': 'HydroPeak Outdoor LLC',
            'part_number': 'HP-32OZ-OCEAN',
            'product_type': 'OUTDOOR_RECREATION_PRODUCT',
            'recommended_browse_nodes': '17416544011',
            'item_type_keyword': 'sports-water-bottles',
            'product_description': '<b>Stay Hydrated, Stay Active</b><br><br>The HydroPeak 32oz insulated water bottle is engineered for athletes, hikers, and anyone who demands the best. Our premium 18/8 stainless steel construction with vacuum double-wall insulation keeps your drinks ice cold for 24 hours or steaming hot for 12 hours.<br><br>The innovative straw lid design allows one-handed drinking while on the move. Wide mouth opening fits ice cubes easily and makes cleaning a breeze.',
            'bullet_point_1': 'TEMPERATURE CONTROL — Vacuum double-wall insulation keeps drinks cold 24 hours, hot 12 hours. No condensation, no sweating on the outside',
            'bullet_point_2': 'PREMIUM MATERIALS — 18/8 food-grade stainless steel, BPA-free, no metallic taste. Powder-coated exterior resists scratches and provides firm grip',
            'bullet_point_3': 'CONVENIENT STRAW LID — One-handed drinking with flip-up straw. Leak-proof seal tested to withstand drops. Includes wide-mouth lid for ice cubes',
            'bullet_point_4': 'PERFECT SIZE — 32oz capacity fits standard cup holders. Lightweight at just 14oz empty. Ideal for gym, hiking, office, school, and travel',
            'bullet_point_5': 'LIFETIME WARRANTY — We stand behind our products. If anything goes wrong, we will replace it free of charge. 100% satisfaction guaranteed',
            'generic_keywords': 'water bottle insulated stainless steel vacuum cold hot straw lid sports gym hiking outdoor bpa free leak proof',
            'standard_price': '24.99',
            'quantity': '800',
            'fulfillment_channel': 'AMAZON_NA',
            'condition_type': 'New',
            'main_image_url': 'https://m.media-amazon.com/images/I/71GKYfVzYwL.jpg',
            'other_image_url_1': 'https://m.media-amazon.com/images/I/61XntLH2loL.jpg',
            'other_image_url_2': 'https://m.media-amazon.com/images/I/61bemgdSCOL.jpg',
            'other_image_url_3': 'https://m.media-amazon.com/images/I/81wH1Y0rrKL.jpg',
            'other_image_url_4': '',
            'other_image_url_5': '',
            'swatch_image_url': 'https://m.media-amazon.com/images/I/71aXuAwSncL.jpg',
            'item_weight': '0.88',
            'item_weight_unit_of_measure': 'pounds',
            'item_length': '10.5',
            'item_width': '3.5',
            'item_height': '3.5',
            'item_dimensions_unit_of_measure': 'inches',
            'package_weight': '1.2',
            'package_length': '12.0',
            'package_width': '4.5',
            'package_height': '4.5',
            'number_of_items': '1',
            'parent_child': 'Child',
            'parent_sku': 'SSB-002',
            'relationship_type': 'Variation',
            'variation_theme': 'Color',
            'color_name': 'Ocean Blue',
            'size_name': '32 oz',
            'material_type': 'Stainless Steel',
            'country_of_origin': 'CN',
            'batteries_required': 'No',
            'are_batteries_included': 'No',
            'cpsia_cautionary_statement': '',
            'target_audience_keywords': 'Unisex Adults',
            'style_name': 'Straw Lid',
            'department_name': 'Unisex',
            'update_delete': 'Update',
        },
        {
            'SKU': 'LED-003-WHT',
            'product_id': '',
            'product_id_type': 'UPC',
            'item_name': 'BrightDesk Pro LED Desk Lamp with Wireless Charger - 5 Color Modes, 7 Brightness Levels, USB-A Port, 45min Auto Timer, Eye-Caring, Foldable Architect Lamp for Home Office (White)',
            'brand_name': 'BrightDesk',
            'manufacturer': 'BrightDesk Home Solutions',
            'part_number': 'BD-PRO-WHT',
            'product_type': 'LIGHTING',
            'recommended_browse_nodes': '1063292',
            'item_type_keyword': 'desk-lamps',
            'product_description': '<b>Light Your Workspace, Charge Your Devices</b><br><br>The BrightDesk Pro combines premium LED lighting with a built-in 10W wireless charger. 5 color temperature modes (2700K-6500K) and 7 brightness levels give you 35 lighting combinations for any task — reading, studying, video calls, or relaxing.<br><br>The eye-caring design eliminates flicker and reduces blue light to protect your vision during long work sessions.',
            'bullet_point_1': 'WIRELESS CHARGING BUILT-IN — 10W Qi wireless charging pad in the base. Just place your phone on it. Also includes USB-A port for wired charging',
            'bullet_point_2': '35 LIGHTING OPTIONS — 5 color temperatures (2700K warm to 6500K daylight) × 7 brightness levels. Memory function remembers your last setting',
            'bullet_point_3': 'EYE-CARING TECHNOLOGY — No flicker, no ghosting, no blue light hazard. Soft, uniform light reduces eye strain during long study and work sessions',
            'bullet_point_4': 'FLEXIBLE & SPACE-SAVING — Multi-joint adjustable arm rotates 180°. Foldable design stores flat when not in use. Weighted base prevents tipping',
            'bullet_point_5': 'SMART FEATURES — 45-minute auto-off timer for bedtime reading. Touch controls on the base. Energy-efficient LEDs last 50,000 hours',
            'generic_keywords': 'desk lamp led wireless charger office lamp adjustable brightness eye care reading light usb charging foldable architect',
            'standard_price': '39.99',
            'quantity': '300',
            'fulfillment_channel': 'AMAZON_NA',
            'condition_type': 'New',
            'main_image_url': 'https://m.media-amazon.com/images/I/71myfUGR+fL.jpg',
            'other_image_url_1': 'https://m.media-amazon.com/images/I/71Lg-bVp8JL.jpg',
            'other_image_url_2': 'https://m.media-amazon.com/images/I/61ORSZg7SaL.jpg',
            'other_image_url_3': 'https://m.media-amazon.com/images/I/515LLATX9gL.jpg',
            'other_image_url_4': '',
            'other_image_url_5': '',
            'swatch_image_url': '',
            'item_weight': '1.8',
            'item_weight_unit_of_measure': 'pounds',
            'item_length': '18.0',
            'item_width': '6.5',
            'item_height': '2.5',
            'item_dimensions_unit_of_measure': 'inches',
            'package_weight': '2.5',
            'package_length': '20.0',
            'package_width': '8.0',
            'package_height': '4.0',
            'number_of_items': '1',
            'parent_child': '',
            'parent_sku': '',
            'relationship_type': '',
            'variation_theme': '',
            'color_name': 'White',
            'size_name': '',
            'material_type': 'Aluminum, ABS Plastic',
            'country_of_origin': 'CN',
            'batteries_required': 'No',
            'are_batteries_included': 'No',
            'cpsia_cautionary_statement': '',
            'target_audience_keywords': 'Unisex Adults',
            'style_name': 'Pro',
            'department_name': 'Unisex',
            'update_delete': 'Update',
        },
    ]

    # 写数据
    for row_idx, product in enumerate(test_products, 3):
        for col_idx, header in enumerate(headers, 1):
            value = product.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 设置列宽
    col_widths = {
        'SKU': 16, 'product_id': 14, 'product_id_type': 12,
        'item_name': 50, 'brand_name': 14, 'manufacturer': 22,
        'product_description': 40,
        'bullet_point_1': 40, 'bullet_point_2': 40, 'bullet_point_3': 40,
        'bullet_point_4': 40, 'bullet_point_5': 40,
        'generic_keywords': 35, 'main_image_url': 30,
    }
    for col_idx, header in enumerate(headers, 1):
        width = col_widths.get(header, 14)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    # 冻结前2行+前2列
    ws.freeze_panes = 'C3'

    # ===== Sheet 2: 字段说明 =====
    ws2 = wb.create_sheet("字段说明")
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 30

    field_docs = [
        ['字段名', '必填/选填', '说明', '示例值'],
        ['SKU', '必填', '卖家自定义的唯一商品编号', 'WBE-001-BLK'],
        ['product_id', '必填*', '产品标识 (UPC/EAN/ASIN)，新品可留空', '012345678901'],
        ['product_id_type', '必填*', '产品ID类型', 'UPC / EAN / ASIN'],
        ['item_name', '必填', '商品标题，200字符以内，含品牌+关键特征', 'ProSound X5 Wireless Bluetooth...'],
        ['brand_name', '必填', '品牌名称（需在亚马逊品牌注册）', 'ProSound'],
        ['manufacturer', '必填', '制造商名称', 'ProSound Technology Co., Ltd.'],
        ['product_type', '必填', '产品类型代码', 'HEADPHONES / LIGHTING / HOME'],
        ['product_description', '重要', '商品详细描述，支持HTML，2000字符', '<b>标题</b><br>描述内容...'],
        ['bullet_point_1~5', '重要', '五点描述，每条150-250字符，大写关键词开头', 'CRYSTAL CLEAR SOUND — ...'],
        ['generic_keywords', '重要', '后台搜索关键词，250字节，空格分隔', 'wireless earbuds bluetooth...'],
        ['standard_price', '必填', '售价(USD)', '29.99'],
        ['quantity', '必填', '库存数量', '500'],
        ['fulfillment_channel', '必填', '配送方式: DEFAULT=自配, AMAZON_NA=FBA', 'AMAZON_NA'],
        ['condition_type', '必填', '商品状况', 'New'],
        ['main_image_url', '必填', '主图URL，白底，≥1000x1000px，JPEG/PNG', 'https://...main.jpg'],
        ['other_image_url_1~5', '推荐', '副图URL，场景图/功能图/尺寸图等', 'https://...lifestyle.jpg'],
        ['item_weight', '推荐', '商品净重(磅)', '0.12'],
        ['item_length/width/height', '推荐', '商品尺寸(英寸)', '2.5'],
        ['package_weight/length/width/height', '推荐', '包装尺寸(影响FBA费用计算)', '6.5'],
        ['parent_child', '变体用', '父子关系: Parent/Child', 'Child'],
        ['parent_sku', '变体用', '子商品的父SKU', 'WBE-001'],
        ['variation_theme', '变体用', '变体维度: Color/Size/SizeColor', 'Color'],
        ['color_name / size_name', '变体用', '变体属性值', 'Black / Large'],
        ['country_of_origin', '推荐', '原产国 ISO代码', 'CN / US / DE'],
        ['update_delete', '必填', '操作类型', 'Update / PartialUpdate / Delete'],
    ]

    header_fill2 = PatternFill(start_color='FF2D3748', end_color='FF2D3748', fill_type='solid')
    for row_idx, row_data in enumerate(field_docs, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = header_fill2
                cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 保存
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'input')
    os.makedirs(output_dir, exist_ok=True)
    
    # 删除旧的简单sample
    old_sample = os.path.join(output_dir, 'sample.xlsx')
    if os.path.exists(old_sample):
        os.remove(old_sample)
    
    output_path = os.path.join(output_dir, 'amazon_test_products.xlsx')
    wb.save(output_path)
    print(f"✅ 亚马逊完整测试数据已生成: {output_path}")
    print("   - Sheet 1: 3条完整商品数据 (51个字段)")
    print("   - Sheet 2: 字段说明文档")
    return output_path


if __name__ == '__main__':
    create_full_sample()
