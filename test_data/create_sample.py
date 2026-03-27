"""
生成测试Excel数据
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_sample():
    wb = Workbook()
    ws = wb.active
    ws.title = "商品数据"

    # 表头
    headers = ['SKU', '标题', '描述', '价格', '图片URL', '品牌', '分类', '关键词']
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # 测试数据
    data = [
        ['TEST-001', 'Wireless Bluetooth Earbuds', 'High quality wireless earbuds with noise cancellation',
         29.99, 'https://example.com/earbuds.jpg', 'TechBrand', 'Electronics', 'wireless earbuds bluetooth'],
        ['TEST-002', 'Stainless Steel Water Bottle', 'Insulated water bottle keeps drinks cold 24h hot 12h',
         19.99, 'https://example.com/bottle.jpg', 'HydroPro', 'Sports', 'water bottle insulated stainless'],
        ['TEST-003', 'LED Desk Lamp', 'Adjustable LED desk lamp with USB charging port',
         35.99, 'https://example.com/lamp.jpg', 'BrightHome', 'Home', 'desk lamp led adjustable'],
        ['TEST-004', 'Yoga Mat', 'Non-slip exercise yoga mat 6mm thick with carry strap',
         24.99, 'https://example.com/yoga.jpg', 'FitLife', 'Sports', 'yoga mat non slip exercise'],
        ['TEST-005', 'Phone Case iPhone 15', 'Shockproof clear case for iPhone 15 Pro Max',
         12.99, 'https://example.com/case.jpg', 'ShieldTech', 'Electronics', 'phone case iphone 15 clear'],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # 保存
    output_dir = os.path.join(os.path.dirname(__file__), '..', 'input')
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, 'sample.xlsx')
    wb.save(output_path)
    print(f"✅ 测试数据已生成: {output_path}")
    return output_path

if __name__ == '__main__':
    create_sample()
