"""
生成SP-API完整测试数据Excel
覆盖所有54列, 包含3组变体+3个独立商品
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = "商品采集模板"

HEADERS = [
    'SKU', 'product_id', 'product_id_type', 'item_name', 'brand_name',
    'manufacturer', 'part_number', 'product_type',
    'recommended_browse_nodes', 'item_type_keyword',
    'product_description', 'bullet_point_1', 'bullet_point_2',
    'bullet_point_3', 'bullet_point_4', 'bullet_point_5',
    'generic_keywords',
    'standard_price', 'quantity', 'fulfillment_channel', 'condition_type',
    'main_image_url', 'other_image_url_1', 'other_image_url_2',
    'other_image_url_3', 'other_image_url_4', 'other_image_url_5',
    'swatch_image_url',
    'item_weight', 'item_weight_unit_of_measure',
    'item_length', 'item_width', 'item_height',
    'item_dimensions_unit_of_measure',
    'package_weight', 'package_length', 'package_width', 'package_height',
    'number_of_items',
    'parent_child', 'parent_sku', 'relationship_type', 'variation_theme',
    'color_name', 'size_name',
    'material_type', 'country_of_origin',
    'batteries_required', 'are_batteries_included',
    'cpsia_cautionary_statement',
    'target_audience_keywords', 'style_name', 'department_name',
    'update_delete',
]

# Row 1: 分组表头
groups = [
    ('基础必填', 1, 8, 'FFDC3545'), ('分类', 9, 10, 'FF6C757D'),
    ('描述文案', 11, 17, 'FFFD7E14'), ('销售信息', 18, 21, 'FF007BFF'),
    ('图片', 22, 28, 'FF28A745'), ('物流尺寸', 29, 38, 'FF6F42C1'),
    ('数量', 39, 39, 'FF6C757D'), ('变体', 40, 45, 'FF17A2B8'),
    ('属性', 46, 47, 'FF6C757D'), ('合规', 48, 50, 'FF795548'),
    ('其他', 51, 54, 'FF6C757D'),
]
for name, s, e, clr in groups:
    for c in range(s, e + 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.fill = PatternFill(start_color=clr, end_color=clr, fill_type='solid')
        cell.font = Font(color='FFFFFFFF', bold=True, size=10)
        cell.alignment = Alignment(horizontal='center')

# Row 2: 列名
for i, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=2, column=i, value=h)
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color='FFE8E8E8', end_color='FFE8E8E8', fill_type='solid')

# ===== 商品数据 =====
products = [
    # ===== 变体组1: 无线耳机 (Color, 1P + 3C) =====
    {
        'SKU': 'WBE-001',
        'item_name': 'ProSound X5 Wireless Bluetooth 5.3 Earbuds - Active Noise Cancelling',
        'brand_name': 'ProSound',
        'manufacturer': 'ProSound Technology Co., Ltd.',
        'product_type': 'HEADPHONES',
        'item_type_keyword': 'in-ear-headphones',
        'recommended_browse_nodes': '172541',
        'parent_child': 'Parent',
        'variation_theme': 'Color',
    },
    {
        'SKU': 'WBE-001-BLK', 'product_id': '012345678901', 'product_id_type': 'UPC',
        'item_name': 'ProSound X5 Wireless Bluetooth 5.3 Earbuds - ANC, 40H Playtime, IPX7 Waterproof (Black)',
        'brand_name': 'ProSound', 'manufacturer': 'ProSound Technology Co., Ltd.',
        'part_number': 'PS-X5-BLK', 'product_type': 'HEADPHONES',
        'recommended_browse_nodes': '172541', 'item_type_keyword': 'in-ear-headphones',
        'product_description': 'Experience premium sound with ProSound X5 earbuds featuring Bluetooth 5.3, active noise cancellation, and 40-hour battery life. IPX7 waterproof rating makes them perfect for workouts.',
        'bullet_point_1': 'ADVANCED BLUETOOTH 5.3 - Stable connection within 50ft range, seamless switching between devices',
        'bullet_point_2': 'ACTIVE NOISE CANCELLATION - Reduce ambient noise by up to 35dB for immersive listening experience',
        'bullet_point_3': '40-HOUR BATTERY LIFE - 8 hours per charge plus 32 hours with USB-C fast charging case',
        'bullet_point_4': 'IPX7 WATERPROOF - Sweat and rain resistant, perfect for workouts and outdoor activities',
        'bullet_point_5': 'TOUCH CONTROL & COMFORT - Intuitive touch controls, ergonomic design with 3 sizes of ear tips',
        'generic_keywords': 'wireless earbuds bluetooth noise cancelling waterproof sport earphones running gym',
        'standard_price': 29.99, 'quantity': 500,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/61SUkFDJR7L._AC_SL1500_.jpg',
        'other_image_url_1': 'https://m.media-amazon.com/images/I/71z5VIqWYQL._AC_SL1500_.jpg',
        'other_image_url_2': 'https://m.media-amazon.com/images/I/71XOz9DLNQL._AC_SL1500_.jpg',
        'other_image_url_3': 'https://m.media-amazon.com/images/I/61x2O0SZodL._AC_SL1500_.jpg',
        'item_weight': 56, 'item_weight_unit_of_measure': 'grams',
        'item_length': 6.5, 'item_width': 5.0, 'item_height': 3.2,
        'item_dimensions_unit_of_measure': 'centimeters',
        'package_weight': 180, 'package_length': 12.0, 'package_width': 10.0, 'package_height': 5.0,
        'number_of_items': 1,
        'parent_child': 'Child', 'parent_sku': 'WBE-001',
        'relationship_type': 'Variation', 'variation_theme': 'Color',
        'color_name': 'Black', 'material_type': 'Plastic', 'country_of_origin': 'CN',
        'batteries_required': 'Yes', 'are_batteries_included': 'Yes',
        'target_audience_keywords': 'unisex-adults', 'department_name': 'Unisex',
    },
    {
        'SKU': 'WBE-001-WHT', 'product_id': '012345678902', 'product_id_type': 'UPC',
        'item_name': 'ProSound X5 Wireless Bluetooth 5.3 Earbuds - ANC, 40H Playtime (White)',
        'brand_name': 'ProSound', 'part_number': 'PS-X5-WHT', 'product_type': 'HEADPHONES',
        'product_description': 'Premium sound with Bluetooth 5.3, ANC, and 40-hour battery.',
        'bullet_point_1': 'BLUETOOTH 5.3 - Stable 50ft range',
        'bullet_point_2': 'NOISE CANCELLATION - 35dB reduction',
        'bullet_point_3': '40-HOUR BATTERY with USB-C charging',
        'bullet_point_4': 'IPX7 WATERPROOF for workouts',
        'bullet_point_5': 'TOUCH CONTROL with 3 ear tip sizes',
        'generic_keywords': 'wireless earbuds bluetooth white noise cancelling',
        'standard_price': 29.99, 'quantity': 350,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/61eVfZMjmKL._AC_SL1500_.jpg',
        'item_weight': 56, 'item_weight_unit_of_measure': 'grams',
        'parent_child': 'Child', 'parent_sku': 'WBE-001',
        'relationship_type': 'Variation', 'variation_theme': 'Color',
        'color_name': 'White', 'material_type': 'Plastic', 'country_of_origin': 'CN',
        'batteries_required': 'Yes', 'are_batteries_included': 'Yes',
    },
    {
        'SKU': 'WBE-001-RED', 'product_id': '012345678903', 'product_id_type': 'UPC',
        'item_name': 'ProSound X5 Wireless Bluetooth 5.3 Earbuds - ANC (Red)',
        'brand_name': 'ProSound', 'part_number': 'PS-X5-RED', 'product_type': 'HEADPHONES',
        'standard_price': 32.99, 'quantity': 200,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/71oBvUFmqYL._AC_SL1500_.jpg',
        'parent_child': 'Child', 'parent_sku': 'WBE-001',
        'relationship_type': 'Variation', 'variation_theme': 'Color',
        'color_name': 'Red', 'country_of_origin': 'CN',
        'batteries_required': 'Yes', 'are_batteries_included': 'Yes',
    },

    # ===== 变体组2: 保温杯 (SizeName-ColorName, 1P + 4C) =====
    {
        'SKU': 'SSB-002',
        'item_name': 'HydroPeak Insulated Stainless Steel Water Bottle with Straw Lid',
        'brand_name': 'HydroPeak', 'manufacturer': 'HydroPeak Outdoor LLC',
        'product_type': 'HOME', 'item_type_keyword': 'water-bottles',
        'parent_child': 'Parent', 'variation_theme': 'SizeName-ColorName',
    },
    {
        'SKU': 'SSB-002-32-BL', 'product_id': '012345678904', 'product_id_type': 'UPC',
        'item_name': 'HydroPeak 32oz Insulated Stainless Steel Water Bottle with Straw Lid (Ocean Blue)',
        'brand_name': 'HydroPeak', 'manufacturer': 'HydroPeak Outdoor LLC',
        'part_number': 'HP-32-BL', 'product_type': 'HOME',
        'product_description': 'Vacuum insulated bottle. Keeps drinks cold 24 hours, hot 12 hours. BPA free.',
        'bullet_point_1': 'VACUUM INSULATED - Double wall keeps cold 24H, hot 12H',
        'bullet_point_2': 'PREMIUM 18/8 STAINLESS STEEL - BPA free, food-grade',
        'bullet_point_3': 'STRAW LID INCLUDED - Easy sip, leak proof design',
        'bullet_point_4': 'PERFECT 32oz SIZE - Fits most cup holders',
        'bullet_point_5': 'SWEAT-FREE EXTERIOR - No condensation, comfortable grip',
        'generic_keywords': 'water bottle insulated stainless steel straw lid 32oz sports',
        'standard_price': 24.99, 'quantity': 800,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/61RwGsmeepL._AC_SL1500_.jpg',
        'other_image_url_1': 'https://m.media-amazon.com/images/I/71FGRW6pIEL._AC_SL1500_.jpg',
        'item_weight': 420, 'item_weight_unit_of_measure': 'grams',
        'item_length': 27.0, 'item_width': 8.5, 'item_height': 8.5,
        'item_dimensions_unit_of_measure': 'centimeters',
        'parent_child': 'Child', 'parent_sku': 'SSB-002',
        'relationship_type': 'Variation', 'variation_theme': 'SizeName-ColorName',
        'color_name': 'Ocean Blue', 'size_name': '32oz',
        'material_type': 'Stainless Steel', 'country_of_origin': 'CN',
        'batteries_required': 'No', 'are_batteries_included': 'No',
    },
    {
        'SKU': 'SSB-002-32-RD', 'product_id': '012345678905', 'product_id_type': 'UPC',
        'item_name': 'HydroPeak 32oz Insulated Water Bottle (Red)',
        'brand_name': 'HydroPeak',
        'standard_price': 24.99, 'quantity': 600,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/71RLh6jJOPL._AC_SL1500_.jpg',
        'parent_child': 'Child', 'parent_sku': 'SSB-002',
        'relationship_type': 'Variation', 'variation_theme': 'SizeName-ColorName',
        'color_name': 'Red', 'size_name': '32oz',
        'material_type': 'Stainless Steel', 'country_of_origin': 'CN',
    },
    {
        'SKU': 'SSB-002-16-BL', 'product_id': '012345678906', 'product_id_type': 'UPC',
        'item_name': 'HydroPeak 16oz Insulated Water Bottle (Ocean Blue)',
        'brand_name': 'HydroPeak',
        'standard_price': 19.99, 'quantity': 400,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/61RwGsmeepL._AC_SL1500_.jpg',
        'parent_child': 'Child', 'parent_sku': 'SSB-002',
        'relationship_type': 'Variation', 'variation_theme': 'SizeName-ColorName',
        'color_name': 'Ocean Blue', 'size_name': '16oz',
        'country_of_origin': 'CN',
    },
    {
        'SKU': 'SSB-002-24-BK', 'product_id': '012345678907', 'product_id_type': 'UPC',
        'item_name': 'HydroPeak 24oz Insulated Water Bottle (Black)',
        'brand_name': 'HydroPeak',
        'standard_price': 22.99, 'quantity': 550,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/71DXCoYyOeL._AC_SL1500_.jpg',
        'parent_child': 'Child', 'parent_sku': 'SSB-002',
        'relationship_type': 'Variation', 'variation_theme': 'SizeName-ColorName',
        'color_name': 'Black', 'size_name': '24oz',
        'country_of_origin': 'CN',
    },

    # ===== 独立商品: LED台灯 =====
    {
        'SKU': 'LED-003-WHT', 'product_id': '012345678908', 'product_id_type': 'UPC',
        'item_name': 'BrightDesk Pro LED Desk Lamp with Wireless Charger - 5 Color Modes, 7 Brightness Levels',
        'brand_name': 'BrightDesk', 'manufacturer': 'BrightDesk Lighting Inc.',
        'part_number': 'BD-PRO-WHT', 'product_type': 'LIGHTING',
        'item_type_keyword': 'desk-lamps',
        'product_description': 'LED desk lamp with built-in 10W Qi wireless charger. 5 color temperatures (2700K-6500K) and 7 brightness levels.',
        'bullet_point_1': 'BUILT-IN WIRELESS CHARGER - 10W Qi fast charging for iPhone and Samsung',
        'bullet_point_2': '5 COLOR MODES - From warm 2700K to cool 6500K daylight',
        'bullet_point_3': '7 BRIGHTNESS LEVELS - Smooth dimming for any task',
        'bullet_point_4': 'USB-A PORT - Extra charging port for second device',
        'bullet_point_5': 'FOLDABLE DESIGN - Space-saving, adjustable arm and head',
        'generic_keywords': 'desk lamp led wireless charger usb eye care foldable office home',
        'standard_price': 39.99, 'quantity': 300,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/61QhSjmwZ4L._AC_SL1500_.jpg',
        'other_image_url_1': 'https://m.media-amazon.com/images/I/71mD-8RMNSL._AC_SL1500_.jpg',
        'other_image_url_2': 'https://m.media-amazon.com/images/I/61UaWZK3URL._AC_SL1500_.jpg',
        'item_weight': 850, 'item_weight_unit_of_measure': 'grams',
        'item_length': 40.0, 'item_width': 12.0, 'item_height': 45.0,
        'item_dimensions_unit_of_measure': 'centimeters',
        'package_weight': 1200, 'package_length': 48.0, 'package_width': 15.0, 'package_height': 10.0,
        'material_type': 'Aluminum', 'country_of_origin': 'CN',
        'batteries_required': 'No', 'are_batteries_included': 'No',
        'target_audience_keywords': 'unisex-adults', 'department_name': 'Home',
    },

    # ===== 独立商品: 礼品盒 =====
    {
        'SKU': 'GBX-004-DLX', 'product_id': '012345678909', 'product_id_type': 'UPC',
        'item_name': "Premium Men's Birthday Gift Box Set - LED Presentation Box (Deluxe Edition)",
        'brand_name': 'GiftCraft', 'manufacturer': 'GiftCraft Design Co.',
        'part_number': 'GC-DLX-004', 'product_type': 'GIFT_SET',
        'product_description': "Luxury men's gift box with curated premium items in elegant LED-lit presentation box.",
        'bullet_point_1': 'PREMIUM GIFT SET - Curated collection of luxury items for him',
        'bullet_point_2': 'LED PRESENTATION BOX - Elegant magnetic closure with built-in LED',
        'bullet_point_3': "PERFECT FOR ANY OCCASION - Birthday, anniversary, Valentine's Day",
        'bullet_point_4': 'HIGH-QUALITY ITEMS - Each item carefully selected',
        'bullet_point_5': 'READY TO GIFT - Beautifully packaged, no wrapping needed',
        'generic_keywords': 'mens gift box birthday anniversary present luxury premium set surprise',
        'standard_price': 45.99, 'quantity': 150,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/81D3cHVDniL._AC_SL1500_.jpg',
        'other_image_url_1': 'https://m.media-amazon.com/images/I/71CZLL8dqvL._AC_SL1500_.jpg',
        'item_weight': 1200, 'item_weight_unit_of_measure': 'grams',
        'item_length': 30.0, 'item_width': 25.0, 'item_height': 12.0,
        'item_dimensions_unit_of_measure': 'centimeters',
        'material_type': 'Mixed', 'country_of_origin': 'CN',
        'batteries_required': 'Yes', 'are_batteries_included': 'Yes',
        'target_audience_keywords': 'mens', 'department_name': 'Mens',
    },

    # ===== 变体组3: 手机壳 (Color, 1P + 2C) =====
    {
        'SKU': 'PHC-005',
        'item_name': 'UltraShield iPhone 15 Pro Max Case - Military Grade Drop Protection',
        'brand_name': 'UltraShield', 'manufacturer': 'UltraShield Electronics',
        'product_type': 'CELLULAR_PHONE_CASE', 'item_type_keyword': 'cell-phone-cases',
        'parent_child': 'Parent', 'variation_theme': 'Color',
    },
    {
        'SKU': 'PHC-005-CLR', 'product_id': '012345678910', 'product_id_type': 'UPC',
        'item_name': 'UltraShield iPhone 15 Pro Max Clear Case - Military Grade, Anti-Yellowing',
        'brand_name': 'UltraShield', 'manufacturer': 'UltraShield Electronics',
        'part_number': 'US-IP15PM-CLR', 'product_type': 'CELLULAR_PHONE_CASE',
        'product_description': 'Crystal clear protection for iPhone 15 Pro Max with military-grade drop tested corners.',
        'bullet_point_1': 'MIL-STD-810G CERTIFIED - 10ft drop tested protection',
        'bullet_point_2': 'ANTI-YELLOWING - UV resistant material stays crystal clear',
        'bullet_point_3': 'SLIM 1.2mm PROFILE - Wireless charging compatible',
        'bullet_point_4': 'PRECISE CUTOUTS - Perfect access to buttons and ports',
        'bullet_point_5': 'RAISED EDGES - Protects screen and camera from scratches',
        'generic_keywords': 'iphone 15 pro max case clear military grade slim protective anti yellowing',
        'standard_price': 15.99, 'quantity': 2000,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/71v4oEz4ldL._AC_SL1500_.jpg',
        'other_image_url_1': 'https://m.media-amazon.com/images/I/71uqMjczAHL._AC_SL1500_.jpg',
        'item_weight': 38, 'item_weight_unit_of_measure': 'grams',
        'item_length': 16.0, 'item_width': 8.0, 'item_height': 1.2,
        'item_dimensions_unit_of_measure': 'centimeters',
        'parent_child': 'Child', 'parent_sku': 'PHC-005',
        'relationship_type': 'Variation', 'variation_theme': 'Color',
        'color_name': 'Clear', 'material_type': 'TPU', 'country_of_origin': 'CN',
    },
    {
        'SKU': 'PHC-005-BLK', 'product_id': '012345678911', 'product_id_type': 'UPC',
        'item_name': 'UltraShield iPhone 15 Pro Max Case - Military Grade (Black)',
        'brand_name': 'UltraShield', 'part_number': 'US-IP15PM-BLK',
        'product_type': 'CELLULAR_PHONE_CASE',
        'standard_price': 16.99, 'quantity': 1500,
        'fulfillment_channel': 'DEFAULT', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/61WMN0e9XEL._AC_SL1500_.jpg',
        'parent_child': 'Child', 'parent_sku': 'PHC-005',
        'relationship_type': 'Variation', 'variation_theme': 'Color',
        'color_name': 'Black', 'material_type': 'TPU', 'country_of_origin': 'CN',
    },

    # ===== 独立商品: 瑜伽垫 (FBA发货) =====
    {
        'SKU': 'YGM-006-PP', 'product_id': '012345678912', 'product_id_type': 'UPC',
        'item_name': 'ZenFlex Premium Yoga Mat 6mm - Non-Slip, Eco-Friendly TPE, Alignment Lines (Purple)',
        'brand_name': 'ZenFlex', 'manufacturer': 'ZenFlex Fitness Inc.',
        'part_number': 'ZF-YM6-PP', 'product_type': 'SPORTING_GOODS',
        'item_type_keyword': 'yoga-mats',
        'product_description': 'Premium 6mm thick TPE yoga mat with alignment guide lines. Non-slip dual texture surface for all yoga styles.',
        'bullet_point_1': 'ECO-FRIENDLY TPE - Non-toxic, recyclable, no PVC or latex',
        'bullet_point_2': 'DUAL TEXTURE SURFACE - Anti-slip on both sides for safety',
        'bullet_point_3': 'ALIGNMENT LINES - Built-in guide for proper form',
        'bullet_point_4': '6MM THICKNESS - Perfect cushioning for joints, only 2.2lbs',
        'bullet_point_5': 'CARRYING STRAP INCLUDED - Easy transport to studio or gym',
        'generic_keywords': 'yoga mat non slip eco friendly tpe thick alignment lines exercise gym',
        'standard_price': 28.99, 'quantity': 1000,
        'fulfillment_channel': 'AMAZON_NA', 'condition_type': 'New',
        'main_image_url': 'https://m.media-amazon.com/images/I/71CZLL8dqvL._AC_SL1500_.jpg',
        'other_image_url_1': 'https://m.media-amazon.com/images/I/81y8jNOMHJL._AC_SL1500_.jpg',
        'item_weight': 1000, 'item_weight_unit_of_measure': 'grams',
        'item_length': 183.0, 'item_width': 61.0, 'item_height': 0.6,
        'item_dimensions_unit_of_measure': 'centimeters',
        'package_weight': 1200, 'package_length': 62.0, 'package_width': 15.0, 'package_height': 15.0,
        'number_of_items': 1,
        'material_type': 'TPE', 'country_of_origin': 'CN',
        'color_name': 'Purple',
        'target_audience_keywords': 'unisex-adults', 'style_name': 'Alignment',
        'department_name': 'Unisex',
    },
]

# 写入数据
for ri, p in enumerate(products, 3):
    for ci, h in enumerate(HEADERS, 1):
        v = p.get(h)
        if v is not None:
            ws.cell(row=ri, column=ci, value=v)

ws.freeze_panes = 'A3'

# 保存到input目录
output_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'input')
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, 'sp_api_full_test.xlsx')
wb.save(output_path)

print(f"Created: {output_path}")
print(f"Total rows: {len(products)}")
print(f"Total columns: {len(HEADERS)}")
print(f"  WBE-001: 1 Parent + 3 Children (Color)")
print(f"  SSB-002: 1 Parent + 4 Children (SizeName-ColorName)")
print(f"  PHC-005: 1 Parent + 2 Children (Color)")
print(f"  LED-003-WHT: Standalone")
print(f"  GBX-004-DLX: Standalone")
print(f"  YGM-006-PP:  Standalone (FBA)")
