from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from core.temp_image_host import populate_ai_image_urls


def _create_excel(path: Path, image_path: str, existing_url: str = ""):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SKU", "AI主图路径", "AI主图URL"])
    worksheet.append(["SKU-1", image_path, existing_url])
    workbook.save(path)


def test_populate_ai_image_urls_uploads_local_paths():
    suffix = uuid4().hex[:8]
    output_dir = Path("output").resolve()
    excel_path = output_dir / f"test_temp_host_{suffix}.xlsx"
    image_path = output_dir / f"test_temp_host_{suffix}.jpg"

    try:
        image_path.write_bytes(b"fake-image-data")
        _create_excel(excel_path, str(image_path))

        uploaded_paths = []

        def fake_uploader(path: str) -> str:
            uploaded_paths.append(path)
            return "https://x0.at/test-image.jpg"

        stats = populate_ai_image_urls(str(excel_path), uploader=fake_uploader)

        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        assert worksheet["C2"].value == "https://x0.at/test-image.jpg"
        workbook.close()

        assert uploaded_paths == [str(image_path)]
        assert stats["uploaded"] == 1
        assert stats["failed"] == 0
    finally:
        if excel_path.exists():
            excel_path.unlink()
        if image_path.exists():
            image_path.unlink()


def test_populate_ai_image_urls_skips_existing_url_by_default():
    suffix = uuid4().hex[:8]
    output_dir = Path("output").resolve()
    excel_path = output_dir / f"test_temp_host_skip_{suffix}.xlsx"
    image_path = output_dir / f"test_temp_host_skip_{suffix}.jpg"

    try:
        image_path.write_bytes(b"fake-image-data")
        _create_excel(excel_path, str(image_path), existing_url="https://x0.at/existing.jpg")

        stats = populate_ai_image_urls(
            str(excel_path),
            uploader=lambda path: "https://x0.at/new.jpg",
        )

        workbook = load_workbook(excel_path)
        worksheet = workbook.active
        assert worksheet["C2"].value == "https://x0.at/existing.jpg"
        workbook.close()

        assert stats["skipped"] == 1
        assert stats["uploaded"] == 0
    finally:
        if excel_path.exists():
            excel_path.unlink()
        if image_path.exists():
            image_path.unlink()
