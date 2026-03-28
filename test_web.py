import os
import json
import time
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from PIL import Image
import xlrd
import xlwt

import web.app as web_app
from web.app import app, config as web_config


def _build_excel_bytes(row_count=1):
    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url'])
    for idx in range(1, row_count + 1):
        ws.append([f'SKU-{idx}', f'Demo Product {idx}', f'https://example.com/demo-{idx}.jpg'])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_xls_bytes():
    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')
    headers = ['SKU', 'item_name', 'main_image_url']
    values = ['SKU-1', 'Demo Product', 'https://example.com/demo.jpg']

    for col_idx, header in enumerate(headers):
        ws.write(0, col_idx, header)
    for col_idx, value in enumerate(values):
        ws.write(1, col_idx, value)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_test_image_base64():
    buffer = BytesIO()
    Image.new('RGB', (2, 2), color='white').save(buffer, format='JPEG')
    return buffer.getvalue()


def _wait_for_task(client, task_id, timeout=3.0):
    deadline = time.time() + timeout
    last_payload = None
    while time.time() < deadline:
        response = client.get(f'/api/tasks/{task_id}')
        assert response.status_code == 200
        last_payload = response.get_json()
        if last_payload.get('status') in ('completed', 'failed'):
            return last_payload
        time.sleep(0.05)
    raise AssertionError(f'任务未在 {timeout} 秒内完成: {task_id}, last={last_payload}')


def test_web_upload_endpoint_reads_excel(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_upload_{uuid4().hex[:8]}.xlsx'
    uploaded_path = input_dir / filename

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_config, 'OUTPUT_DIR', str(output_dir))

        client = app.test_client()
        response = client.post(
            '/api/upload',
            data={'file': (_build_excel_bytes(), filename)},
            content_type='multipart/form-data',
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['total_rows'] == 1
        assert payload['column_mapping']['sku'] == 'SKU'
        assert payload['column_mapping']['title'] == 'item_name'
    finally:
        if uploaded_path.exists():
            uploaded_path.unlink()


def test_web_upload_endpoint_accepts_uppercase_extension(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_upload_{uuid4().hex[:8]}.XLSX'
    uploaded_path = input_dir / filename

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_config, 'OUTPUT_DIR', str(output_dir))

        client = app.test_client()
        response = client.post(
            '/api/upload',
            data={'file': (_build_excel_bytes(), filename)},
            content_type='multipart/form-data',
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['filename'] == filename
    finally:
        if uploaded_path.exists():
            uploaded_path.unlink()


def test_web_upload_endpoint_reads_xls(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_upload_{uuid4().hex[:8]}.xls'
    uploaded_path = input_dir / filename

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_config, 'OUTPUT_DIR', str(output_dir))

        client = app.test_client()
        response = client.post(
            '/api/upload',
            data={'file': (_build_xls_bytes(), filename)},
            content_type='multipart/form-data',
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['filename'] == filename
        assert payload['column_mapping']['sku'] == 'SKU'
        assert payload['file_type'] == 'xls'
        assert '兼容模式' in payload['warning']
    finally:
        if uploaded_path.exists():
            uploaded_path.unlink()


def test_web_upload_endpoint_enforces_batch_limit(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_upload_limit_{uuid4().hex[:8]}.xlsx'
    uploaded_path = input_dir / filename
    original_limit = web_config.BATCH_LIMIT

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_config, 'OUTPUT_DIR', str(output_dir))
        monkeypatch.setattr(web_config, 'BATCH_LIMIT', 1)

        client = app.test_client()
        response = client.post(
            '/api/upload',
            data={'file': (_build_excel_bytes(row_count=2), filename)},
            content_type='multipart/form-data',
        )

        assert response.status_code == 400
        payload = response.get_json()
        assert '超过当前导入上限 1 条' in payload['error']
        assert not uploaded_path.exists()
    finally:
        web_config.BATCH_LIMIT = original_limit
        if uploaded_path.exists():
            uploaded_path.unlink()


def test_template_recommend_endpoint_returns_candidates(monkeypatch):
    monkeypatch.setattr(web_app, 'recommend_product_types', lambda **kwargs: {
        'query': 'wireless mouse',
        'title': 'Wireless Mouse',
        'source': 'amazon_api',
        'marketplace': 'US',
        'candidates': [
            {'product_type': 'COMPUTER_INPUT_DEVICE', 'display_name': 'Computer Input Device', 'score': 0.98},
            {'product_type': 'KEYBOARDS', 'display_name': 'Keyboards', 'score': 0.42},
        ],
    })

    client = app.test_client()
    response = client.post('/api/templates/recommend', json={'keyword': 'wireless mouse'})

    assert response.status_code == 200
    payload = response.get_json()
    assert payload['success'] is True
    assert payload['candidates'][0]['product_type'] == 'COMPUTER_INPUT_DEVICE'


def test_templates_generate_route_creates_task_and_downloadable_workbook(monkeypatch):
    from core.template_generator import build_template_definition, generate_template

    temp_dir = Path(web_config.OUTPUT_DIR).resolve() / f'tmp_template_generate_{uuid4().hex[:8]}'
    temp_dir.mkdir(parents=True, exist_ok=True)
    schema_fields = {
        'required_fields': [
            {'name': 'item_name', 'title': 'Item Name', 'description': 'Title', 'group': 'product_identity'},
            {'name': 'manufacturer', 'title': 'Manufacturer', 'description': 'Maker', 'group': 'product_identity'},
        ],
        'optional_fields': [
            {'name': 'generic_keyword', 'title': 'Generic Keyword', 'description': 'Search terms', 'group': 'product_description'},
        ],
        'enum_fields': {},
        'field_groups': {},
    }
    holder = {}

    def fake_ensure_definition(product_type, marketplace='US', variation_mode='single', refresh=False):
        definition = build_template_definition(schema_fields, product_type, marketplace, variation_mode)
        holder['definition'] = definition
        return definition

    def fake_ensure_workbook(template_id):
        definition = holder['definition']
        workbook_path = temp_dir / f'{template_id}.xlsx'
        generate_template(definition, str(workbook_path))
        return str(workbook_path), definition

    monkeypatch.setattr(web_app, 'ensure_template_definition', fake_ensure_definition)
    monkeypatch.setattr(web_app, 'ensure_template_workbook', fake_ensure_workbook)

    client = app.test_client()
    response = client.post('/api/templates/generate', json={
        'product_type': 'COMPUTER_INPUT_DEVICE',
        'variation_mode': 'variation',
    })

    assert response.status_code == 200
    task_id = response.get_json()['task_id']
    task_payload = _wait_for_task(client, task_id)
    assert task_payload['status'] == 'completed'
    assert task_payload['result']['template_id']
    assert task_payload['result']['variation_mode'] == 'variation'

    download_response = client.get(task_payload['result']['download_url'])
    assert download_response.status_code == 200
    wb = load_workbook(BytesIO(download_response.data))
    ws = wb.active
    assert '[必填]' in str(ws.cell(row=1, column=1).value)
    headers = [str(cell.value or '') for cell in ws[2]]
    assert 'sku' in headers
    assert 'product_type' in headers
    assert 'manufacturer' in headers
    assert '__template_meta__' in wb.sheetnames
    wb.close()
    download_response.close()


def test_template_diagnose_route_persists_template_diagnostics(monkeypatch):
    from amazon.mapper import FieldMapper
    from core.template_generator import build_template_definition, generate_template

    temp_dir = Path(web_config.OUTPUT_DIR).resolve() / f'tmp_template_diag_{uuid4().hex[:8]}'
    temp_dir.mkdir(parents=True, exist_ok=True)
    input_path = temp_dir / f'template_diag_{uuid4().hex[:8]}.xlsx'
    schema_fields = {
        'required_fields': [
            {'name': 'item_name', 'title': 'Item Name', 'description': 'Title', 'group': 'product_identity'},
            {'name': 'manufacturer', 'title': 'Manufacturer', 'description': 'Maker', 'group': 'product_identity'},
        ],
        'optional_fields': [
            {'name': 'brand', 'title': 'Brand', 'description': 'Brand', 'group': 'product_identity'},
        ],
        'enum_fields': {},
        'field_groups': {},
    }
    definition = build_template_definition(schema_fields, 'COMPUTER_INPUT_DEVICE', 'US', 'single')
    generate_template(definition, str(input_path))

    wb = load_workbook(input_path)
    ws = wb.active
    header_map = {str(ws.cell(row=2, column=idx).value): idx for idx in range(1, ws.max_column + 1)}
    row_idx = 3
    ws.cell(row=row_idx, column=header_map['sku'], value='SKU-1')
    ws.cell(row=row_idx, column=header_map['product_type'], value='COMPUTER_INPUT_DEVICE')
    ws.cell(row=row_idx, column=header_map['item_name'], value='Demo Mouse')
    ws.cell(row=row_idx, column=header_map['upc'], value='123456789012')
    ws.cell(row=row_idx, column=header_map['external_product_id_type'], value='UPC')
    ws.cell(row=row_idx, column=header_map['product_identity_mode'], value='real_gtin')
    wb.save(input_path)
    wb.close()

    monkeypatch.setattr(web_app, '_load_template_definition_for_file', lambda filepath: ({
        'template_id': definition['template_id'],
        'product_type': 'COMPUTER_INPUT_DEVICE',
        'variation_mode': 'single',
    }, definition))
    monkeypatch.setattr(web_app, '_build_listing_api_context', lambda account_id='': (None, FieldMapper('ATVPDKIKX0DER'), None, '未配置亚马逊账号'))
    monkeypatch.setattr(FieldMapper, '_load_schema_fields', lambda self, product_type: schema_fields)
    monkeypatch.setattr(web_app, '_probe_product_media', lambda product: {
        'status': 'none', 'total': 0, 'passed': 0, 'failed': 0, 'checks': [],
    })

    client = app.test_client()
    response = client.post('/api/template-diagnose', json={'file': str(input_path)})
    assert response.status_code == 200
    task_id = response.get_json()['task_id']
    task_payload = _wait_for_task(client, task_id)

    assert task_payload['status'] == 'completed'
    assert task_payload['result']['fail'] == 1
    result_entry = task_payload['result']['results'][0]
    assert result_entry['template']['required_missing'][0]['key'] == 'manufacturer'

    wb2 = load_workbook(input_path)
    ws2 = wb2.active
    headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[2]]
    header_map = {header: idx + 1 for idx, header in enumerate(headers)}
    assert ws2.cell(row=3, column=header_map['template_required_missing_count']).value == 1
    assert 'manufacturer' in str(ws2.cell(row=3, column=header_map['template_blocking_issues']).value or '')
    assert ws2.cell(row=3, column=header_map['listing_check_status']).value == 'fail'
    wb2.close()

    products_response = client.get('/api/products', query_string={'file': str(input_path)})
    sku_entry = products_response.get_json()['products'][0]['skus'][0]
    assert sku_entry['template_product_type'] == 'COMPUTER_INPUT_DEVICE'
    assert sku_entry['template_required_missing_count'] == '1'
    assert 'manufacturer' in sku_entry['template_blocking_issues']
    input_path.unlink()
    temp_dir.rmdir()


def test_merge_template_diagnostic_promotes_preview_missing_fields_to_blocking():
    result_entry = {
        'summary_text': '预览 INVALID',
        'missing_fields': [
            {'name': 'compatible_devices', 'title': 'Compatible Devices', 'source': 'amazon_preview'},
            {'name': 'hand_orientation', 'title': 'Hand Orientation', 'source': 'amazon_preview'},
        ],
    }
    template_eval = {
        'template_id': 'tpl_input_mouse',
        'required_total': 13,
        'required_filled': 13,
        'required_missing': [],
        'recommended_missing': [],
        'optional_missing': [],
        'blocking_issues': [],
        'variation_issues': [],
    }

    merged = web_app._merge_template_diagnostic(result_entry, template_eval, [])

    assert merged['template']['blocking_issues']
    assert any('compatible_devices' in msg for msg in merged['template']['blocking_issues'])
    assert any('hand_orientation' in msg for msg in merged['template']['blocking_issues'])
    assert '建议补充 0 项' not in merged['summary_text']


def test_template_overlay_promotes_preview_missing_fields_for_future_templates(monkeypatch):
    import amazon.schema_manager as schema_manager
    import core.template_service as template_service
    from openpyxl import load_workbook

    temp_root = Path(web_config.OUTPUT_DIR).resolve() / f'tmp_template_overlay_{uuid4().hex[:8]}'
    def_dir = temp_root / 'defs'
    gen_dir = temp_root / 'gen'
    overlay_dir = temp_root / 'overlay'
    for directory in (def_dir, gen_dir, overlay_dir):
        directory.mkdir(parents=True, exist_ok=True)

    schema_fields = {
        'required_fields': [{'name': 'item_name', 'title': 'Item Name', 'description': 'Title', 'group': 'product_identity'}],
        'optional_fields': [],
        'enum_fields': {},
        'field_groups': {},
    }

    monkeypatch.setattr(template_service, '_template_definition_dir', lambda: str(def_dir))
    monkeypatch.setattr(template_service, '_generated_template_dir', lambda: str(gen_dir))
    monkeypatch.setattr(template_service, '_template_overlay_dir', lambda: str(overlay_dir))
    monkeypatch.setattr(schema_manager, 'fetch_schema', lambda product_type, marketplace='US', sp_client=None: {'schema': {}, 'property_groups': {}})
    monkeypatch.setattr(schema_manager, 'parse_schema', lambda raw: schema_fields)

    definition = template_service.ensure_template_definition('INPUT_MOUSE', 'US', 'single', refresh=True)
    keys = {item['key'] for item in definition['columns']}
    assert 'compatible_devices' not in keys

    template_service.update_template_overlay('INPUT_MOUSE', 'US', [{'name': 'compatible_devices', 'title': 'Compatible Devices'}])
    updated = template_service.ensure_template_definition('INPUT_MOUSE', 'US', 'single')
    columns = {item['key']: item for item in updated['columns']}
    assert 'compatible_devices' in columns
    assert columns['compatible_devices']['level'] == 'required'

    workbook_path, _ = template_service.ensure_template_workbook(updated['template_id'])
    wb = load_workbook(workbook_path)
    ws = wb.active
    headers = [str(ws.cell(row=2, column=idx).value or '') for idx in range(1, ws.max_column + 1)]
    idx = headers.index('compatible_devices') + 1
    assert str(ws.cell(row=1, column=idx).value or '').startswith('[必填]')
    wb.close()


def test_template_upload_route_starts_auto_diagnosis(monkeypatch):
    def fake_execute(task_id, input_file, account_id='', selected_skus=None):
        web_app._complete_task_record(task_id, message='模板诊断完成', result={
            'input_file': input_file,
            'pass': 1,
            'warn': 0,
            'fail': 0,
            'total': 1,
            'results': [],
            'template': {'template_id': 'tpl_demo', 'product_type': 'DRINKING_CUP', 'variation_mode': 'single'},
        }, progress=1, total=1)

    monkeypatch.setattr(web_app, '_load_template_definition_for_file', lambda filepath: ({
        'template_id': 'tpl_demo',
        'product_type': 'DRINKING_CUP',
        'variation_mode': 'single',
    }, {
        'template_id': 'tpl_demo',
        'product_type': 'DRINKING_CUP',
        'variation_mode': 'single',
        'required_total': 5,
        'recommended_total': 2,
        'columns': [{'key': 'sku', 'level': 'required'}],
    }))
    monkeypatch.setattr(web_app, '_execute_template_diagnosis', fake_execute)

    client = app.test_client()
    response = client.post(
        '/api/template-upload',
        data={'file': (_build_excel_bytes(), f'template_upload_{uuid4().hex[:8]}.xlsx')},
        content_type='multipart/form-data',
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload['success'] is True
    assert payload['template']['product_type'] == 'DRINKING_CUP'
    task_payload = _wait_for_task(client, payload['task_id'])
    assert task_payload['status'] == 'completed'
    uploaded_path = Path(payload['filepath'])
    if uploaded_path.exists():
        uploaded_path.unlink()


def test_template_diagnose_persists_preview_status_fields(monkeypatch):
    from amazon.mapper import FieldMapper
    from core.template_generator import build_template_definition, generate_template

    temp_dir = Path(web_config.OUTPUT_DIR).resolve() / f'tmp_template_preview_status_{uuid4().hex[:8]}'
    temp_dir.mkdir(parents=True, exist_ok=True)
    input_path = temp_dir / f'template_preview_status_{uuid4().hex[:8]}.xlsx'
    schema_fields = {
        'required_fields': [{'name': 'item_name', 'title': 'Item Name', 'description': 'Title', 'group': 'product_identity'}],
        'optional_fields': [],
        'enum_fields': {},
        'field_groups': {},
    }
    definition = build_template_definition(schema_fields, 'INPUT_MOUSE', 'US', 'single')
    generate_template(definition, str(input_path))

    wb = load_workbook(input_path)
    ws = wb.active
    header_map = {str(ws.cell(row=2, column=idx).value): idx for idx in range(1, ws.max_column + 1)}
    ws.cell(row=3, column=header_map['sku'], value='SKU-1')
    ws.cell(row=3, column=header_map['product_type'], value='INPUT_MOUSE')
    ws.cell(row=3, column=header_map['item_name'], value='Demo Mouse')
    ws.cell(row=3, column=header_map['product_identity_mode'], value='gtin_exemption')
    wb.save(input_path)
    wb.close()

    class FakeListingsAPI:
        def put_listings_item(self, sku, product, preview=False):
            return {
                'status': 'INVALID',
                'issues': [{'severity': 'ERROR', 'message': 'Preview failed', 'attributeNames': ['compatible_devices'], 'categories': ['MISSING_ATTRIBUTE'], 'code': '90220'}],
            }

        def get_listings_item(self, sku):
            return None

    monkeypatch.setattr(web_app, '_load_template_definition_for_file', lambda filepath: ({
        'template_id': definition['template_id'],
        'product_type': 'INPUT_MOUSE',
        'variation_mode': 'single',
    }, definition))
    monkeypatch.setattr(web_app, '_build_listing_api_context', lambda account_id='': ({'name': 'US Test'}, FieldMapper('ATVPDKIKX0DER'), FakeListingsAPI(), ''))
    monkeypatch.setattr(FieldMapper, '_load_schema_fields', lambda self, product_type: schema_fields)
    monkeypatch.setattr(web_app, '_probe_product_media', lambda product: {'status': 'none', 'total': 0, 'passed': 0, 'failed': 0, 'checks': []})

    client = app.test_client()
    response = client.post('/api/template-diagnose', json={'file': str(input_path)})
    assert response.status_code == 200
    _wait_for_task(client, response.get_json()['task_id'])

    wb2 = load_workbook(input_path)
    ws2 = wb2.active
    headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[2]]
    header_map = {header: idx + 1 for idx, header in enumerate(headers)}
    assert ws2.cell(row=3, column=header_map['preview_status']).value == 'INVALID'
    assert 'Preview failed' in str(ws2.cell(row=3, column=header_map['preview_message']).value or '')
    wb2.close()

    products_response = client.get('/api/products', query_string={'file': str(input_path)})
    sku_entry = products_response.get_json()['products'][0]['skus'][0]
    assert sku_entry['preview_status'] == 'INVALID'


def test_template_diagnose_uses_template_labels_for_preview_missing_fields(monkeypatch):
    from amazon.mapper import FieldMapper
    from core.template_generator import build_template_definition, generate_template

    temp_dir = Path(web_config.OUTPUT_DIR).resolve() / f'tmp_template_labels_{uuid4().hex[:8]}'
    temp_dir.mkdir(parents=True, exist_ok=True)
    input_path = temp_dir / f'template_labels_{uuid4().hex[:8]}.xlsx'
    schema_fields = {
        'required_fields': [{'name': 'item_name', 'title': 'Item Name', 'description': 'Title', 'group': 'product_identity'}],
        'optional_fields': [{'name': 'compatible_devices', 'title': 'Compatible Devices', 'description': 'Provide the devices that are compatible with this item.', 'group': 'other'}],
        'enum_fields': {},
        'field_groups': {},
    }
    definition = build_template_definition(schema_fields, 'INPUT_MOUSE', 'US', 'single')
    generate_template(definition, str(input_path))

    wb = load_workbook(input_path)
    ws = wb.active
    header_map = {str(ws.cell(row=2, column=idx).value): idx for idx in range(1, ws.max_column + 1)}
    ws.cell(row=3, column=header_map['sku'], value='SKU-1')
    ws.cell(row=3, column=header_map['product_type'], value='INPUT_MOUSE')
    ws.cell(row=3, column=header_map['item_name'], value='Demo Mouse')
    ws.cell(row=3, column=header_map['product_identity_mode'], value='gtin_exemption')
    wb.save(input_path)
    wb.close()

    class FakeListingsAPI:
        def put_listings_item(self, sku, product, preview=False):
            return {
                'status': 'INVALID',
                'issues': [{'severity': 'ERROR', 'message': 'Compatible Devices missing', 'attributeNames': ['compatible_devices'], 'categories': ['MISSING_ATTRIBUTE'], 'code': '90220'}],
            }

        def get_listings_item(self, sku):
            return None

    monkeypatch.setattr(web_app, '_load_template_definition_for_file', lambda filepath: ({
        'template_id': definition['template_id'],
        'product_type': 'INPUT_MOUSE',
        'variation_mode': 'single',
    }, definition))
    monkeypatch.setattr(web_app, '_build_listing_api_context', lambda account_id='': ({'name': 'US Test'}, FieldMapper('ATVPDKIKX0DER'), FakeListingsAPI(), ''))
    monkeypatch.setattr(FieldMapper, '_load_schema_fields', lambda self, product_type: schema_fields)
    monkeypatch.setattr(web_app, '_probe_product_media', lambda product: {'status': 'none', 'total': 0, 'passed': 0, 'failed': 0, 'checks': []})

    client = app.test_client()
    response = client.post('/api/template-diagnose', json={'file': str(input_path)})
    assert response.status_code == 200
    task_payload = _wait_for_task(client, response.get_json()['task_id'])
    result_entry = task_payload['result']['results'][0]
    assert result_entry['missing_fields'][0]['title'] == 'Compatible Devices'
    assert 'Compatible Devices' in result_entry['template']['blocking_issues'][0]


def test_template_definition_uses_official_dg_required_field():
    from core.template_generator import build_template_definition

    schema_fields = {
        'required_fields': [
            {'name': 'item_name', 'title': 'Item Name', 'description': 'Title', 'group': 'product_identity'},
            {
                'name': 'supplier_declared_dg_hz_regulation',
                'title': 'Dangerous Goods Regulations',
                'description': 'Hazmat regulation declaration',
                'group': 'compliance',
            },
        ],
        'optional_fields': [],
        'enum_fields': {'supplier_declared_dg_hz_regulation': ['not_applicable', 'ghs']},
        'field_groups': {},
    }

    definition = build_template_definition(schema_fields, 'INPUT_MOUSE', 'US', 'single')
    columns_by_key = {column['key']: column for column in definition['columns']}

    assert 'supplier_declared_dg_hz_regulation' in columns_by_key
    assert columns_by_key['supplier_declared_dg_hz_regulation']['level'] == 'required'
    assert columns_by_key['supplier_declared_dg_hz_regulation']['source_attribute'] == 'supplier_declared_dg_hz_regulation'
    assert 'hazmat_declaration' not in columns_by_key


def test_mapper_maps_dg_alias_and_official_field_to_schema_attribute():
    from amazon.mapper import FieldMapper

    mapper = FieldMapper('ATVPDKIKX0DER')

    # 新模板字段：直接透传到 schema 阻断字段
    attrs_official = mapper.build_listing_attributes({
        'sku': 'SKU-DG-1',
        'title': 'Demo Mouse',
        'main_image_url': 'https://example.com/demo.jpg',
        'batteries_required': 'Yes',
        'supplier_declared_dg_hz_regulation': 'ghs',
    })
    assert attrs_official['supplier_declared_dg_hz_regulation'][0]['value'] == 'ghs'

    # 兼容旧模板字段别名：hazmat_declaration 仍映射到同一阻断字段
    attrs_alias = mapper.build_listing_attributes({
        'sku': 'SKU-DG-2',
        'title': 'Demo Mouse',
        'main_image_url': 'https://example.com/demo.jpg',
        'batteries_required': 'Yes',
        'hazmat_declaration': 'not_applicable',
    })
    assert attrs_alias['supplier_declared_dg_hz_regulation'][0]['value'] == 'not_applicable'
    assert 'hazmat_declaration' not in attrs_alias


def test_mapper_skips_template_and_diagnostic_runtime_columns():
    from amazon.mapper import FieldMapper

    mapper = FieldMapper('ATVPDKIKX0DER')
    row = {
        'SKU': 'SKU-1',
        'item_name': 'Demo Mouse',
        'product_type': 'INPUT_MOUSE',
        'template_id': 'tpl_mouse',
        'template_blocking_issues': 'Compatible Devices missing',
        'listing_check_status': 'fail',
        'preview_status': 'INVALID',
        'submit_status': 'PENDING',
    }
    col_map = {
        'sku': 'SKU',
        'title': 'item_name',
        'product_type': 'product_type',
    }

    product = mapper.map_excel_row(row, col_map)

    assert 'template_id' not in product
    assert 'template_blocking_issues' not in product
    assert 'listing_check_status' not in product
    assert 'preview_status' not in product
    assert 'submit_status' not in product

def test_config_endpoint_round_trips_general_settings(monkeypatch):
    temp_dir = Path(web_config.OUTPUT_DIR).resolve() / f'tmp_config_test_{uuid4().hex[:8]}'
    temp_dir.mkdir(parents=True, exist_ok=True)
    env_path = temp_dir / '.env'
    env_path.write_text('', encoding='utf-8')
    tracked_keys = ['OUTPUT_DIR', 'DEFAULT_LANG', 'BATCH_LIMIT', 'AI_CONCURRENCY', 'IMAGE_CONCURRENCY']
    original_env = {key: os.environ.get(key) for key in tracked_keys}

    try:
        monkeypatch.setattr(web_app, '_env_file_path', lambda: str(env_path))

        client = app.test_client()
        response = client.post(
            '/api/config',
            json={
                'output_dir': 'custom-output',
                'default_lang': 'en',
                'batch_limit': 321,
                'ai_concurrency': 4,
                'image_concurrency': 6,
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True

        response = client.get('/api/config')
        assert response.status_code == 200
        data = response.get_json()
        assert data['default_lang'] == 'en'
        assert data['batch_limit'] == 321
        assert data['ai_concurrency'] == 4
        assert data['image_concurrency'] == 6
        assert data['output_dir'].endswith('custom-output')
        assert Path(data['output_dir']).is_absolute()

        env_text = env_path.read_text(encoding='utf-8')
        assert 'OUTPUT_DIR=custom-output' in env_text
        assert 'DEFAULT_LANG=en' in env_text
        assert 'BATCH_LIMIT=321' in env_text
        assert 'AI_CONCURRENCY=4' in env_text
        assert 'IMAGE_CONCURRENCY=6' in env_text
    finally:
        for key, value in original_env.items():
            if value is None:
                os.environ.pop(key, None)
            else:
                os.environ[key] = value
        from config import reload_config
        reload_config()
        if env_path.exists():
            env_path.unlink()
        if temp_dir.exists():
            temp_dir.rmdir()


def test_process_single_rewrite_returns_ai_fields_and_persists(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_process_single_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url'])
    ws.append(['SKU-1', 'Demo Product', 'https://example.com/demo.jpg'])
    wb.save(input_path)
    wb.close()

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))

        from stage1_pipeline import Stage1Pipeline
        monkeypatch.setattr(Stage1Pipeline, '_detect_product_type', lambda *args, **kwargs: 'PRODUCT')
        monkeypatch.setattr(Stage1Pipeline, '_ai_text', lambda *args, **kwargs: 'AI Title')

        client = app.test_client()
        response = client.post(
            '/api/process-single',
            json={
                'file': str(input_path),
                'sku': 'SKU-1',
                'field': 'title',
                'text_fields': ['title'],
                'action': 'rewrite',
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['result']['ai_title'] == 'AI Title'
        assert payload['result']['ai_status'] == 'completed'

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        assert 'AI标题' in headers
        assert 'AI状态' in headers
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        assert ws2.cell(row=2, column=header_map['AI标题']).value == 'AI Title'
        assert ws2.cell(row=2, column=header_map['AI状态']).value == 'completed'
        wb2.close()
    finally:
        if input_path.exists():
            input_path.unlink()


def test_products_endpoint_exposes_ai_fields(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_products_ai_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url', 'AI标题', 'AI商品描述', 'AI搜索关键词', 'AI卖点1'])
    ws.append(['SKU-1', 'Demo Product', 'https://example.com/demo.jpg', 'AI Title', 'AI Desc', 'kw1 kw2', 'AI Bullet 1'])
    wb.save(input_path)
    wb.close()

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))

        client = app.test_client()
        response = client.get('/api/products', query_string={'file': str(input_path)})

        assert response.status_code == 200
        payload = response.get_json()
        sku_entry = payload['products'][0]['skus'][0]
        assert sku_entry['ai_title'] == 'AI Title'
        assert sku_entry['ai_description'] == 'AI Desc'
        assert sku_entry['ai_keywords'] == 'kw1 kw2'
        assert sku_entry['ai_bullet_1'] == 'AI Bullet 1'
        assert sku_entry['ai_status'] == 'completed'
    finally:
        if input_path.exists():
            input_path.unlink()


def test_process_single_image_uses_unique_output_paths(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_process_single_image_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url'])
    ws.append(['SKU-1', 'Demo Product 1', 'https://example.com/demo-1.jpg'])
    ws.append(['SKU-2', 'Demo Product 2', 'https://example.com/demo-2.jpg'])
    wb.save(input_path)
    wb.close()

    generated_paths = []

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_config, 'OUTPUT_DIR', str(output_dir))

        from config import get_config
        cfg = get_config()
        monkeypatch.setattr(cfg, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(cfg, 'OUTPUT_DIR', str(output_dir))
        monkeypatch.setattr(cfg, 'AI_API_KEY', 'test-key')

        import base64
        import stage1_pipeline
        image_base64 = base64.b64encode(_build_test_image_base64()).decode('utf-8')
        monkeypatch.setattr(stage1_pipeline, 'ai_image_edit_url', lambda *args, **kwargs: image_base64)

        client = app.test_client()
        for sku in ('SKU-1', 'SKU-2'):
            response = client.post(
                '/api/process-single',
                json={
                    'file': str(input_path),
                    'sku': sku,
                    'action': 'image',
                },
            )
            assert response.status_code == 200
            assert response.get_json()['success'] is True

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        path_1 = ws2.cell(row=2, column=header_map['AI主图路径']).value
        path_2 = ws2.cell(row=3, column=header_map['AI主图路径']).value
        wb2.close()

        assert path_1
        assert path_2
        assert path_1 != path_2
        assert Path(path_1).exists()
        assert Path(path_2).exists()
        generated_paths.extend([path_1, path_2])
    finally:
        if input_path.exists():
            input_path.unlink()
        for generated_path in generated_paths:
            generated_file = Path(generated_path)
            if generated_file.exists():
                generated_file.unlink()


def test_process_single_image_all_persists_sub_images(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_process_single_image_all_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url', 'image_2', 'image_3'])
    ws.append([
        'SKU-1',
        'Demo Product',
        'https://example.com/main.jpg',
        'https://example.com/sub-2.jpg',
        'https://example.com/sub-3.jpg',
    ])
    wb.save(input_path)
    wb.close()

    generated_paths = []

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_config, 'OUTPUT_DIR', str(output_dir))

        from config import get_config
        cfg = get_config()
        monkeypatch.setattr(cfg, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(cfg, 'OUTPUT_DIR', str(output_dir))
        monkeypatch.setattr(cfg, 'AI_API_KEY', 'test-key')

        import base64
        import stage1_pipeline
        image_base64 = base64.b64encode(_build_test_image_base64()).decode('utf-8')
        monkeypatch.setattr(stage1_pipeline, 'ai_image_edit_url', lambda *args, **kwargs: image_base64)

        client = app.test_client()
        response = client.post(
            '/api/process-single',
            json={
                'file': str(input_path),
                'sku': 'SKU-1',
                'action': 'image',
                'scope': 'all',
                'bg_style': 'gradient',
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['result']['ai_status'] == 'completed'
        assert payload['result']['ai_image_2']
        assert payload['result']['ai_image_3']

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        for header in ('AI主图路径', 'AI副图2路径', 'AI副图3路径'):
            assert header in header_map
            path_value = ws2.cell(row=2, column=header_map[header]).value
            assert path_value
            generated_paths.append(path_value)
        wb2.close()

        products_response = client.get('/api/products', query_string={'file': str(input_path)})
        sku_entry = products_response.get_json()['products'][0]['skus'][0]
        assert sku_entry['ai_image_2']
        assert sku_entry['ai_image_3']
    finally:
        if input_path.exists():
            input_path.unlink()
        for generated_path in generated_paths:
            generated_file = Path(generated_path)
            if generated_file.exists():
                generated_file.unlink()


def test_process_single_image_persists_media_locator_fields(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_process_single_media_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url'])
    ws.append(['SKU-1', 'Demo Product', 'https://example.com/demo-1.jpg'])
    wb.save(input_path)
    wb.close()

    generated_paths = []

    class FakeMediaStore:
        def enabled(self):
            return True

        def upload_image(self, local_path, *, sku, slot, marketplace='US'):
            return type('UploadResult', (), {
                'to_dict': lambda self: {
                    'success': True,
                    'locator': f's3://demo-bucket/amazon28/{sku}/{slot}.jpg',
                    'preview_url': 'https://cdn.example.com/demo.jpg',
                    'provider': 's3',
                    'bucket': 'demo-bucket',
                    'key': f'amazon28/{sku}/{slot}.jpg',
                    'etag': 'etag-1',
                    'error': '',
                }
            })()

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_config, 'OUTPUT_DIR', str(output_dir))

        from config import get_config
        cfg = get_config()
        monkeypatch.setattr(cfg, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(cfg, 'OUTPUT_DIR', str(output_dir))
        monkeypatch.setattr(cfg, 'AI_API_KEY', 'test-key')
        monkeypatch.setattr(cfg, 'MEDIA_STORE_ENABLED', True)
        monkeypatch.setattr(cfg, 'MEDIA_STORE_PROVIDER', 's3')

        import base64
        import stage1_pipeline
        image_base64 = base64.b64encode(_build_test_image_base64()).decode('utf-8')
        monkeypatch.setattr(stage1_pipeline, 'ai_image_edit_url', lambda *args, **kwargs: image_base64)
        monkeypatch.setattr(stage1_pipeline, 'get_media_store', lambda: FakeMediaStore())

        client = app.test_client()
        response = client.post(
            '/api/process-single',
            json={
                'file': str(input_path),
                'sku': 'SKU-1',
                'action': 'image',
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['result']['ai_media_locator'] == 's3://demo-bucket/amazon28/SKU-1/main.jpg'
        assert payload['result']['ai_upload_status'] == 'uploaded'
        assert payload['result']['ai_main_image']

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        assert ws2.cell(row=2, column=header_map['AI主图URL']).value == 's3://demo-bucket/amazon28/SKU-1/main.jpg'
        assert ws2.cell(row=2, column=header_map['AI主图上传状态']).value == 'uploaded'
        generated_path = ws2.cell(row=2, column=header_map['AI主图路径']).value
        wb2.close()
        generated_paths.append(generated_path)

        products_response = client.get('/api/products', query_string={'file': str(input_path)})
        sku_entry = products_response.get_json()['products'][0]['skus'][0]
        assert sku_entry['ai_media_locator'] == 's3://demo-bucket/amazon28/SKU-1/main.jpg'
        assert sku_entry['ai_upload_status'] == 'uploaded'
        assert sku_entry['ai_main_image']
    finally:
        if input_path.exists():
            input_path.unlink()
        for generated_path in generated_paths:
            if generated_path and Path(generated_path).exists():
                Path(generated_path).unlink()


def test_update_product_endpoint_persists_multiple_fields_xlsx(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_update_product_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'UPC', 'item_weight'])
    ws.append(['SKU-1', 'Demo Product', '111111111111', '0.5'])
    wb.save(input_path)
    wb.close()

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))

        client = app.test_client()
        response = client.post(
            '/api/update-product',
            json={
                'file': str(input_path),
                'sku': 'SKU-1',
                'updates': {
                    'title': 'Updated Product',
                    'upc': '222222222222',
                    'item_weight': '1.25',
                    'color': 'Blue',
                },
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        assert ws2.cell(row=2, column=header_map['item_name']).value == 'Updated Product'
        assert ws2.cell(row=2, column=header_map['UPC']).value == '222222222222'
        assert ws2.cell(row=2, column=header_map['item_weight']).value == '1.25'
        assert ws2.cell(row=2, column=header_map['color']).value == 'Blue'
        wb2.close()
    finally:
        if input_path.exists():
            input_path.unlink()


def test_update_product_endpoint_persists_multiple_fields_xls(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_update_product_{uuid4().hex[:8]}.xls'
    input_path = input_dir / filename

    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')
    headers = ['SKU', 'item_name', 'UPC']
    values = ['SKU-1', 'Demo Product', '111111111111']
    for col_idx, header in enumerate(headers):
        ws.write(0, col_idx, header)
    for col_idx, value in enumerate(values):
        ws.write(1, col_idx, value)
    wb.save(str(input_path))

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))

        client = app.test_client()
        response = client.post(
            '/api/update-product',
            json={
                'file': str(input_path),
                'sku': 'SKU-1',
                'updates': {
                    'title': 'Updated XLS Product',
                    'upc': '333333333333',
                },
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True

        book = xlrd.open_workbook(str(input_path))
        sheet = book.sheet_by_index(0)
        headers_after = [sheet.cell_value(0, col_idx) for col_idx in range(sheet.ncols)]
        assert 'upc' not in headers_after
        header_map = {str(header): idx for idx, header in enumerate(headers_after)}
        assert sheet.cell_value(1, header_map['item_name']) == 'Updated XLS Product'
        assert sheet.cell_value(1, header_map['UPC']) == '333333333333'
    finally:
        if input_path.exists():
            input_path.unlink()


def test_persist_bulk_row_updates_xls_prefers_preserve_format(monkeypatch):
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_xls_preserve_{uuid4().hex[:8]}.xls'
    input_path = output_dir / filename

    wb = xlwt.Workbook()
    ws = wb.add_sheet('Sheet1')
    ws.write(0, 0, 'SKU')
    ws.write(0, 1, 'item_name')
    ws.write(1, 0, 'SKU-1')
    ws.write(1, 1, 'Demo Product')
    wb.save(str(input_path))

    called = {}

    try:
        def fake_preserve(file_path, updates):
            called['file_path'] = file_path
            called['updates'] = updates
            return True

        monkeypatch.setattr(web_app, '_persist_bulk_row_updates_xls_preserve_format', fake_preserve)
        monkeypatch.setattr(
            web_app,
            '_persist_bulk_row_updates_xls_rebuild',
            lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError('不应进入回退逻辑')),
        )

        web_app._persist_bulk_row_updates_xls(str(input_path), {'SKU-1': {'item_name': 'Updated'}})

        assert called['file_path'] == str(input_path)
        assert called['updates']['SKU-1']['item_name'] == 'Updated'
    finally:
        if input_path.exists():
            input_path.unlink()


def test_validate_endpoint_persists_results_to_excel(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_validate_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url'])
    ws.append(['SKU-1', 'Warn Product', 'https://example.com/1.jpg'])
    ws.append(['SKU-2', 'Fail Product', 'https://example.com/2.jpg'])
    wb.save(input_path)
    wb.close()

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))

        from amazon.mapper import FieldMapper

        def fake_validate(self, product, schema_fields=None):
            if product.get('sku') == 'SKU-1':
                return {
                    'valid': True,
                    'errors': [],
                    'warnings': ['建议填写: 品牌(brand)'],
                    'info': [],
                    'schema_required_missing': [],
                }
            return {
                'valid': False,
                'errors': ['缺少必填字段: 价格'],
                'warnings': ['建议填写: 品牌(brand)'],
                'info': [],
                'schema_required_missing': [{'name': 'price', 'title': '价格'}],
            }

        monkeypatch.setattr(FieldMapper, 'validate_required_fields', fake_validate)

        client = app.test_client()
        response = client.post(
            '/api/validate',
            json={'input_file': str(input_path), 'skus': ['SKU-1', 'SKU-2']},
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['total'] == 2
        assert payload['invalid'] == 1

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}

        assert ws2.cell(row=2, column=header_map['validation_status']).value == 'warn'
        assert ws2.cell(row=2, column=header_map['validation_warnings']).value == '建议填写: 品牌(brand)'
        assert ws2.cell(row=2, column=header_map['validation_errors']).value in ('', None)

        assert ws2.cell(row=3, column=header_map['validation_status']).value == 'fail'
        assert ws2.cell(row=3, column=header_map['validation_errors']).value == '缺少必填字段: 价格'
        assert ws2.cell(row=3, column=header_map['validation_warnings']).value == '建议填写: 品牌(brand)'
        wb2.close()
    finally:
        if input_path.exists():
            input_path.unlink()


def test_listing_check_endpoint_persists_diagnostics(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_listing_check_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url', 'product_type'])
    ws.append(['SKU-1', 'Demo Product', 'https://example.com/demo.jpg', 'HEADPHONES'])
    wb.save(input_path)
    wb.close()

    try:
        monkeypatch.setattr(web_app.time, 'sleep', lambda *_args, **_kwargs: None)

        from amazon.accounts import AccountManager
        from amazon.listings import ListingsAPI
        from amazon.mapper import FieldMapper

        monkeypatch.setattr(AccountManager, 'get_default_account', lambda self: {
            'name': 'US Check',
            'seller_id': 'SELLER-3',
            'marketplace_id': 'ATVPDKIKX0DER',
            'lwa_client_id': 'client',
            'lwa_client_secret': 'secret',
            'refresh_token': 'refresh',
        })
        monkeypatch.setattr(FieldMapper, '_load_schema_fields', lambda self, product_type: {
            'required_fields': [{'name': 'manufacturer', 'title': 'Manufacturer'}]
        })
        monkeypatch.setattr(FieldMapper, 'validate_required_fields', lambda self, product, schema_fields=None: {
            'valid': False,
            'errors': ['缺少类目必填字段: Manufacturer (manufacturer)'],
            'warnings': ['建议填写: 品牌(brand)'],
            'info': [],
            'schema_required_missing': [{'name': 'manufacturer', 'title': 'Manufacturer', 'group': 'identity'}],
        })
        monkeypatch.setattr(web_app, '_probe_product_media', lambda product: {
            'status': 'fail',
            'total': 1,
            'passed': 0,
            'failed': 1,
            'checks': [{
                'field': 'main_image_url',
                'url': product.get('main_image_url'),
                'ok': False,
                'status_code': 404,
                'content_type': 'text/plain',
                'message': 'HTTP 404',
            }],
        })
        monkeypatch.setattr(ListingsAPI, 'put_listings_item', lambda self, sku, product, preview=False: {
            'status': 'INVALID',
            'issues': [{
                'severity': 'ERROR',
                'code': '90220',
                'message': "'Manufacturer' is required but missing.",
                'attributeNames': ['manufacturer'],
                'categories': ['MISSING_ATTRIBUTE'],
            }],
        })
        monkeypatch.setattr(ListingsAPI, 'get_listings_item', lambda self, sku: {
            'summaries': [{'asin': 'B00CHECK123'}],
            'issues': [{
                'severity': 'WARNING',
                'code': 'W100',
                'message': 'Listing still processing',
                'attributeNames': ['manufacturer'],
            }],
        })

        client = app.test_client()
        response = client.post(
            '/api/listing-check',
            json={'file': str(input_path), 'skus': ['SKU-1']},
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['fail'] == 1
        assert payload['results'][0]['status'] == 'fail'
        assert payload['results'][0]['missing_fields'][0]['name'] == 'manufacturer'

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        assert ws2.cell(row=2, column=header_map['listing_check_status']).value == 'fail'
        assert '缺少字段' in str(ws2.cell(row=2, column=header_map['listing_check_summary']).value or '')
        assert 'manufacturer' in str(ws2.cell(row=2, column=header_map['listing_check_missing_fields']).value or '')
        assert 'HTTP 404' in str(ws2.cell(row=2, column=header_map['listing_check_issues']).value or '')
        assert ws2.cell(row=2, column=header_map['listing_check_account']).value == 'US Check'
        wb2.close()

        products_response = client.get('/api/products', query_string={'file': str(input_path)})
        sku_entry = products_response.get_json()['products'][0]['skus'][0]
        assert sku_entry['listing_check_status'] == 'fail'
        assert 'manufacturer' in sku_entry['listing_check_missing_fields']
    finally:
        if input_path.exists():
            input_path.unlink()


def test_export_endpoint_accepts_sp_api_alias_and_selected_skus(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    filename = f'test_export_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url', 'price'])
    ws.append(['SKU-1', 'Demo Product 1', 'https://example.com/1.jpg', '12.99'])
    ws.append(['SKU-2', 'Demo Product 2', 'https://example.com/2.jpg', '18.99'])
    wb.save(input_path)
    wb.close()

    exported_path = None

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_config, 'OUTPUT_DIR', str(output_dir))

        client = app.test_client()
        response = client.post(
            '/api/export',
            json={
                'file': str(input_path),
                'format': 'sp_api',
                'selected_skus': ['SKU-2'],
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['total_exported'] == 1

        exported_path = output_dir / payload['filename']
        assert exported_path.exists()

        wb2 = load_workbook(exported_path)
        ws2 = wb2.active
        assert ws2.max_row == 2
        assert ws2.cell(row=2, column=1).value == 'SKU-2'
        assert ws2.cell(row=2, column=3).value == 'Demo Product 2'
        wb2.close()
    finally:
        if input_path.exists():
            input_path.unlink()
        if exported_path and exported_path.exists():
            exported_path.unlink()


def test_submit_endpoint_persists_results_to_excel(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_submit_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url', 'standard_price'])
    ws.append(['SKU-1', 'Demo Product', 'https://example.com/demo.jpg', '19.99'])
    wb.save(input_path)
    wb.close()

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_app.time, 'sleep', lambda *_args, **_kwargs: None)

        from amazon.accounts import AccountManager
        from amazon.listings import ListingsAPI
        from amazon.mapper import FieldMapper

        monkeypatch.setattr(AccountManager, 'get_default_account', lambda self: {
            'name': 'US Test',
            'seller_id': 'SELLER-1',
            'marketplace_id': 'ATVPDKIKX0DER',
            'lwa_client_id': 'client',
            'lwa_client_secret': 'secret',
            'refresh_token': 'refresh',
        })
        monkeypatch.setattr(FieldMapper, 'validate_required_fields', lambda self, product, schema_fields=None: {
            'valid': True,
            'errors': [],
            'warnings': [],
            'info': [],
            'schema_required_missing': [],
        })
        monkeypatch.setattr(ListingsAPI, 'put_listings_item', lambda self, sku, product, preview=False: {
            'status': 'ACCEPTED',
            'submissionId': 'SUB-001',
            'issues': [{'severity': 'WARNING', 'code': 'W1', 'message': 'Need review'}],
            'identifiers': [{'asin': 'B00TEST123'}],
        })

        client = app.test_client()
        response = client.post(
            '/api/submit',
            json={
                'file': str(input_path),
                'skus': ['SKU-1'],
                'preview': False,
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['accepted'] == 1
        assert payload['persist_warning'] == ''

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        assert ws2.cell(row=2, column=header_map['submit_status']).value == 'ACCEPTED'
        assert ws2.cell(row=2, column=header_map['submission_id']).value == 'SUB-001'
        assert ws2.cell(row=2, column=header_map['asin']).value == 'B00TEST123'
        assert 'Need review' in str(ws2.cell(row=2, column=header_map['submit_message']).value or '')
        assert ws2.cell(row=2, column=header_map['submit_time']).value
        wb2.close()

        products_response = client.get('/api/products', query_string={'file': str(input_path)})
        sku_entry = products_response.get_json()['products'][0]['skus'][0]
        assert sku_entry['submit_status'] == 'ACCEPTED'
        assert sku_entry['asin'] == 'B00TEST123'
    finally:
        if input_path.exists():
            input_path.unlink()


def test_submit_endpoint_blocks_when_precheck_fails(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_submit_blocked_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url', 'standard_price'])
    ws.append(['SKU-1', 'Blocked Product', 'https://example.com/demo.jpg', '19.99'])
    wb.save(input_path)
    wb.close()

    submit_calls = {'count': 0}

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_app.time, 'sleep', lambda *_args, **_kwargs: None)

        from amazon.accounts import AccountManager
        from amazon.listings import ListingsAPI
        from amazon.mapper import FieldMapper

        monkeypatch.setattr(AccountManager, 'get_default_account', lambda self: {
            'name': 'US Block',
            'seller_id': 'SELLER-BLOCK',
            'marketplace_id': 'ATVPDKIKX0DER',
            'lwa_client_id': 'client',
            'lwa_client_secret': 'secret',
            'refresh_token': 'refresh',
        })
        monkeypatch.setattr(FieldMapper, 'validate_required_fields', lambda self, product, schema_fields=None: {
            'valid': True,
            'errors': [],
            'warnings': [],
            'info': [],
            'schema_required_missing': [],
        })
        monkeypatch.setattr(web_app, '_run_listing_check_for_product', lambda *args, **kwargs: {
            'sku': 'SKU-1',
            'status': 'fail',
            'summary_text': '模板必填缺失 1 项；预览 INVALID',
            'missing_fields': [{'name': 'manufacturer', 'title': 'Manufacturer', 'source': 'schema'}],
            'preview': {'status': 'INVALID', 'issues': []},
            'listing': {'issues': [], 'exists': False},
            'media': {'checks': [], 'failed': 0},
            'template': {'required_total': 5, 'required_filled': 4, 'required_missing': [{'key': 'manufacturer', 'label_zh': '制造商'}], 'recommended_missing': [], 'blocking_issues': ['制造商 (manufacturer)']},
        })

        def fake_submit(self, sku, product, preview=False):
            submit_calls['count'] += 1
            return {'status': 'ACCEPTED'}

        monkeypatch.setattr(ListingsAPI, 'put_listings_item', fake_submit)

        client = app.test_client()
        response = client.post('/api/submit', json={
            'file': str(input_path),
            'skus': ['SKU-1'],
            'preview': False,
        })

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['accepted'] == 0
        assert payload['failed'] == 1
        assert payload['results'][0]['status'] == 'PREVIEW_BLOCKED'
        assert submit_calls['count'] == 0

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        assert ws2.cell(row=2, column=header_map['submit_status']).value == 'PREVIEW_BLOCKED'
        wb2.close()
    finally:
        if input_path.exists():
            input_path.unlink()


def test_submit_preview_persists_results_to_excel(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_submit_preview_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url', 'standard_price'])
    ws.append(['SKU-1', 'Demo Product', 'https://example.com/demo.jpg', '19.99'])
    wb.save(input_path)
    wb.close()

    try:
        monkeypatch.setattr(web_config, 'INPUT_DIR', str(input_dir))
        monkeypatch.setattr(web_app.time, 'sleep', lambda *_args, **_kwargs: None)

        from amazon.accounts import AccountManager
        from amazon.listings import ListingsAPI
        from amazon.mapper import FieldMapper

        monkeypatch.setattr(AccountManager, 'get_default_account', lambda self: {
            'name': 'US Preview',
            'seller_id': 'SELLER-2',
            'marketplace_id': 'ATVPDKIKX0DER',
            'lwa_client_id': 'client',
            'lwa_client_secret': 'secret',
            'refresh_token': 'refresh',
        })
        monkeypatch.setattr(FieldMapper, 'validate_required_fields', lambda self, product, schema_fields=None: {
            'valid': True,
            'errors': [],
            'warnings': [],
            'info': [],
            'schema_required_missing': [],
        })
        monkeypatch.setattr(ListingsAPI, 'put_listings_item', lambda self, sku, product, preview=False: {
            'status': 'VALID',
            'issues': [{'severity': 'WARNING', 'message': 'Review title formatting'}],
        })

        client = app.test_client()
        response = client.post(
            '/api/submit',
            json={
                'file': str(input_path),
                'skus': ['SKU-1'],
                'preview': True,
            },
        )

        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        assert payload['valid'] == 1

        wb2 = load_workbook(input_path)
        ws2 = wb2.active
        headers = [str(cell.value) if cell.value is not None else '' for cell in ws2[1]]
        header_map = {header: idx + 1 for idx, header in enumerate(headers)}
        assert ws2.cell(row=2, column=header_map['preview_status']).value == 'VALID'
        assert 'Review title formatting' in str(ws2.cell(row=2, column=header_map['preview_message']).value or '')
        assert ws2.cell(row=2, column=header_map['preview_account']).value == 'US Preview'
        assert ws2.cell(row=2, column=header_map['preview_time']).value
        wb2.close()

        products_response = client.get('/api/products', query_string={'file': str(input_path)})
        sku_entry = products_response.get_json()['products'][0]['skus'][0]
        assert sku_entry['preview_status'] == 'VALID'
        assert sku_entry['preview_account'] == 'US Preview'
    finally:
        if input_path.exists():
            input_path.unlink()


def test_submit_task_endpoint_returns_task_and_result(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_submit_task_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url', 'standard_price'])
    ws.append(['SKU-1', 'Demo Product', 'https://example.com/demo.jpg', '19.99'])
    wb.save(input_path)
    wb.close()

    def fake_execute(task_id, input_file, skus, preview, account_id):
        web_app._complete_task_record(task_id, message='预览完成：1/1 可提交', result={
            'success': True,
            'mode': 'preview',
            'total': 1,
            'valid': 1,
            'invalid': 0,
            'results': [{'sku': 'SKU-1', 'valid': True, 'status': 'VALID', 'errors': [], 'warnings': []}],
            'message': '预校验完成: 1/1 个商品可提交',
        }, progress=1, total=1, current_item='完成')

    try:
        monkeypatch.setattr(web_app, '_execute_submit_task', fake_execute)
        client = app.test_client()
        response = client.post('/api/submit-task', json={
            'file': str(input_path),
            'skus': ['SKU-1'],
            'preview': True,
        })
        assert response.status_code == 200
        payload = response.get_json()
        assert payload['success'] is True
        task_payload = _wait_for_task(client, payload['task_id'])
        assert task_payload['status'] == 'completed'
        assert task_payload['result']['valid'] == 1
        assert task_payload['stage_name'] == '解析文件'
    finally:
        if input_path.exists():
            input_path.unlink()


def test_account_manager_test_connection_uses_listings_probe(monkeypatch):
    output_dir = Path(web_config.OUTPUT_DIR).resolve()
    accounts_path = output_dir / f'test_accounts_{uuid4().hex[:8]}.json'
    accounts_path.write_text(json.dumps({
        'accounts': [{
            'name': 'US Test',
            'seller_id': 'SELLER-1',
            'marketplace_id': 'ATVPDKIKX0DER',
            'marketplace_name': 'Amazon US',
            'lwa_client_id': 'client',
            'lwa_client_secret': 'secret',
            'refresh_token': 'refresh',
            'is_default': True,
            'enabled': True,
        }]
    }), encoding='utf-8')

    try:
        from amazon.accounts import AccountManager
        from amazon.auth import AmazonAuth
        from amazon.listings import ListingsAPI

        monkeypatch.setattr(AmazonAuth, 'get_access_token', lambda self: 'token')

        probe_calls = {}

        def fake_probe(self, probe_sku=None):
            probe_calls['seller_id'] = self.seller_id
            probe_calls['marketplace_id'] = self.marketplace_id
            return {'success': True, 'status_code': 404, 'message': 'Listings API 可访问'}

        monkeypatch.setattr(ListingsAPI, 'probe_connection', fake_probe)

        manager = AccountManager(str(accounts_path))
        result = manager.test_connection('SELLER-1')

        assert result['success'] is True
        assert result['probe_status'] == 404
        assert probe_calls['seller_id'] == 'SELLER-1'
        assert probe_calls['marketplace_id'] == 'ATVPDKIKX0DER'
    finally:
        if accounts_path.exists():
            accounts_path.unlink()


def test_self_check_endpoint_reports_core_checks(monkeypatch):
    from amazon.accounts import AccountManager
    import core.ai_client as ai_client

    class FakeMediaStore:
        def enabled(self):
            return True

        def healthcheck(self):
            return {
                'success': True,
                'message': 'S3 媒体存储可用',
                'bucket': 'demo-bucket',
                'prefix': 'amazon28',
            }

    monkeypatch.setattr(ai_client, 'ai_text', lambda *args, **kwargs: 'OK')
    monkeypatch.setattr(ai_client, 'ai_image_generate', lambda *args, **kwargs: 'base64-image')
    monkeypatch.setattr(AccountManager, 'get_default_account', lambda self: {'seller_id': 'SELLER-1', 'name': 'Default'})
    monkeypatch.setattr(AccountManager, 'test_connection', lambda self, seller_id=None: {'success': True, 'message': '连接成功'})
    monkeypatch.setattr(web_app, '_detect_xls_preserve_support', lambda: {'supported': True, 'app': 'Excel.Application'})
    monkeypatch.setattr(web_app, 'get_media_store', lambda: FakeMediaStore())

    client = app.test_client()
    response = client.post('/api/self-check', json={})

    assert response.status_code == 200
    payload = response.get_json()
    names = {check['name'] for check in payload['checks']}
    assert {'output_dir', 'text_ai', 'image_ai', 'amazon_account', 'xls_preserve', 'media_store'} <= names
    assert payload['status'] == 'pass'


def test_cancel_endpoint_updates_status():
    client = app.test_client()
    web_app._task_cancel_event.clear()
    web_app.update_status(
        running=True,
        stage=1,
        progress=1,
        total=3,
        current_item='SKU-1',
        logs=[],
        result_file=None,
        error=None,
        cancel_requested=False,
        cancelled=False,
    )

    try:
        response = client.post('/api/process/cancel')
        assert response.status_code == 200
        assert response.get_json()['success'] is True
        assert web_app._task_cancel_event.is_set() is True

        status_response = client.get('/api/status')
        status_payload = status_response.get_json()
        assert status_payload['status'] == 'cancelling'

        web_app.update_status(running=False, cancel_requested=False, cancelled=True)
        cancelled_response = client.get('/api/status')
        assert cancelled_response.get_json()['status'] == 'cancelled'
    finally:
        web_app._task_cancel_event.clear()
        web_app.update_status(
            running=False,
            stage=None,
            progress=0,
            total=0,
            current_item='',
            logs=[],
            result_file=None,
            error=None,
            cancel_requested=False,
            cancelled=False,
        )


def test_task_history_endpoint_includes_validate_record(monkeypatch):
    input_dir = Path(web_config.INPUT_DIR).resolve()
    filename = f'test_task_history_{uuid4().hex[:8]}.xlsx'
    input_path = input_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url'])
    ws.append(['SKU-1', 'Demo Product', 'https://example.com/demo.jpg'])
    wb.save(input_path)
    wb.close()

    with web_app._task_lock:
        snapshot = list(web_app.task_history)
        web_app.task_history.clear()

    try:
        from amazon.mapper import FieldMapper

        monkeypatch.setattr(FieldMapper, 'validate_required_fields', lambda self, product, schema_fields=None: {
            'valid': True,
            'errors': [],
            'warnings': [],
            'info': [],
            'schema_required_missing': [],
        })

        client = app.test_client()
        response = client.post('/api/validate', json={'input_file': str(input_path), 'skus': ['SKU-1']})
        assert response.status_code == 200

        tasks_response = client.get('/api/tasks')
        tasks_payload = tasks_response.get_json()
        assert any(task.get('kind') == 'validate' for task in tasks_payload['history'])
    finally:
        with web_app._task_lock:
            web_app.task_history.clear()
            web_app.task_history.extend(snapshot)
        if input_path.exists():
            input_path.unlink()
