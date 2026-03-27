import json
import threading
import time
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook

from stage1_pipeline import Stage1Pipeline


def _create_excel(path, prefix):
    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'item_name', 'main_image_url'])
    ws.append([f'{prefix}-1', f'{prefix} Title 1', 'https://example.com/1.jpg'])
    ws.append([f'{prefix}-2', f'{prefix} Title 2', 'https://example.com/2.jpg'])
    wb.save(path)


def test_stage1_resume_ignores_progress_from_different_input(monkeypatch):
    output_dir = Path('output').resolve()
    suffix = uuid4().hex[:8]
    input_a = output_dir / f'test_resume_input_a_{suffix}.xlsx'
    input_b = output_dir / f'test_resume_input_b_{suffix}.xlsx'
    result_file = output_dir / f'test_resume_result_{suffix}.xlsx'
    progress_file = output_dir / '.progress.json'
    original_progress = progress_file.read_text(encoding='utf-8') if progress_file.exists() else None

    try:
        _create_excel(input_a, 'A')
        _create_excel(input_b, 'B')

        progress_file.write_text(json.dumps({
            'last_completed': 1,
            'input': str(input_a.resolve()),
            'input_mtime': input_a.stat().st_mtime,
            'rows': None,
        }), encoding='utf-8')

        pipeline = Stage1Pipeline()
        pipeline.config.OUTPUT_DIR = str(output_dir)

        processed_skus = []
        monkeypatch.setattr(pipeline, '_detect_product_type', lambda *args: 'PRODUCT')
        monkeypatch.setattr(
            pipeline,
            '_generate_text_v2',
            lambda item, *args: processed_skus.append(item['SKU']) or item.__setitem__('AI标题', 'ok'),
        )
        monkeypatch.setattr(pipeline, '_process_images', lambda *args: None)
        monkeypatch.setattr(pipeline.excel, 'write_comparison_output', lambda *args: None)

        pipeline.run(
            input_file=str(input_b),
            output_file=str(result_file),
            process_images=False,
            process_text=True,
            resume=True,
        )

        assert processed_skus == ['B-1', 'B-2']
    finally:
        for path in (input_a, input_b, result_file):
            if path.exists():
                path.unlink()
        if original_progress is None:
            if progress_file.exists():
                progress_file.unlink()
        else:
            progress_file.write_text(original_progress, encoding='utf-8')


def test_stage1_respects_text_concurrency(monkeypatch):
    output_dir = Path('output').resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = uuid4().hex[:8]
    input_file = output_dir / f'test_text_concurrency_{suffix}.xlsx'
    result_file = output_dir / f'test_text_concurrency_result_{suffix}.xlsx'

    try:
        _create_excel(input_file, 'TXT')
        wb = Workbook()
        ws = wb.active
        ws.append(['SKU', 'item_name', 'main_image_url'])
        for idx in range(1, 5):
            ws.append([f'TXT-{idx}', f'Text Title {idx}', f'https://example.com/{idx}.jpg'])
        wb.save(input_file)

        pipeline = Stage1Pipeline()
        pipeline.config.OUTPUT_DIR = str(output_dir)
        pipeline.config.AI_CONCURRENCY = 2
        pipeline.config.IMAGE_CONCURRENCY = 1

        state = {'active': 0, 'max_active': 0}
        lock = threading.Lock()

        monkeypatch.setattr(pipeline, '_detect_product_type', lambda *args: 'PRODUCT')

        def fake_generate(item, *args, **kwargs):
            with lock:
                state['active'] += 1
                state['max_active'] = max(state['max_active'], state['active'])
            time.sleep(0.06)
            item['AI标题'] = 'ok'
            with lock:
                state['active'] -= 1
            return True

        monkeypatch.setattr(pipeline, '_generate_text_v2', fake_generate)
        monkeypatch.setattr(pipeline, '_process_images', lambda *args, **kwargs: False)
        monkeypatch.setattr(pipeline, '_save_progress', lambda *args, **kwargs: None)
        monkeypatch.setattr(pipeline.excel, 'write_comparison_output', lambda *args, **kwargs: None)

        pipeline.run(
            input_file=str(input_file),
            output_file=str(result_file),
            process_images=False,
            process_text=True,
        )

        assert state['max_active'] == 2
    finally:
        for path in (input_file, result_file):
            if path.exists():
                path.unlink()


def test_stage1_respects_image_concurrency(monkeypatch):
    output_dir = Path('output').resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = uuid4().hex[:8]
    input_file = output_dir / f'test_image_concurrency_{suffix}.xlsx'
    result_file = output_dir / f'test_image_concurrency_result_{suffix}.xlsx'

    try:
        wb = Workbook()
        ws = wb.active
        ws.append(['SKU', 'item_name', 'main_image_url'])
        for idx in range(1, 6):
            ws.append([f'IMG-{idx}', f'Image Title {idx}', f'https://example.com/{idx}.jpg'])
        wb.save(input_file)

        pipeline = Stage1Pipeline()
        pipeline.config.OUTPUT_DIR = str(output_dir)
        pipeline.config.AI_CONCURRENCY = 1
        pipeline.config.IMAGE_CONCURRENCY = 3

        state = {'active': 0, 'max_active': 0}
        lock = threading.Lock()

        def fake_images(item, *args, **kwargs):
            with lock:
                state['active'] += 1
                state['max_active'] = max(state['max_active'], state['active'])
            time.sleep(0.06)
            item['AI主图路径'] = str(output_dir / f"{item['SKU']}.jpg")
            with lock:
                state['active'] -= 1
            return True

        monkeypatch.setattr(pipeline, '_process_images', fake_images)
        monkeypatch.setattr(pipeline, '_save_progress', lambda *args, **kwargs: None)
        monkeypatch.setattr(pipeline.excel, 'write_comparison_output', lambda *args, **kwargs: None)

        pipeline.run(
            input_file=str(input_file),
            output_file=str(result_file),
            process_images=True,
            process_text=False,
        )

        assert state['max_active'] == 3
    finally:
        for path in (input_file, result_file):
            if path.exists():
                path.unlink()
