"""
亚马逊商品处理工具 - Web界面
Flask后端 + 简洁前端
"""
import os
import sys
import json
import re
import subprocess
import tempfile
import time
import logging
import threading
import requests
from datetime import datetime
from uuid import uuid4
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import get_config, reload_config
from core.excel.processor import ExcelProcessor
from core.media_store import get_media_store
from core.template_service import (
    DEFAULT_MARKETPLACE,
    ensure_template_definition,
    ensure_template_workbook,
    evaluate_template_row,
    load_template_definition,
    read_template_metadata,
    recommend_product_types,
    summarize_template_issues,
    template_definition_summary,
    update_template_overlay,
)

logger = logging.getLogger(__name__)

app = Flask(__name__,
            template_folder='templates',
            static_folder='static')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

config = get_config()


# ===== 统一响应格式 =====

def ok(data=None, message=''):
    """统一成功响应"""
    resp = {'success': True}
    if message:
        resp['message'] = message
    if data is not None:
        resp.update(data)
    return jsonify(resp)


def err(message, code=400):
    """统一错误响应"""
    return jsonify({'success': False, 'error': message}), code


def _env_file_path() -> str:
    return os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env')


def _write_env_updates(updates: dict):
    """更新 .env 并同步到当前进程环境。"""
    env_path = _env_file_path()
    lines = []
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

    new_lines = []
    updated_keys = set()
    for line in lines:
        key = line.split('=')[0].strip() if '=' in line else ''
        if key in updates:
            new_lines.append(f"{key}={updates[key]}\n")
            updated_keys.add(key)
        else:
            new_lines.append(line)

    for key, val in updates.items():
        if key not in updated_keys:
            new_lines.append(f"{key}={val}\n")

    with open(env_path, 'w', encoding='utf-8') as f:
        f.writelines(new_lines)

    for key, val in updates.items():
        os.environ[str(key)] = str(val)


# ===== Excel 文件读取缓存 =====

_excel_cache = {}
_excel_cache_lock = threading.Lock()
_EXCEL_CACHE_MAX = 5


def read_excel_cached(filepath):
    """以 (filepath, mtime) 为 key 缓存 Excel 读取结果"""
    mtime = os.path.getmtime(filepath)
    key = (filepath, mtime)
    with _excel_cache_lock:
        if key in _excel_cache:
            return _excel_cache[key]
    processor = ExcelProcessor()
    data = processor.read_input(filepath)
    col_map = processor.detect_columns()
    headers = list(processor.headers)
    result = {'data': data, 'col_map': col_map, 'headers': headers}
    with _excel_cache_lock:
        # Evict oldest if cache full
        if len(_excel_cache) >= _EXCEL_CACHE_MAX:
            oldest_key = next(iter(_excel_cache))
            del _excel_cache[oldest_key]
        _excel_cache[key] = result
    return result


def invalidate_excel_cache(filepath=None):
    """写入操作后清除对应缓存"""
    with _excel_cache_lock:
        if filepath:
            keys_to_remove = [k for k in _excel_cache if k[0] == filepath]
            for k in keys_to_remove:
                del _excel_cache[k]
        else:
            _excel_cache.clear()


def _detect_header_row(ws) -> int:
    """自动检测表头行。"""
    for row_idx in range(1, min(10, ws.max_row + 1)):
        row_values = [str(cell.value or '').strip() for cell in ws[row_idx]]
        if any(ind in row_values for ind in ['SKU', 'sku', 'item_name', 'product_id', 'brand_name']):
            return row_idx
    return 1


def _build_header_index(ws, header_row: int):
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        if value is None:
            continue
        headers[str(value).strip()] = col_idx
    return headers


def _ensure_header(ws, header_row: int, headers: dict, header_name: str) -> int:
    """获取列索引，不存在则在表头末尾创建。"""
    if header_name in headers:
        return headers[header_name]

    col_idx = ws.max_column + 1
    ws.cell(row=header_row, column=col_idx, value=header_name)
    headers[header_name] = col_idx
    return col_idx


def _persist_row_updates(input_file: str, sku: str, updates: dict):
    """按 SKU 批量更新 Excel 行字段，缺失列自动创建。"""
    _persist_bulk_row_updates(input_file, {sku: updates})


def _persist_bulk_row_updates(input_file: str, updates_by_sku: dict):
    """批量更新多个 SKU，xlsx 保持原表结构，xls 使用 xlwt 重写。"""
    ext = os.path.splitext(str(input_file))[1].lower()
    if ext == '.xls':
        _persist_bulk_row_updates_xls(input_file, updates_by_sku)
    else:
        _persist_bulk_row_updates_xlsx(input_file, updates_by_sku)

    invalidate_excel_cache(input_file)


def _persist_bulk_row_updates_xlsx(input_file: str, updates_by_sku: dict):
    from openpyxl import load_workbook

    wb = load_workbook(input_file)
    ws = wb.active
    header_row = _detect_header_row(ws)
    headers = _build_header_index(ws, header_row)

    sku_col_idx = None
    for candidate in ('SKU', 'sku', 'seller_sku', 'seller-sku'):
        if candidate in headers:
            sku_col_idx = headers[candidate]
            break

    if sku_col_idx is None:
        wb.close()
        raise ValueError('找不到SKU列')

    row_map = {}
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell_sku = str(ws.cell(row=row_idx, column=sku_col_idx).value or '').strip()
        if cell_sku:
            row_map[cell_sku] = row_idx

    missing_skus = [sku for sku in updates_by_sku if sku not in row_map]
    if missing_skus:
        wb.close()
        raise ValueError(f"SKU未找到: {', '.join(missing_skus)}")

    for sku, updates in updates_by_sku.items():
        target_row_idx = row_map[sku]
        for header_name, value in updates.items():
            col_idx = _ensure_header(ws, header_row, headers, header_name)
            ws.cell(row=target_row_idx, column=col_idx, value=value)

    wb.save(input_file)
    wb.close()


def _persist_bulk_row_updates_xls(input_file: str, updates_by_sku: dict):
    if _persist_bulk_row_updates_xls_preserve_format(input_file, updates_by_sku):
        return
    logger.warning("`.xls` 保样式写回不可用，回退到重建工作簿模式，原样式/公式可能丢失")
    _persist_bulk_row_updates_xls_rebuild(input_file, updates_by_sku)


def _persist_bulk_row_updates_xls_preserve_format(input_file: str, updates_by_sku: dict) -> bool:
    payload_path = ''
    script_path = ''
    try:
        with tempfile.NamedTemporaryFile('w', encoding='utf-8-sig', suffix='.json', delete=False) as payload_file:
            payload_path = payload_file.name
            json.dump({
                'input_file': os.path.abspath(input_file),
                'updates_by_sku': updates_by_sku,
            }, payload_file, ensure_ascii=False)

        script = r"""
param([string]$PayloadPath)
$ErrorActionPreference = 'Stop'

function Get-CellText([object]$value) {
    if ($null -eq $value) { return '' }
    return ([string]$value).Trim()
}

function Get-HeaderRow($sheet, [int]$lastRow, [int]$lastCol) {
    $indicators = @('SKU', 'sku', 'item_name', 'product_id', 'brand_name', 'standard_price', '商品编号', '标题', '品牌', 'ASIN', 'title', 'brand', 'price')
    $scanTo = [Math]::Min($lastRow, 10)
    for ($row = 1; $row -le $scanTo; $row++) {
        $values = @()
        for ($col = 1; $col -le $lastCol; $col++) {
            $values += Get-CellText $sheet.Cells.Item($row, $col).Value2
        }
        foreach ($indicator in $indicators) {
            if ($values -contains $indicator) {
                return $row
            }
        }
    }
    return 1
}

function Copy-Formats($excel, $fromCell, $toCell) {
    if ($null -eq $fromCell -or $null -eq $toCell) { return }
    $fromCell.Copy() | Out-Null
    $toCell.PasteSpecial(-4122) | Out-Null
    $excel.CutCopyMode = 0
}

function Invoke-WorkbookUpdate([string]$progId, $payload) {
    $app = $null
    $workbook = $null
    $worksheet = $null
    try {
        $app = New-Object -ComObject $progId
        $app.Visible = $false
        $app.DisplayAlerts = $false
        $workbook = $app.Workbooks.Open($payload.input_file)
        $worksheet = $workbook.Worksheets.Item(1)

        $usedRange = $worksheet.UsedRange
        $lastRow = [int]($usedRange.Row + $usedRange.Rows.Count - 1)
        $lastCol = [int]($usedRange.Column + $usedRange.Columns.Count - 1)
        if ($lastRow -lt 1 -or $lastCol -lt 1) {
            throw 'Excel内容为空'
        }

        $headerRow = Get-HeaderRow $worksheet $lastRow $lastCol
        $headers = @{}
        for ($col = 1; $col -le $lastCol; $col++) {
            $headerText = Get-CellText $worksheet.Cells.Item($headerRow, $col).Value2
            if ($headerText) {
                $headers[$headerText] = $col
            }
        }

        $skuCol = $null
        foreach ($candidate in @('SKU', 'sku', 'seller_sku', 'seller-sku')) {
            if ($headers.ContainsKey($candidate)) {
                $skuCol = [int]$headers[$candidate]
                break
            }
        }
        if ($null -eq $skuCol) {
            throw '找不到SKU列'
        }

        $rowMap = @{}
        for ($row = $headerRow + 1; $row -le $lastRow; $row++) {
            $sku = Get-CellText $worksheet.Cells.Item($row, $skuCol).Value2
            if ($sku) {
                $rowMap[$sku] = $row
            }
        }

        $missingSkus = New-Object System.Collections.Generic.List[string]
        foreach ($skuProp in $payload.updates_by_sku.PSObject.Properties) {
            if (-not $rowMap.ContainsKey($skuProp.Name)) {
                $missingSkus.Add($skuProp.Name)
            }
        }
        if ($missingSkus.Count -gt 0) {
            throw ('SKU未找到: ' + ($missingSkus -join ', '))
        }

        foreach ($skuProp in $payload.updates_by_sku.PSObject.Properties) {
            $targetRow = [int]$rowMap[$skuProp.Name]
            foreach ($updateProp in $skuProp.Value.PSObject.Properties) {
                $headerName = [string]$updateProp.Name
                if (-not $headers.ContainsKey($headerName)) {
                    $sourceCol = [Math]::Max(1, $lastCol)
                    $lastCol++
                    $headers[$headerName] = $lastCol
                    Copy-Formats $app $worksheet.Cells.Item($headerRow, $sourceCol) $worksheet.Cells.Item($headerRow, $lastCol)
                    $worksheet.Columns.Item($lastCol).ColumnWidth = $worksheet.Columns.Item($sourceCol).ColumnWidth
                    $worksheet.Cells.Item($headerRow, $lastCol).Value2 = $headerName
                    if ($lastRow -gt $headerRow) {
                        for ($row = $headerRow + 1; $row -le $lastRow; $row++) {
                            Copy-Formats $app $worksheet.Cells.Item($row, $sourceCol) $worksheet.Cells.Item($row, $lastCol)
                        }
                    }
                }
                $targetCol = [int]$headers[$headerName]
                $worksheet.Cells.Item($targetRow, $targetCol).Value2 = $updateProp.Value
            }
        }

        $workbook.Save()
        $workbook.Close($false)
        $app.Quit()
        return $true
    } finally {
        if ($worksheet -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($worksheet) }
        if ($workbook -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) }
        if ($app -ne $null) { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($app) }
        [GC]::Collect()
        [GC]::WaitForPendingFinalizers()
    }
}

$payload = Get-Content -Raw -LiteralPath $PayloadPath | ConvertFrom-Json -Depth 20
$progIds = @('Excel.Application', 'Ket.Application')
$lastError = ''
foreach ($progId in $progIds) {
    try {
        if (Invoke-WorkbookUpdate $progId $payload) {
            exit 0
        }
    } catch {
        $lastError = $_.Exception.Message
    }
}

if ($lastError) {
    Write-Error $lastError
} else {
    Write-Error 'No compatible spreadsheet COM application found.'
}
exit 1
"""

        with tempfile.NamedTemporaryFile('w', encoding='utf-8-sig', suffix='.ps1', delete=False) as script_file:
            script_path = script_file.name
            script_file.write(script)

        result = subprocess.run(
            ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass', '-File', script_path, payload_path],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='ignore',
            timeout=120,
        )
        if result.returncode == 0:
            return True

        detail = (result.stderr or result.stdout or '').strip()
        if detail:
            logger.warning(f'`.xls` 保样式写回失败: {detail}')
        return False
    except Exception as e:
        logger.warning(f'`.xls` 保样式写回异常: {e}')
        return False
    finally:
        for temp_path in (payload_path, script_path):
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass


def _persist_bulk_row_updates_xls_rebuild(input_file: str, updates_by_sku: dict):
    import xlrd
    import xlwt

    workbook = xlrd.open_workbook(input_file)
    sheet = workbook.sheet_by_index(0)
    rows = [
        [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]
        for row_idx in range(sheet.nrows)
    ]

    if not rows:
        raise ValueError('Excel内容为空')

    header_row = None
    indicators = ['SKU', 'sku', 'item_name', 'product_id', 'brand_name',
                  'standard_price', '商品编号', '标题', '品牌', 'ASIN', 'title', 'brand', 'price']
    for row_idx, row_values in enumerate(rows[:10], start=1):
        normalized = [str(value or '').strip() for value in row_values]
        if any(indicator in normalized for indicator in indicators):
            header_row = row_idx
            break
    if header_row is None:
        header_row = 1

    header_idx = header_row - 1
    headers = {
        str(value).strip(): idx + 1
        for idx, value in enumerate(rows[header_idx])
        if str(value or '').strip()
    }

    sku_col_idx = None
    for candidate in ('SKU', 'sku', 'seller_sku', 'seller-sku'):
        if candidate in headers:
            sku_col_idx = headers[candidate]
            break
    if sku_col_idx is None:
        raise ValueError('找不到SKU列')

    row_map = {}
    for row_idx in range(header_idx + 1, len(rows)):
        sku = str(rows[row_idx][sku_col_idx - 1] if sku_col_idx - 1 < len(rows[row_idx]) else '').strip()
        if sku:
            row_map[sku] = row_idx

    missing_skus = [sku for sku in updates_by_sku if sku not in row_map]
    if missing_skus:
        raise ValueError(f"SKU未找到: {', '.join(missing_skus)}")

    max_cols = max((len(row) for row in rows), default=0)
    for row in rows:
        if len(row) < max_cols:
            row.extend([''] * (max_cols - len(row)))

    for sku, updates in updates_by_sku.items():
        target_row_idx = row_map[sku]
        for header_name, value in updates.items():
            if header_name not in headers:
                max_cols += 1
                headers[header_name] = max_cols
                rows[header_idx].append(header_name)
                for row_idx, row in enumerate(rows):
                    if row_idx != header_idx:
                        row.append('')
            rows[target_row_idx][headers[header_name] - 1] = value

    new_workbook = xlwt.Workbook()
    new_sheet = new_workbook.add_sheet(sheet.name or 'Sheet1')
    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row):
            new_sheet.write(row_idx, col_idx, value)
    new_workbook.save(input_file)


def _detect_xls_preserve_support() -> dict:
    """检测本机是否有可用于保样式写回 .xls 的电子表格 COM 应用。"""
    script = r"""
$ErrorActionPreference = 'Stop'

function Test-App([string]$progId) {
    $app = $null
    try {
        $app = New-Object -ComObject $progId
        $app.Visible = $false
        if ($app.PSObject.Methods.Name -contains 'Quit') {
            $app.Quit()
        }
        return $true
    } catch {
        return $false
    } finally {
        if ($app -ne $null) {
            try { [void][System.Runtime.InteropServices.Marshal]::ReleaseComObject($app) } catch {}
        }
        [GC]::Collect()
        [GC]::WaitForPendingFinalizers()
    }
}

if (Test-App 'Excel.Application') {
    Write-Output 'Excel.Application'
    exit 0
}
if (Test-App 'Ket.Application') {
    Write-Output 'Ket.Application'
    exit 0
}
Write-Output ''
exit 1
"""
    try:
        result = subprocess.run(
            ['powershell', '-NoProfile', '-ExecutionPolicy', 'Bypass', '-Command', script],
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='ignore',
            timeout=20,
        )
        app_name = (result.stdout or '').strip()
        return {
            'supported': result.returncode == 0 and bool(app_name),
            'app': app_name,
        }
    except Exception as e:
        logger.warning(f'检测 .xls 保格式写回能力失败: {e}')
        return {'supported': False, 'app': '', 'error': str(e)}


def _map_regenerate_field(field: str):
    mapping = {
        'title': ['title'],
        'bullet_1': ['bullets'],
        'bullet_2': ['bullets'],
        'bullet_3': ['bullets'],
        'bullet_4': ['bullets'],
        'bullet_5': ['bullets'],
        'bullets': ['bullets'],
        'description': ['description'],
        'keywords': ['keywords'],
        'all': ['title', 'bullets', 'description', 'keywords'],
    }
    return mapping.get(str(field or '').strip(), ['title', 'bullets', 'description', 'keywords'])


def _logical_field_to_excel_header(field: str, col_map: dict):
    """将逻辑字段名映射为 Excel 中的实际列名。"""
    field = str(field or '').strip()
    direct_map = {
        'item_name': ('title', 'item_name'),
        'title': ('title', 'item_name'),
        'bullet_point_1': ('bullet_point_1', 'bullet_point_1'),
        'bullet_point_2': ('bullet_point_2', 'bullet_point_2'),
        'bullet_point_3': ('bullet_point_3', 'bullet_point_3'),
        'bullet_point_4': ('bullet_point_4', 'bullet_point_4'),
        'bullet_point_5': ('bullet_point_5', 'bullet_point_5'),
        'product_description': ('description', 'product_description'),
        'description': ('description', 'product_description'),
        'generic_keywords': ('keywords', 'generic_keywords'),
        'keywords': ('keywords', 'generic_keywords'),
        'main_image_url': ('image_url', 'main_image_url'),
        'price': ('price', 'price'),
        'standard_price': ('price', 'standard_price'),
        'quantity': ('quantity', 'quantity'),
        'brand': ('brand', 'brand'),
        'product_type': ('product_type', 'product_type'),
        'condition_type': ('condition_type', 'condition_type'),
        'fulfillment_channel': ('fulfillment_channel', 'fulfillment_channel'),
        'color': ('color', 'color'),
        'size': ('size', 'size'),
        'material': ('material', 'material'),
        'manufacturer': ('manufacturer', 'manufacturer'),
        'model_number': ('model_number', 'model_number'),
        'country_of_origin': ('country_of_origin', 'country_of_origin'),
        'upc': ('upc', 'upc'),
        'ean': ('upc', 'ean'),
        'gtin': ('upc', 'gtin'),
        'item_weight': ('weight', 'item_weight'),
        'item_weight_unit': ('item_weight_unit', 'item_weight_unit'),
        'item_length': ('item_length', 'item_length'),
        'item_width': ('item_width', 'item_width'),
        'item_height': ('item_height', 'item_height'),
        'dimension_unit': ('dimension_unit', 'dimension_unit'),
        'package_weight': ('package_weight', 'package_weight'),
        'package_weight_unit': ('package_weight_unit', 'package_weight_unit'),
        'package_length': ('package_length', 'package_length'),
        'package_width': ('package_width', 'package_width'),
        'package_height': ('package_height', 'package_height'),
        'batteries_required': ('batteries_required', 'batteries_required'),
        'batteries_included': ('batteries_included', 'batteries_included'),
        'battery_type': ('battery_type', 'battery_type'),
        'number_of_batteries': ('number_of_batteries', 'number_of_batteries'),
        'battery_cell_composition': ('battery_cell_composition', 'battery_cell_composition'),
        'lithium_battery_packaging': ('lithium_battery_packaging', 'lithium_battery_packaging'),
        'lithium_battery_energy_content': ('lithium_battery_energy_content', 'lithium_battery_energy_content'),
        'lithium_battery_weight': ('lithium_battery_weight', 'lithium_battery_weight'),
        'hazmat_declaration': ('hazmat_declaration', 'hazmat_declaration'),
        'cpsia_cautionary_statement': ('cpsia_cautionary_statement', 'cpsia_cautionary_statement'),
        'safety_warning': ('safety_warning', 'safety_warning'),
        'legal_disclaimer': ('legal_disclaimer', 'legal_disclaimer'),
        'parent_sku': ('parent_sku', 'parent_sku'),
        'parentage_level': ('parentage_level', 'parentage_level'),
        'variation_theme': ('variation_theme', 'variation_theme'),
        'target_gender': ('target_gender', 'target_gender'),
        'age_range': ('age_range', 'age_range'),
        'department': ('department', 'department'),
        'currency': ('currency', 'currency'),
        'sale_price': ('sale_price', 'sale_price'),
        'sale_from_date': ('sale_from_date', 'sale_from_date'),
        'sale_end_date': ('sale_end_date', 'sale_end_date'),
        'list_price': ('list_price', 'list_price'),
        'max_order_quantity': ('max_order_quantity', 'max_order_quantity'),
        'handling_time': ('handling_time', 'handling_time'),
        'merchant_shipping_group_name': ('merchant_shipping_group_name', 'merchant_shipping_group_name'),
        'product_tax_code': ('product_tax_code', 'product_tax_code'),
        'item_type_keyword': ('item_type_keyword', 'item_type_keyword'),
        'external_product_id_type': ('external_product_id_type', 'external_product_id_type'),
        'asin': ('asin', 'asin'),
        'preview_status': ('preview_status', 'preview_status'),
        'preview_message': ('preview_message', 'preview_message'),
        'preview_time': ('preview_time', 'preview_time'),
        'preview_account': ('preview_account', 'preview_account'),
        'submit_status': ('submit_status', 'submit_status'),
        'submission_id': ('submission_id', 'submission_id'),
        'submit_time': ('submit_time', 'submit_time'),
        'submit_message': ('submit_message', 'submit_message'),
        'product_identity_mode': ('product_identity_mode', 'product_identity_mode'),
        'number_of_items': ('number_of_items', 'number_of_items'),
        'style_name': ('style_name', 'style_name'),
        'pattern_type': ('pattern_type', 'pattern_type'),
        'closure_type': ('closure_type', 'closure_type'),
        'unit_count': ('unit_count', 'unit_count'),
        'unit_count_type': ('unit_count_type', 'unit_count_type'),
    }

    if field in (
        'AI标题', 'AI商品描述', 'AI搜索关键词', 'AI状态',
        'AI主图路径', 'AI主图URL', 'AI主图预览URL', 'AI主图上传状态', 'AI主图上传错误',
    ):
        return field
    if field.startswith('AI卖点'):
        return field
    if field.startswith('AI副图') and any(
        field.endswith(suffix) for suffix in ('路径', 'URL', '预览URL', '上传状态', '上传错误')
    ):
        return field

    mapping = direct_map.get(field)
    if mapping:
        col_key, fallback = mapping
        return col_map.get(col_key) or fallback
    return col_map.get(field) or field


def _derive_ai_status(item: dict, ai_fields: list):
    explicit = str(item.get('ai_status', '') or item.get('AI状态', '') or '').strip()
    if explicit:
        return explicit
    return 'completed' if any(str(item.get(field, '') or '').strip() for field in ai_fields) else 'PENDING'


def _build_output_image_preview_url(local_path: str) -> str:
    local_path = str(local_path or '').strip()
    if not local_path or not os.path.exists(local_path):
        return ''
    return f"/api/output-image/{os.path.basename(local_path)}"


def _is_http_url(value: str) -> bool:
    text = str(value or '').strip().lower()
    return text.startswith('http://') or text.startswith('https://')


def _is_media_locator(value: str) -> bool:
    text = str(value or '').strip().lower()
    return text.startswith('s3://') or _is_http_url(text)


def _resolve_ai_media_locator(item: dict, slot='main') -> str:
    field_name = 'AI主图URL' if slot == 'main' else f'AI副图{slot}URL'
    ai_locator = str(item.get(field_name, '') or '').strip()
    if _is_media_locator(ai_locator):
        return ai_locator
    return ''


def _resolve_ai_public_image_url(item: dict, slot='main') -> str:
    locator = _resolve_ai_media_locator(item, slot)
    if _is_http_url(locator):
        return locator
    return ''


def _resolve_ai_preview_image_url(item: dict, slot='main') -> str:
    preview_field = 'AI主图预览URL' if slot == 'main' else f'AI副图{slot}预览URL'
    preview_url = str(item.get(preview_field, '') or '').strip()
    if _is_http_url(preview_url) or preview_url.startswith('/'):
        return preview_url

    path_field = 'AI主图路径' if slot == 'main' else f'AI副图{slot}路径'
    return _build_output_image_preview_url(item.get(path_field, ''))


def _resolve_ai_upload_state(item: dict, slot='main') -> dict:
    status_field = 'AI主图上传状态' if slot == 'main' else f'AI副图{slot}上传状态'
    error_field = 'AI主图上传错误' if slot == 'main' else f'AI副图{slot}上传错误'
    status = str(item.get(status_field, '') or '').strip()
    error = str(item.get(error_field, '') or '').strip()
    return {'status': status, 'error': error}


def _resolve_current_image_preview(locator: str, ai_locator: str = '', ai_preview: str = '') -> str:
    locator = str(locator or '').strip()
    ai_locator = str(ai_locator or '').strip()
    ai_preview = str(ai_preview or '').strip()
    if _is_http_url(locator):
        return locator
    if locator and ai_locator and locator == ai_locator and ai_preview:
        return ai_preview
    return ''


def _build_ai_image_result_data(item: dict) -> dict:
    main_preview = _resolve_ai_preview_image_url(item, 'main')
    main_locator = _resolve_ai_media_locator(item, 'main')
    main_state = _resolve_ai_upload_state(item, 'main')
    result = {
        'ai_main_image': main_preview,
        'ai_main_image_preview': main_preview,
        'ai_media_locator': main_locator,
        'ai_public_image_url': _resolve_ai_public_image_url(item, 'main'),
        'ai_upload_status': main_state['status'],
        'ai_upload_error': main_state['error'],
    }
    result['ai_image_url'] = result['ai_public_image_url'] or result['ai_main_image']

    for slot in range(2, 10):
        preview_url = _resolve_ai_preview_image_url(item, slot)
        locator = _resolve_ai_media_locator(item, slot)
        public_url = _resolve_ai_public_image_url(item, slot)
        state = _resolve_ai_upload_state(item, slot)
        result[f'ai_image_{slot}'] = public_url or preview_url
        result[f'ai_image_{slot}_preview'] = preview_url
        result[f'ai_media_locator_{slot}'] = locator
        result[f'ai_public_image_{slot}'] = public_url
        result[f'ai_image_{slot}_upload_status'] = state['status']
        result[f'ai_image_{slot}_upload_error'] = state['error']

    return result


def _collect_ai_image_persist_updates(item: dict) -> dict:
    updates = {}
    for slot in ['main', *range(2, 10)]:
        path_field = 'AI主图路径' if slot == 'main' else f'AI副图{slot}路径'
        url_field = 'AI主图URL' if slot == 'main' else f'AI副图{slot}URL'
        preview_field = 'AI主图预览URL' if slot == 'main' else f'AI副图{slot}预览URL'
        status_field = 'AI主图上传状态' if slot == 'main' else f'AI副图{slot}上传状态'
        error_field = 'AI主图上传错误' if slot == 'main' else f'AI副图{slot}上传错误'
        path_value = str(item.get(path_field, '') or '').strip()
        url_value = str(item.get(url_field, '') or '').strip()
        preview_value = str(item.get(preview_field, '') or '').strip()
        status_value = str(item.get(status_field, '') or '').strip()
        error_value = str(item.get(error_field, '') or '').strip()
        has_slot_data = any(
            field in item for field in (path_field, url_field, preview_field, status_field, error_field)
        ) or any((path_value, url_value, preview_value, status_value, error_value))
        if has_slot_data:
            updates[path_field] = path_value
            updates[url_field] = url_value
            updates[preview_field] = preview_value
            updates[status_field] = status_value
            updates[error_field] = error_value
    return updates


def _pick_existing_header(headers, *candidates, default=''):
    header_set = {str(header).strip() for header in headers if str(header or '').strip()}
    for candidate in candidates:
        if candidate in header_set:
            return candidate
    return default or (candidates[0] if candidates else '')


def _format_submit_issues(result_entry: dict) -> str:
    issues = result_entry.get('issues') or []
    messages = []

    for issue in issues:
        if isinstance(issue, dict):
            severity = str(issue.get('severity', '') or '').strip().upper()
            code = str(issue.get('code', '') or '').strip()
            message = str(issue.get('message', '') or issue.get('details', '') or '').strip()
            prefix_parts = [part for part in (severity, code) if part]
            prefix = f"[{'/'.join(prefix_parts)}] " if prefix_parts else ''
            if message:
                messages.append(prefix + message)
        else:
            text = str(issue or '').strip()
            if text:
                messages.append(text)

    if messages:
        return '; '.join(messages)
    return str(result_entry.get('message', '') or '').strip()


def _build_submit_persist_updates(headers, result_entry: dict, submit_time: str) -> dict:
    status = str(result_entry.get('status', '') or 'UNKNOWN').strip() or 'UNKNOWN'
    submission_id = str(
        result_entry.get('submission_id', '')
        or result_entry.get('submissionId', '')
        or ''
    ).strip()
    asin = str(result_entry.get('asin', '') or '').strip()
    submit_message = _format_submit_issues(result_entry)

    updates = {
        _pick_existing_header(headers, 'submit_status', '提交状态', default='submit_status'): status,
        _pick_existing_header(headers, 'submission_id', '提交ID', default='submission_id'): submission_id,
        _pick_existing_header(headers, 'submit_time', '提交时间', default='submit_time'): submit_time,
        _pick_existing_header(headers, 'submit_message', '提交信息', '问题详情', default='submit_message'): submit_message,
    }
    if asin:
        updates[_pick_existing_header(headers, 'asin', 'ASIN', default='asin')] = asin
    return updates


def _build_preview_persist_updates(headers, result_entry: dict, preview_time: str, account_name: str) -> dict:
    status = str(result_entry.get('status', '') or '').strip().upper()
    if not status:
        status = 'VALID' if result_entry.get('valid') else 'INVALID'

    details = []
    for message in result_entry.get('errors', []) or []:
        text = str(message or '').strip()
        if text:
            details.append(f'[ERROR] {text}')
    for message in result_entry.get('warnings', []) or []:
        text = str(message or '').strip()
        if text:
            details.append(f'[WARN] {text}')

    return {
        _pick_existing_header(headers, 'preview_status', '预览状态', default='preview_status'): status,
        _pick_existing_header(headers, 'preview_message', '预览信息', default='preview_message'): '; '.join(details),
        _pick_existing_header(headers, 'preview_time', '预览时间', default='preview_time'): preview_time,
        _pick_existing_header(headers, 'preview_account', '预览账号', default='preview_account'): account_name,
    }


def _build_listing_api_context(account_id: str = ''):
    from amazon.accounts import AccountManager
    from amazon.auth import AmazonAuth
    from amazon.listings import ListingsAPI
    from amazon.mapper import FieldMapper

    mgr = AccountManager()
    acc = mgr.get_account(account_id) if account_id else mgr.get_default_account()
    mapper = FieldMapper(acc.get('marketplace_id', config.AMAZON_MARKETPLACE)) if acc else FieldMapper(config.AMAZON_MARKETPLACE)

    if not acc:
        return acc, mapper, None, '未配置亚马逊账号'

    required = [acc.get('lwa_client_id'), acc.get('lwa_client_secret'), acc.get('refresh_token'), acc.get('seller_id')]
    if not all(required):
        return acc, mapper, None, '亚马逊账号凭证不完整'

    auth = AmazonAuth(
        client_id=acc['lwa_client_id'],
        client_secret=acc['lwa_client_secret'],
        refresh_token=acc['refresh_token'],
    )
    listings_api = ListingsAPI(
        auth=auth,
        seller_id=acc['seller_id'],
        marketplace_id=acc.get('marketplace_id', config.AMAZON_MARKETPLACE),
    )
    return acc, mapper, listings_api, ''


def _normalize_diagnostic_issue(issue: dict, source: str) -> dict:
    severity = str(issue.get('severity') or issue.get('level') or 'ERROR').strip().upper()
    level_map = {'ERROR': 'error', 'WARNING': 'warning', 'WARN': 'warning', 'INFO': 'info'}
    level = level_map.get(severity, 'error')
    attrs = issue.get('attributeNames') or issue.get('attributes') or []
    if isinstance(attrs, str):
        attrs = [attrs]
    categories = issue.get('categories') or []
    if isinstance(categories, str):
        categories = [categories]
    return {
        'source': source,
        'level': level,
        'severity': severity,
        'code': str(issue.get('code', '') or '').strip(),
        'message': str(issue.get('message', '') or issue.get('details', '') or '').strip(),
        'attributeNames': [str(attr).strip() for attr in attrs if str(attr).strip()],
        'categories': [str(cat).strip().upper() for cat in categories if str(cat).strip()],
    }


def _probe_remote_media_url(field_name: str, url: str) -> dict:
    text = str(url or '').strip()
    result = {
        'field': field_name,
        'url': text,
        'ok': False,
        'status_code': None,
        'content_type': '',
        'message': '',
    }
    if not text:
        result['message'] = '未提供媒体地址'
        return result
    if text.lower().startswith('s3://'):
        result['ok'] = True
        result['message'] = 's3:// 媒体定位符，跳过 HTTP 连通性探测'
        return result
    if not _is_http_url(text):
        result['message'] = '媒体地址必须是 http(s) 或 s3://'
        return result

    try:
        response = requests.head(text, allow_redirects=True, timeout=8)
        status_code = int(response.status_code)
        content_type = str(response.headers.get('Content-Type', '') or '').strip()
        if status_code >= 400 or status_code in (401, 403, 405):
            response = requests.get(text, allow_redirects=True, timeout=12, stream=True)
            status_code = int(response.status_code)
            content_type = str(response.headers.get('Content-Type', '') or '').strip()
        result['status_code'] = status_code
        result['content_type'] = content_type
        result['ok'] = status_code < 400
        result['message'] = '可抓取' if result['ok'] else f'HTTP {status_code}'
        if result['ok'] and content_type and not content_type.lower().startswith('image/'):
            result['message'] = f'返回内容不是图片: {content_type}'
            result['ok'] = False
    except Exception as exc:
        result['message'] = str(exc)
    return result


def _probe_product_media(product: dict) -> dict:
    checks = []
    main_url = str(product.get('main_image_url', '') or '').strip()
    if main_url:
        checks.append(_probe_remote_media_url('main_image_url', main_url))
    for idx in range(1, 9):
        image_url = str(product.get(f'other_image_{idx}', '') or '').strip()
        if image_url:
            checks.append(_probe_remote_media_url(f'other_image_{idx}', image_url))

    failed = [item for item in checks if not item.get('ok')]
    passed = [item for item in checks if item.get('ok')]
    status = 'pass' if checks and not failed else ('fail' if failed else 'none')
    return {
        'status': status,
        'total': len(checks),
        'passed': len(passed),
        'failed': len(failed),
        'checks': checks,
    }


def _extract_listing_asin(payload: dict) -> str:
    payload = payload or {}
    for summary in payload.get('summaries', []) or []:
        asin = str(summary.get('asin', '') or '').strip()
        if asin:
            return asin
    for ident in payload.get('identifiers', []) or []:
        asin = str(ident.get('asin', '') or '').strip()
        if asin:
            return asin
    return ''


def _collect_missing_fields(validation: dict, preview_issues: list, listing_issues: list, field_meta: dict) -> list:
    merged = {}

    for item in validation.get('schema_required_missing', []) or []:
        name = str(item.get('name', '') or '').strip()
        if not name:
            continue
        meta = field_meta.get(name, {})
        merged[name] = {
            'name': name,
            'title': str(item.get('title', '') or meta.get('title', name)).strip() or name,
            'group': str(item.get('group', '') or meta.get('group', '') or 'other').strip() or 'other',
            'description': str(item.get('description', '') or meta.get('description', '') or '').strip(),
            'source': 'schema',
        }

    for source, issues in (('amazon_preview', preview_issues), ('listing_runtime', listing_issues)):
        for issue in issues:
            categories = set(issue.get('categories', []) or [])
            code = str(issue.get('code', '') or '').strip()
            if 'MISSING_ATTRIBUTE' not in categories and code != '90220':
                continue
            for attr_name in issue.get('attributeNames', []) or []:
                attr_name = str(attr_name or '').strip()
                if not attr_name:
                    continue
                meta = field_meta.get(attr_name, {})
                merged.setdefault(attr_name, {
                    'name': attr_name,
                    'title': str(meta.get('title', attr_name)).strip() or attr_name,
                    'group': str(meta.get('group', '') or 'other').strip() or 'other',
                    'description': str(meta.get('description', '') or '').strip(),
                    'source': source,
                })

    return list(merged.values())


def _build_listing_check_summary(result: dict) -> str:
    parts = []
    if result.get('missing_fields'):
        parts.append(f"缺少字段 {len(result['missing_fields'])} 项")
    preview_status = str(result.get('preview', {}).get('status', '') or '').strip().upper()
    if preview_status:
        parts.append(f'预览 {preview_status}')
    media = result.get('media', {})
    if media.get('failed'):
        parts.append(f"图片异常 {media['failed']} 项")
    listing = result.get('listing', {})
    if listing.get('exists'):
        parts.append('后台已存在该 SKU')
    if not parts:
        parts.append('未发现明显缺项')
    return '；'.join(parts)


def _build_listing_check_persist_updates(headers, result_entry: dict, checked_time: str, account_name: str) -> dict:
    preview_messages = [issue.get('message', '') for issue in result_entry.get('preview', {}).get('issues', []) or [] if issue.get('message')]
    listing_messages = [issue.get('message', '') for issue in result_entry.get('listing', {}).get('issues', []) or [] if issue.get('message')]
    media_messages = [
        f"{item.get('field')}: {item.get('message')}"
        for item in result_entry.get('media', {}).get('checks', []) or []
        if not item.get('ok') and item.get('message')
    ]
    missing_fields = [
        f"{str(item.get('title', '') or item.get('name', '')).strip()} ({str(item.get('name', '') or '').strip()})"
        for item in result_entry.get('missing_fields', []) or []
        if str(item.get('name', '') or '').strip()
    ]

    issue_messages = []
    if preview_messages:
        issue_messages.append('预览: ' + '; '.join(preview_messages))
    if listing_messages:
        issue_messages.append('后台: ' + '; '.join(listing_messages))
    if media_messages:
        issue_messages.append('媒体: ' + '; '.join(media_messages))

    return {
        _pick_existing_header(headers, 'listing_check_status', '缺项诊断状态', default='listing_check_status'): str(result_entry.get('status', '') or '').strip(),
        _pick_existing_header(headers, 'listing_check_summary', '缺项诊断摘要', default='listing_check_summary'): str(result_entry.get('summary_text', '') or '').strip(),
        _pick_existing_header(headers, 'listing_check_missing_fields', '缺失字段清单', default='listing_check_missing_fields'): '; '.join(missing_fields),
        _pick_existing_header(headers, 'listing_check_issues', '缺项诊断问题', default='listing_check_issues'): ' | '.join(issue_messages),
        _pick_existing_header(headers, 'listing_check_time', '缺项诊断时间', default='listing_check_time'): checked_time,
        _pick_existing_header(headers, 'listing_check_account', '缺项诊断账号', default='listing_check_account'): account_name,
    }


def _apply_listing_check_results_to_file(input_file: str, headers: list, results: list, account_name: str):
    checked_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    updates_by_sku = {}
    for result in results:
        sku = str(result.get('sku', '') or '').strip()
        if not sku:
            continue
        result['checked_at'] = checked_time
        updates_by_sku[sku] = _build_listing_check_persist_updates(headers, result, checked_time, account_name)

    if updates_by_sku:
        _persist_bulk_row_updates(input_file, updates_by_sku)


def _run_listing_check_for_product(product: dict, mapper, listings_api=None, include_listing_lookup: bool = True,
                                   include_media_probe: bool = True) -> dict:
    sku = str(product.get('sku', '') or '').strip() or 'N/A'
    schema_fields = mapper._load_schema_fields(product.get('product_type', ''))
    field_meta = {}
    if schema_fields:
        for field in (schema_fields.get('required_fields', []) or []) + (schema_fields.get('optional_fields', []) or []):
            name = str(field.get('name', '') or '').strip()
            if name:
                field_meta[name] = {
                    'title': str(field.get('title', '') or name).strip() or name,
                    'group': str(field.get('group', '') or 'other').strip() or 'other',
                    'description': str(field.get('description', '') or '').strip(),
                }

    validation = mapper.validate_required_fields(product, schema_fields=schema_fields)
    media = _probe_product_media(product) if include_media_probe else {
        'status': 'none',
        'total': 0,
        'passed': 0,
        'failed': 0,
        'checks': [],
    }

    preview = {
        'attempted': False,
        'status': '',
        'issues': [],
        'message': '',
    }
    if listings_api is not None:
        if validation.get('valid'):
            preview_result = listings_api.put_listings_item(sku, product, preview=True)
            preview['attempted'] = True
            preview['status'] = str(preview_result.get('status', '') or '').strip().upper()
            preview['issues'] = [
                _normalize_diagnostic_issue(issue, 'amazon_preview')
                for issue in (preview_result.get('issues') or [])
            ]
        else:
            preview['status'] = 'SKIPPED_LOCAL_INVALID'
            preview['message'] = '本地必填校验未通过，已跳过 Amazon 预览'
    else:
        preview['status'] = 'UNAVAILABLE'
        preview['message'] = '当前未配置可用亚马逊账号，无法执行 Amazon 预览'

    listing_state = {
        'queried': False,
        'exists': False,
        'asin': '',
        'issues': [],
        'summaries_count': 0,
    }
    if include_listing_lookup and listings_api is not None and sku and sku != 'N/A':
        listing_payload = listings_api.get_listings_item(sku)
        listing_state['queried'] = True
        if listing_payload:
            listing_state['exists'] = True
            listing_state['asin'] = _extract_listing_asin(listing_payload)
            listing_state['summaries_count'] = len(listing_payload.get('summaries', []) or [])
            listing_state['issues'] = [
                _normalize_diagnostic_issue(issue, 'listing_runtime')
                for issue in (listing_payload.get('issues') or [])
            ]

    missing_fields = _collect_missing_fields(
        validation=validation,
        preview_issues=preview.get('issues', []),
        listing_issues=listing_state.get('issues', []),
        field_meta=field_meta,
    )

    local_errors = [str(message or '').strip() for message in validation.get('errors', []) or [] if str(message or '').strip()]
    local_warnings = [str(message or '').strip() for message in validation.get('warnings', []) or [] if str(message or '').strip()]
    preview_errors = [issue for issue in preview.get('issues', []) if issue.get('level') == 'error']
    preview_warnings = [issue for issue in preview.get('issues', []) if issue.get('level') == 'warning']
    listing_errors = [issue for issue in listing_state.get('issues', []) if issue.get('level') == 'error']
    listing_warnings = [issue for issue in listing_state.get('issues', []) if issue.get('level') == 'warning']

    error_count = len(local_errors) + len(preview_errors) + len(listing_errors) + int(media.get('failed', 0))
    warning_count = len(local_warnings) + len(preview_warnings) + len(listing_warnings)
    status = 'fail' if error_count > 0 else ('warn' if warning_count > 0 else 'pass')

    result = {
        'sku': sku,
        'status': status,
        'summary_text': '',
        'summary_counts': {
            'errors': error_count,
            'warnings': warning_count,
            'missing_fields': len(missing_fields),
            'media_failed': int(media.get('failed', 0)),
        },
        'local': {
            'valid': bool(validation.get('valid')),
            'errors': local_errors,
            'warnings': local_warnings,
            'schema_required_missing': validation.get('schema_required_missing', []),
        },
        'preview': preview,
        'listing': listing_state,
        'media': media,
        'missing_fields': missing_fields,
    }
    result['summary_text'] = _build_listing_check_summary(result)
    return result


def _resolve_ai_status_from_result(result_data: dict, action: str) -> str:
    if action == 'image':
        image_keys = ['ai_main_image']
        image_keys.extend([f'ai_image_{i}' for i in range(2, 10)])
        return 'completed' if any(str(result_data.get(key, '') or '').strip() for key in image_keys) else 'failed'

    ai_value_keys = ['ai_title', 'ai_description', 'ai_keywords']
    ai_value_keys.extend([f'ai_bullet_{i}' for i in range(1, 6)])
    has_content = any(str(result_data.get(key, '') or '').strip() for key in ai_value_keys)
    return 'completed' if has_content else 'failed'


def _normalize_endpoint_template(value: str) -> str:
    """将错误的 {具体模型名} 归一化为 {model} 模板。"""
    text = str(value or '').strip()
    if not text:
        return text
    return re.sub(r'\{([^{}]+)\}', lambda m: '{model}' if m.group(1) != 'model' else m.group(0), text)


def _build_task_title(stage: int, action: str = '', scope: str = '') -> str:
    action = str(action or '').strip().lower()
    scope = str(scope or '').strip().lower()
    if stage == 1:
        if action == 'image':
            return '批量AI生图（全部图片）' if scope == 'all' else '批量AI生图（主图）'
        if action == 'rewrite':
            if scope == 'title':
                return '批量AI文案（标题）'
            if scope == 'bullets':
                return '批量AI文案（卖点）'
            return '批量AI文案'
        return '阶段1批量处理'
    if stage == 2:
        return '批量提交到亚马逊'
    return '任务处理'

# 全局任务状态
_task_lock = threading.Lock()
_task_cancel_event = threading.Event()
_TASK_HISTORY_MAX = 30
task_status = {
    'running': False,
    'stage': None,
    'progress': 0,
    'total': 0,
    'current_item': '',
    'logs': [],
    'result_file': None,
    'error': None,
    'cancel_requested': False,
    'cancelled': False,
    'task_id': None,
    'task_kind': '',
    'task_title': '',
    'started_at': '',
}
task_history = []

# 单行处理任务追踪（支持多任务并发）
single_tasks = {}


def update_status(**kwargs):
    """线程安全地更新task_status"""
    with _task_lock:
        task_status.update(kwargs)
        snapshot = dict(task_status)
    _sync_task_history_from_status(snapshot)


def get_task_status():
    """线程安全地读取task_status"""
    with _task_lock:
        return dict(task_status)


def add_log(msg):
    """添加日志"""
    with _task_lock:
        task_status['logs'].append({
            'time': datetime.now().strftime('%H:%M:%S'),
            'msg': msg,
        })
        # 只保留最近100条
        if len(task_status['logs']) > 100:
            task_status['logs'] = task_status['logs'][-100:]
        snapshot = dict(task_status)
    _sync_task_history_from_status(snapshot)


def _append_task_record(record: dict):
    with _task_lock:
        task_history.insert(0, record)
        if len(task_history) > _TASK_HISTORY_MAX:
            del task_history[_TASK_HISTORY_MAX:]


def _update_task_record(task_id: str, **updates):
    if not task_id:
        return
    with _task_lock:
        for record in task_history:
            if record.get('id') == task_id:
                record.update(updates)
                record['updated_at'] = datetime.now().isoformat(timespec='seconds')
                return


def _start_task_record(kind: str, title: str, input_file: str = '', **extra) -> dict:
    task_id = uuid4().hex[:12]
    record = {
        'id': task_id,
        'kind': kind,
        'title': title,
        'status': 'running',
        'progress': 0,
        'total': 0,
        'current_item': '',
        'message': '',
        'input_file': input_file,
        'file_name': os.path.basename(input_file) if input_file else '',
        'result_file': '',
        'error': '',
        'started_at': datetime.now().isoformat(timespec='seconds'),
        'ended_at': '',
        'updated_at': datetime.now().isoformat(timespec='seconds'),
        'logs': [],
        **extra,
    }
    _append_task_record(record)
    return record


def _record_instant_task(kind: str, title: str, status: str, input_file: str = '', **extra) -> dict:
    record = _start_task_record(kind=kind, title=title, input_file=input_file, **extra)
    _update_task_record(
        record['id'],
        status=status,
        ended_at=datetime.now().isoformat(timespec='seconds'),
    )
    return record


def _sync_task_history_from_status(snapshot: dict):
    task_id = snapshot.get('task_id')
    if not task_id:
        return
    recent_logs = snapshot.get('logs', [])[-20:]
    _update_task_record(
        task_id,
        progress=int(snapshot.get('progress', 0) or 0),
        total=int(snapshot.get('total', 0) or 0),
        current_item=snapshot.get('current_item', ''),
        result_file=snapshot.get('result_file') or '',
        error=snapshot.get('error') or '',
        cancel_requested=bool(snapshot.get('cancel_requested')),
        cancelled=bool(snapshot.get('cancelled')),
        logs=recent_logs,
    )


def _get_task_record(task_id: str):
    task_id = str(task_id or '').strip()
    if not task_id:
        return None
    with _task_lock:
        for record in task_history:
            if record.get('id') == task_id:
                return dict(record)
    return None


def _append_task_log(task_id: str, message: str):
    record = _get_task_record(task_id)
    if not record:
        return
    logs = list(record.get('logs') or [])
    logs.append({
        'time': datetime.now().strftime('%H:%M:%S'),
        'msg': str(message or '').strip(),
    })
    _update_task_record(task_id, logs=logs[-30:])


def _set_task_stage(task_id: str, stage_name: str, stage_index: int, stages: list, progress: int = None,
                    total: int = None, current_item: str = ''):
    updates = {
        'stage_name': stage_name,
        'stage_index': stage_index,
        'stages': list(stages or []),
    }
    if progress is not None:
        updates['progress'] = int(progress)
    if total is not None:
        updates['total'] = int(total)
    if current_item:
        updates['current_item'] = current_item
    _update_task_record(task_id, **updates)


def _complete_task_record(task_id: str, message: str = '', result: dict = None, **extra):
    _update_task_record(
        task_id,
        status='completed',
        message=message,
        result=result or {},
        ended_at=datetime.now().isoformat(timespec='seconds'),
        **extra,
    )


def _fail_task_record(task_id: str, message: str, **extra):
    _update_task_record(
        task_id,
        status='failed',
        message=str(message or '').strip(),
        error=str(message or '').strip(),
        ended_at=datetime.now().isoformat(timespec='seconds'),
        **extra,
    )


def _finalize_current_task_record():
    snapshot = get_task_status()
    task_id = snapshot.get('task_id')
    if not task_id:
        return

    if snapshot.get('cancelled'):
        status = 'cancelled'
        message = '任务已取消'
    elif snapshot.get('error'):
        status = 'failed'
        message = str(snapshot.get('error') or '')
    else:
        status = 'completed'
        message = str(snapshot.get('current_item', '') or '任务完成')

    _update_task_record(
        task_id,
        status=status,
        message=message,
        ended_at=datetime.now().isoformat(timespec='seconds'),
        result_file=snapshot.get('result_file') or '',
        error=snapshot.get('error') or '',
        cancelled=bool(snapshot.get('cancelled')),
    )


# ===== 路由 =====

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/upload', methods=['POST'])
def upload_excel():
    """上传Excel文件"""
    if 'file' not in request.files:
        return jsonify({'error': '请选择文件'}), 400

    file = request.files['file']
    original_filename = file.filename or ''
    if not original_filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': '仅支持 .xlsx 或 .xls 文件'}), 400

    # MIME type check
    allowed_mimes = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
        'application/octet-stream',  # some browsers send this
    }
    if file.content_type and file.content_type not in allowed_mimes:
        return jsonify({'error': f'不支持的文件类型: {file.content_type}'}), 400

    filename = secure_filename(original_filename) or 'upload.xlsx'
    # 保持原始中文文件名
    if original_filename:
        filename = original_filename.replace('/', '_').replace('\\', '_')

    filepath = os.path.join(config.INPUT_DIR, filename)
    os.makedirs(config.INPUT_DIR, exist_ok=True)
    file.save(filepath)

    try:
        processor = ExcelProcessor()
        data = processor.read_input(filepath)
        col_map = processor.detect_columns()
        ext = os.path.splitext(filename)[1].lower()
        warning = ''
        xls_support = None
        if ext == '.xls':
            xls_support = _detect_xls_preserve_support()
            warning = '已进入 .xls 兼容模式，建议尽快另存为 .xlsx。复杂样式保留依赖本机 Excel/WPS。'
            if xls_support.get('supported'):
                warning += f" 当前已检测到 {xls_support.get('app')}，可优先保留原样式。"
            else:
                warning += ' 当前未检测到可用的 Excel/WPS COM，写回时可能回退为重建工作簿。'

        if len(data) > config.BATCH_LIMIT:
            os.remove(filepath)
            return jsonify({
                'error': f'文件包含 {len(data)} 条商品，超过当前导入上限 {config.BATCH_LIMIT} 条，请调整设置后重试'
            }), 400

        preview = []
        for item in data[:10]:
            row = {}
            for key, val in item.items():
                if not key.startswith('_'):
                    row[key] = str(val) if val is not None else ''
            preview.append(row)

        return jsonify({
            'success': True,
            'filename': filename,
            'filepath': filepath,
            'total_rows': len(data),
            'file_type': ext.lstrip('.'),
            'columns': [h for h in processor.headers],
            'column_mapping': col_map,
            'preview': preview,
            'warning': warning,
            'xls_preserve_supported': bool(xls_support and xls_support.get('supported')),
        })
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'error': f'读取Excel失败: {str(e)}'}), 400


@app.route('/api/process', methods=['POST'])
def start_processing():
    """启动处理任务"""
    if task_status['running']:
        return jsonify({'error': '已有任务在运行'}), 400

    data = request.json or {}
    input_file = data.get('input_file') or data.get('file', '')
    stage = data.get('stage', 1)
    options = dict(data.get('options', {}) or {})

    # 兼容当前前端协议: { skus, action, scope, file }
    if data.get('skus') and 'selected_skus' not in options:
        options['selected_skus'] = data.get('skus', [])

    action = data.get('action')
    scope = data.get('scope')
    if action == 'rewrite':
        options.setdefault('process_text', True)
        options.setdefault('process_images', False)
        options.setdefault('text_fields', _map_regenerate_field(scope or 'all'))
        stage = 1
    elif action == 'image':
        options.setdefault('process_text', False)
        options.setdefault('process_images', True)
        options.setdefault('image_scope', scope or 'main')
        options.setdefault('image_style', data.get('bg_style') or data.get('image_style') or 'white')
        stage = 1

    # Support top-level params in addition to nested options
    for key in ('text_fields', 'overwrite_existing', 'selected_skus',
                'process_text', 'process_images', 'image_scope', 'image_style'):
        if key in data and key not in options:
            options[key] = data[key]

    if not input_file or not os.path.exists(input_file):
        return jsonify({'error': '输入文件不存在'}), 400

    task_record = _start_task_record(
        kind='batch_process',
        title=_build_task_title(stage, action, scope),
        input_file=input_file,
        stage=stage,
        action=action,
        scope=scope,
    )

    # 重置状态
    _task_cancel_event.clear()
    update_status(
        running=True,
        stage=stage,
        progress=0,
        total=0,
        current_item='初始化...',
        logs=[],
        result_file=None,
        error=None,
        cancel_requested=False,
        cancelled=False,
        task_id=task_record['id'],
        task_kind=task_record['kind'],
        task_title=task_record['title'],
        started_at=task_record['started_at'],
    )

    # 后台线程执行
    thread = threading.Thread(
        target=_run_task,
        args=(input_file, stage, options),
        daemon=True,
    )
    thread.start()

    return jsonify({
        'success': True,
        'message': f'阶段{stage}任务已启动',
        'task_id': task_record['id'],
    })


@app.route('/api/process/cancel', methods=['POST'])
def cancel_processing():
    """请求取消当前后台任务。"""
    if not get_task_status().get('running'):
        return jsonify({'success': True, 'message': '当前没有正在运行的任务'})

    _task_cancel_event.set()
    update_status(cancel_requested=True, current_item='正在取消...')
    add_log('⏹️ 收到取消请求，等待当前任务安全停止...')
    return jsonify({'success': True, 'message': '已发送取消请求，等待当前任务收尾'})


@app.route('/api/process-single', methods=['POST'])
def process_single():
    """单行处理端点 — 对单个SKU执行仿写或改图，同步返回结果"""
    try:
        data = request.json
        input_file = data.get('file', '') or data.get('input_file', '')
        sku = data.get('sku', '')
        action = data.get('action', 'rewrite')  # rewrite | image
        requested_field = data.get('field', '')
        text_fields = data.get('text_fields') or _map_regenerate_field(requested_field)

        if not input_file or not os.path.exists(input_file):
            return jsonify({'error': '输入文件不存在'}), 400
        if not sku:
            return jsonify({'error': '请指定SKU'}), 400
        if action not in ('rewrite', 'image'):
            return jsonify({'error': f'无效的action: {action}，支持 rewrite 或 image'}), 400

        # 读取Excel并定位目标SKU
        processor = ExcelProcessor()
        all_data = processor.read_input(input_file)
        col_map = processor.detect_columns()
        sku_col = col_map.get('sku', '')

        target_item = None
        for item in all_data:
            item_sku = str(item.get(sku_col, '') or item.get('SKU', '')).strip()
            if item_sku == sku:
                target_item = item
                break

        if target_item is None:
            return jsonify({'error': f'未找到SKU: {sku}'}), 404

        cfg = get_config()
        if not cfg.AI_API_KEY:
            return jsonify({'error': '未配置API Key，无法执行AI处理'}), 400

        result_data = {'sku': sku, 'action': action}

        if action == 'rewrite':
            from stage1_pipeline import Stage1Pipeline
            pipeline = Stage1Pipeline()

            # 构建产品信息
            product_info = pipeline._build_product_info(target_item, col_map)
            product_type = pipeline._detect_product_type(target_item, col_map, product_info)

            # 选择性仿写
            if 'title' in text_fields:
                from core.prompts.amazon_prompts import TITLE_PROMPT
                title = pipeline._ai_text(TITLE_PROMPT.format(
                    product_info=product_info, product_type=product_type))
                if len(title) > 200:
                    title = title[:197] + "..."
                target_item['AI标题'] = title
                result_data['ai_title'] = title

            if 'bullets' in text_fields:
                from core.prompts.amazon_prompts import BULLET_POINTS_PROMPT
                bullets_raw = pipeline._ai_text(BULLET_POINTS_PROMPT.format(
                    product_info=product_info, product_type=product_type))
                bullets = pipeline._parse_bullets(bullets_raw)
                for i, bp in enumerate(bullets, 1):
                    if len(bp) > 500:
                        bp = bp[:497] + "..."
                    target_item[f'AI卖点{i}'] = bp
                    result_data[f'ai_bullet_{i}'] = bp

            if 'description' in text_fields:
                from core.prompts.amazon_prompts import DESCRIPTION_PROMPT
                desc = pipeline._ai_text(DESCRIPTION_PROMPT.format(
                    product_info=product_info, product_type=product_type))
                if len(desc) > 2000:
                    desc = desc[:1997] + "..."
                target_item['AI商品描述'] = desc
                result_data['ai_description'] = desc

            if 'keywords' in text_fields:
                from core.prompts.amazon_prompts import SEARCH_TERMS_PROMPT
                existing_title = target_item.get('AI标题', '') or target_item.get(col_map.get('title', ''), '')
                keywords = pipeline._ai_text(SEARCH_TERMS_PROMPT.format(
                    product_info=product_info, product_type=product_type,
                    title=existing_title))
                kw_bytes = len(keywords.encode('utf-8'))
                if kw_bytes > 250:
                    words = keywords.split()
                    trimmed = []
                    total = 0
                    for w in words:
                        wb = len(w.encode('utf-8')) + 1
                        if total + wb > 249:
                            break
                        trimmed.append(w)
                        total += wb
                    keywords = " ".join(trimmed)
                target_item['AI搜索关键词'] = keywords
                result_data['ai_keywords'] = keywords

            ai_status = _resolve_ai_status_from_result(result_data, 'rewrite')
            target_item['AI状态'] = ai_status
            result_data['ai_status'] = ai_status

            persist_updates = {'AI状态': ai_status}
            if 'ai_title' in result_data:
                persist_updates['AI标题'] = result_data['ai_title']
            if 'ai_description' in result_data:
                persist_updates['AI商品描述'] = result_data['ai_description']
            if 'ai_keywords' in result_data:
                persist_updates['AI搜索关键词'] = result_data['ai_keywords']
            for i in range(1, 6):
                bullet_key = f'ai_bullet_{i}'
                if bullet_key in result_data:
                    persist_updates[f'AI卖点{i}'] = result_data[bullet_key]
            _persist_row_updates(input_file, sku, persist_updates)

        elif action == 'image':
            from stage1_pipeline import Stage1Pipeline
            pipeline = Stage1Pipeline()
            scope = str(data.get('scope', 'main') or 'main').strip().lower()
            bg_style = str(data.get('bg_style', data.get('image_style', 'white')) or 'white').strip().lower()
            image_sources = pipeline._collect_image_sources(target_item, col_map, scope)
            if not image_sources:
                return jsonify({'error': '该SKU没有有效的图片URL'}), 400

            pipeline._process_images(target_item, col_map, 0, scope=scope, bg_style=bg_style)
            result_data.update(_build_ai_image_result_data(target_item))
            result_data['ai_status'] = _resolve_ai_status_from_result(result_data, 'image')
            target_item['AI状态'] = result_data['ai_status']
            persist_updates = {'AI状态': result_data['ai_status']}
            persist_updates.update(_collect_ai_image_persist_updates(target_item))

            if result_data['ai_status'] == 'completed':
                _persist_row_updates(input_file, sku, persist_updates)
            else:
                _persist_row_updates(input_file, sku, {'AI状态': 'failed'})
                result_data['warning'] = '图片处理未产生结果'

        return jsonify({'success': True, 'result': result_data})

    except Exception as e:
        logger.error(f'单行处理失败: {e}')
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


def _run_task(input_file, stage, options):
    """后台执行任务"""
    try:
        if stage == 1:
            add_log('🚀 第一阶段启动：AI内容处理')

            # 先读取数据获取总数
            processor = ExcelProcessor()
            data = processor.read_input(input_file)
            col_map = processor.detect_columns()

            # 按选中的SKU过滤 (新增)
            actual_input = input_file
            selected_skus = options.get('selected_skus')
            if selected_skus:
                sku_col = col_map.get('sku', '')
                sku_set = set(selected_skus)
                filtered = [item for item in data
                            if str(item.get(sku_col, '') or item.get('SKU', '')).strip() in sku_set]
                if filtered:
                    temp_file = os.path.join(config.OUTPUT_DIR, f'_temp_filtered_{int(time.time())}.xlsx')
                    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
                    processor.write_output(filtered, temp_file)
                    actual_input = temp_file
                    data = filtered
                    add_log(f'🔍 已过滤选中的 {len(filtered)} 个SKU')
                else:
                    add_log(f'⚠️ 未匹配到选中的SKU，处理全部数据')

            update_status(total=len(data))
            add_log(f'📖 读取到 {len(data)} 条商品, {len(processor.headers)} 列')

            # 记录增强选项
            if options.get('text_fields'):
                add_log(f'📝 选择性仿写字段: {", ".join(options["text_fields"])}')
            if options.get('overwrite_existing') is not None:
                add_log(f'🔄 覆盖已有AI结果: {"是" if options["overwrite_existing"] else "否"}')

            # 检查API Key
            cfg = get_config()
            if not cfg.AI_API_KEY:
                add_log('⚠️ 未配置API Key，跳过AI处理，直接输出原始数据')

            timestamp = time.strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(config.OUTPUT_DIR, f'处理结果_{timestamp}.xlsx')
            os.makedirs(config.OUTPUT_DIR, exist_ok=True)

            # 注入进度回调到stage1 logger
            import logging as _logging
            class ProgressHandler(_logging.Handler):
                def emit(self, record):
                    msg = record.getMessage()
                    add_log(msg)
                    if '完成' in msg:
                        update_status(current_item=msg[:50])

            s1_logger = _logging.getLogger('stage1_pipeline')
            handler = ProgressHandler()
            handler.setLevel(_logging.INFO)
            s1_logger.addHandler(handler)

            try:
                from stage1_pipeline import Stage1Pipeline
                pipeline = Stage1Pipeline()
                run_kwargs = dict(
                    input_file=actual_input,
                    output_file=output_file,
                    process_images=options.get('process_images', True),
                    process_text=options.get('process_text', True),
                    image_scope=options.get('image_scope', 'main'),
                    image_style=options.get('image_style', 'white'),
                    cancel_event=_task_cancel_event,
                    progress_callback=lambda done, total, current_item: update_status(
                        progress=done,
                        total=total,
                        current_item=current_item or f'已完成 {done}/{total}',
                    ),
                )
                # Pass enhanced options to pipeline if provided
                if options.get('text_fields'):
                    run_kwargs['text_fields'] = options['text_fields']
                if options.get('overwrite_existing') is not None:
                    run_kwargs['overwrite_existing'] = options['overwrite_existing']
                try:
                    result = pipeline.run(**run_kwargs)
                except TypeError as te:
                    # Fallback if pipeline doesn't support newer kwargs
                    if any(token in str(te) for token in ('text_fields', 'overwrite_existing', 'progress_callback',
                                                         'image_scope', 'image_style', 'cancel_event')):
                        add_log('⚠️ Pipeline不支持增强参数，使用默认设置')
                        result = pipeline.run(
                            input_file=actual_input,
                            output_file=output_file,
                            process_images=options.get('process_images', True),
                            process_text=options.get('process_text', True),
                        )
                    else:
                        raise
                if _task_cancel_event.is_set():
                    update_status(result_file=result, cancelled=True, current_item='已取消')
                    add_log(f'⏹️ 第一阶段已取消，当前结果已保存: {os.path.basename(result)}')
                else:
                    update_status(result_file=result, progress=task_status['total'], current_item='完成')
                    add_log(f'✅ 第一阶段完成! 输出: {os.path.basename(result)}')
            finally:
                s1_logger.removeHandler(handler)
                # 清理临时过滤文件
                if actual_input != input_file and os.path.exists(actual_input):
                    try:
                        os.remove(actual_input)
                    except Exception:
                        pass

        elif stage == 2:
            add_log('🚀 第二阶段启动：SP-API提交')
            from stage2_pipeline import Stage2Pipeline
            pipeline = Stage2Pipeline()
            result = pipeline.run(
                input_file=input_file,
                mode=options.get('mode', 'individual'),
            )
            update_status(result_file=result)
            add_log(f'✅ 第二阶段完成! 报告: {result}')

    except Exception as e:
        update_status(error=str(e))
        add_log(f'❌ 任务失败: {str(e)}')
        import traceback
        add_log(f'详细: {traceback.format_exc()[:300]}')
    finally:
        update_status(running=False)
        _finalize_current_task_record()


@app.route('/api/auto-preview')
def auto_preview():
    """自动预览input目录中的文件"""
    filepath = request.args.get('file', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 400

    try:
        processor = ExcelProcessor()
        data = processor.read_input(filepath)
        col_map = processor.detect_columns()

        preview = []
        for item in data[:10]:
            row = {}
            for key, val in item.items():
                if not key.startswith('_'):
                    row[key] = str(val) if val is not None else ''
            preview.append(row)

        return jsonify({
            'success': True,
            'filename': os.path.basename(filepath),
            'filepath': filepath,
            'total_rows': len(data),
            'columns': [h for h in processor.headers],
            'column_mapping': col_map,
            'preview': preview,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/image-process', methods=['POST'])
def image_process():
    """单张图片背景处理 — 使用 /v1/images/edits API"""
    data = request.json
    image_url = data.get('image_url', '')
    bg_style = data.get('bg_style', 'white')
    prompt = data.get('prompt', '')

    cfg = get_config()
    if not cfg.AI_API_KEY:
        return jsonify({'error': '请先在API Configuration中配置API Key'}), 400

    try:
        from core.ai_client import ai_image_edit_url

        # 构建提示
        if not prompt:
            bg_prompts = {
                'white': 'Remove the background and replace with a pure white background (#FFFFFF). Keep the product perfectly intact with clean edges. This is for an Amazon product listing.',
                'lifestyle': 'Place the product in a natural lifestyle setting. Show it being used in a modern, well-lit home environment.',
                'studio': 'Place the product on a clean surface with professional studio lighting. Soft shadows, neutral background.',
                'gradient': 'Place the product on a smooth light gradient background (white to light gray). Professional product photography look.',
            }
            prompt = bg_prompts.get(bg_style, bg_prompts['white'])

        result = ai_image_edit_url(image_url, prompt)

        if result:
            return jsonify({
                'success': True,
                'result_b64': result[:200] + '...',
                'message': 'Image processed successfully',
            })
        else:
            return jsonify({'error': '图片处理失败，AI未返回有效图片'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-copy', methods=['POST'])
def generate_copy():
    """AI生成亚马逊文案"""
    data = request.json
    name = data.get('name', '')
    brand = data.get('brand', '')
    category = data.get('category', '')
    features = data.get('features', '')
    audience = data.get('audience', '')
    price = data.get('price', '')
    competitors = data.get('competitors', '')

    if not name:
        return jsonify({'error': '请输入产品名称'}), 400

    cfg = get_config()
    if not cfg.AI_API_KEY:
        return jsonify({'error': '请先配置API Key'}), 400

    product_info = f"""
Product: {name}
Brand: {brand}
Category: {category}
Key Features: {features}
Target Audience: {audience}
Price Range: {price}
Competitors: {competitors}
""".strip()

    try:
        from core.ai_client import ai_text

        prompt = f"""You are an expert Amazon product listing copywriter. Generate optimized content for the following product.

{product_info}

Generate the following in JSON format:
{{
  "title": "Amazon product title (under 200 chars, include brand + key features + benefit keywords)",
  "bullets": [
    "FEATURE KEYWORD — Detailed benefit description (150-250 chars each)",
    "FEATURE KEYWORD — ...",
    "FEATURE KEYWORD — ...",
    "FEATURE KEYWORD — ...",
    "FEATURE KEYWORD — ..."
  ],
  "description": "Compelling product description with HTML formatting (under 2000 chars)",
  "keywords": "backend search terms separated by spaces (under 250 bytes, no brand names, no repeating title words)"
}}

Rules:
- Title: Brand first, then key features, then benefit keywords
- Bullets: Start each with ALL CAPS keyword, then em dash, then benefit
- Description: Use <b>, <br>, <ul><li> for formatting
- Keywords: Only words NOT in the title, no commas, single spaces
- Write in English for US marketplace
- Be specific, data-driven, conversion-focused

Return ONLY the JSON, no markdown code blocks."""

        content = ai_text(prompt, temperature=0.7, max_tokens=2000)
        # 清理可能的markdown代码块
        if content.startswith('```'):
            content = content.split('\n', 1)[1]
        if content.endswith('```'):
            content = content.rsplit('```', 1)[0]
        content = content.strip()

        result = json.loads(content)
        return jsonify(result)

    except json.JSONDecodeError:
        return jsonify({
            'title': content[:200] if content else '',
            'bullets': [],
            'description': content,
            'keywords': '',
            'warning': 'AI returned non-JSON, showing raw text'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/status')
def api_get_status():
    """获取任务状态"""
    status = get_task_status()
    completed = int(status.get('progress', 0) or 0)
    total = int(status.get('total', 0) or 0)
    running = bool(status.get('running'))
    payload = dict(status)
    payload['completed'] = completed
    payload['current_sku'] = status.get('current_item', '')
    if status.get('cancelled'):
        payload['status'] = 'cancelled'
    elif running and status.get('cancel_requested'):
        payload['status'] = 'cancelling'
    elif running:
        payload['status'] = 'running'
    elif total and completed >= total and not status.get('error'):
        payload['status'] = 'completed'
    else:
        payload['status'] = 'idle'
    return jsonify(payload)


@app.route('/api/tasks')
def api_get_tasks():
    """返回当前任务与最近任务历史。"""
    status = get_task_status()
    with _task_lock:
        history = [dict(record) for record in task_history]

    current_task = None
    if status.get('task_id'):
        current_task = next((record for record in history if record.get('id') == status.get('task_id')), None)
    if current_task is None:
        current_task = next((record for record in history if record.get('status') == 'running'), None)

    return jsonify({
        'current': current_task,
        'history': history,
    })


@app.route('/api/tasks/<task_id>')
def api_get_task_detail(task_id):
    """按任务 ID 获取详细进度。"""
    record = _get_task_record(task_id)
    if not record:
        return jsonify({'error': '任务不存在'}), 404
    return jsonify(record)


@app.route('/api/download/<path:filename>')
def download_file(filename):
    """下载结果文件"""
    if '..' in filename:
        return jsonify({'error': '非法文件路径'}), 400
    filepath = os.path.realpath(os.path.join(config.OUTPUT_DIR, filename))
    allowed_dir = os.path.realpath(config.OUTPUT_DIR)
    if not filepath.startswith(allowed_dir + os.sep) and filepath != allowed_dir:
        return jsonify({'error': '非法文件路径'}), 400
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': '文件不存在'}), 404


@app.route('/api/files')
def list_files():
    """列出输入输出文件"""
    input_files = []
    output_files = []

    for d, lst in [(config.INPUT_DIR, input_files), (config.OUTPUT_DIR, output_files)]:
        if os.path.exists(d):
            for f in os.listdir(d):
                if f.endswith(('.xlsx', '.xls', '.json')) and not f.startswith(('~', '.')):
                    filepath = os.path.join(d, f)
                    lst.append({
                        'name': f,
                        'path': filepath,
                        'size': os.path.getsize(filepath),
                        'modified': datetime.fromtimestamp(
                            os.path.getmtime(filepath)
                        ).strftime('%Y-%m-%d %H:%M'),
                    })

    return jsonify({
        'input': sorted(input_files, key=lambda x: x['modified'], reverse=True),
        'output': sorted(output_files, key=lambda x: x['modified'], reverse=True),
    })


@app.route('/api/config', methods=['GET', 'POST'])
def manage_config():
    """查看/修改配置"""
    if request.method == 'GET':
        return jsonify({
            'text_api_key': '',
            'text_api_key_masked': '***' + config.AI_TEXT_API_KEY[-4:] if config.AI_TEXT_API_KEY else '',
            'text_api_base': config.AI_TEXT_API_BASE,
            'text_endpoint_template': config.AI_TEXT_ENDPOINT_TEMPLATE,
            'text_model': config.AI_TEXT_MODEL,
            'image_api_key': '',
            'image_api_key_masked': '***' + config.AI_IMAGE_API_KEY[-4:] if config.AI_IMAGE_API_KEY else '',
            'image_api_base': config.AI_IMAGE_API_BASE,
            'image_endpoint_template': config.AI_IMAGE_ENDPOINT_TEMPLATE,
            'image_model': config.AI_IMAGE_MODEL,
            'output_dir': config.OUTPUT_DIR,
            'media_store_enabled': bool(config.MEDIA_STORE_ENABLED),
            'media_store_provider': config.MEDIA_STORE_PROVIDER,
            'media_s3_bucket': config.MEDIA_S3_BUCKET,
            'media_s3_region': config.MEDIA_S3_REGION,
            'media_s3_prefix': config.MEDIA_S3_PREFIX,
            'media_s3_submit_scheme': config.MEDIA_S3_SUBMIT_SCHEME,
            'media_s3_preview_base': config.MEDIA_S3_PREVIEW_BASE,
            'default_lang': config.DEFAULT_LANG,
            'batch_limit': config.BATCH_LIMIT,
            'ai_concurrency': config.AI_CONCURRENCY,
            'image_concurrency': config.IMAGE_CONCURRENCY,
            'concurrency': config.AI_CONCURRENCY,
            'amazon_configured': bool(config.AMAZON_CLIENT_ID),
        })
    else:
        data = request.json
        updates = {}
        field_map = {
            'text_api_key': 'AI_TEXT_API_KEY',
            'text_api_base': 'AI_TEXT_API_BASE',
            'text_endpoint_template': 'AI_TEXT_ENDPOINT_TEMPLATE',
            'text_model': 'AI_TEXT_MODEL',
            'image_api_key': 'AI_IMAGE_API_KEY',
            'image_api_base': 'AI_IMAGE_API_BASE',
            'image_endpoint_template': 'AI_IMAGE_ENDPOINT_TEMPLATE',
            'image_model': 'AI_IMAGE_MODEL',
            'output_dir': 'OUTPUT_DIR',
            'media_store_enabled': 'MEDIA_STORE_ENABLED',
            'media_store_provider': 'MEDIA_STORE_PROVIDER',
            'media_s3_bucket': 'MEDIA_S3_BUCKET',
            'media_s3_region': 'MEDIA_S3_REGION',
            'media_s3_prefix': 'MEDIA_S3_PREFIX',
            'media_s3_submit_scheme': 'MEDIA_S3_SUBMIT_SCHEME',
            'media_s3_preview_base': 'MEDIA_S3_PREVIEW_BASE',
            'default_lang': 'DEFAULT_LANG',
            'batch_limit': 'BATCH_LIMIT',
            'ai_concurrency': 'AI_CONCURRENCY',
            'image_concurrency': 'IMAGE_CONCURRENCY',
        }
        for payload_key, env_key in field_map.items():
            if payload_key in data:
                value = data[payload_key]
                if env_key.endswith('_API_KEY') and str(value or '').strip() == '':
                    continue
                if env_key.endswith('_ENDPOINT_TEMPLATE'):
                    value = _normalize_endpoint_template(value)
                if env_key == 'DEFAULT_LANG':
                    value = str(value or 'zh').strip().lower() or 'zh'
                    if value not in ('zh', 'en'):
                        value = 'zh'
                if env_key in ('BATCH_LIMIT', 'AI_CONCURRENCY', 'IMAGE_CONCURRENCY'):
                    try:
                        value = max(1, int(value))
                    except (TypeError, ValueError):
                        continue
                if env_key == 'MEDIA_STORE_ENABLED':
                    value = 'true' if str(value or '').strip().lower() in ('1', 'true', 'yes', 'on') else 'false'
                if env_key == 'MEDIA_STORE_PROVIDER':
                    value = str(value or '').strip().lower()
                if env_key == 'MEDIA_S3_SUBMIT_SCHEME':
                    value = str(value or 's3').strip().lower()
                    if value not in ('s3', 'https'):
                        value = 's3'
                updates[env_key] = value

        if 'concurrency' in data and 'AI_CONCURRENCY' not in updates:
            try:
                updates['AI_CONCURRENCY'] = max(1, int(data.get('concurrency')))
            except (TypeError, ValueError):
                pass

        if 'text_endpoint_template' in data:
            endpoint = str(data.get('text_endpoint_template', '')).strip().lower()
            updates['AI_TEXT_PROTOCOL'] = (
                'gemini_generate_content' if 'generatecontent' in endpoint else 'openai_chat_completions'
            )
        if 'image_endpoint_template' in data:
            endpoint = str(data.get('image_endpoint_template', '')).strip().lower()
            updates['AI_IMAGE_PROTOCOL'] = (
                'gemini_generate_content' if 'generatecontent' in endpoint else 'openai_images'
            )

        _write_env_updates(updates)
        reload_config()
        return jsonify({'success': True, 'message': '配置已更新，已即时生效'})


@app.route('/api/sp-config', methods=['POST'])
def save_sp_config():
    """保存SP-API配置"""
    data = request.json
    updates = {}
    for key in ['AMAZON_CLIENT_ID', 'AMAZON_CLIENT_SECRET', 'AMAZON_REFRESH_TOKEN', 'AMAZON_SELLER_ID']:
        web_key = key.replace('AMAZON_', '').lower()
        if web_key in data and data[web_key]:
            updates[key] = data[web_key]
    _write_env_updates(updates)
    reload_config()
    return jsonify({'success': True, 'message': 'SP-API配置已保存，已即时生效'})


@app.route('/api/config/test', methods=['POST'])
def test_ai_config():
    """测试当前 AI 配置是否可用。"""
    try:
        from core.ai_client import ai_text

        result = ai_text('请只回复 OK', temperature=0.0, max_tokens=8)
        if result:
            return jsonify({'success': True, 'message': f'AI连接成功: {result[:60]}'})
        return jsonify({'success': False, 'message': 'AI请求已发送，但未返回有效文本'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'AI连接失败: {str(e)}'})


@app.route('/api/self-check', methods=['POST'])
def run_self_check():
    """执行一键环境自检。"""
    data = request.json or {}
    account_id = str(data.get('account_id', '') or '').strip()
    checks = []

    def append_check(name: str, status: str, message: str, detail: str = ''):
        checks.append({
            'name': name,
            'status': status,
            'message': message,
            'detail': detail,
        })

    # 输出目录
    try:
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        probe_path = os.path.join(config.OUTPUT_DIR, f'.write_probe_{uuid4().hex[:8]}.tmp')
        with open(probe_path, 'w', encoding='utf-8') as fh:
            fh.write('ok')
        os.remove(probe_path)
        append_check('output_dir', 'pass', '输出目录可写', config.OUTPUT_DIR)
    except Exception as e:
        append_check('output_dir', 'fail', '输出目录不可写', str(e))

    # 文本AI
    try:
        from core.ai_client import ai_text, ai_image_generate
        text_result = ai_text('请只回复 OK', temperature=0.0, max_tokens=8)
        if text_result:
            append_check('text_ai', 'pass', '文本 AI 接口可用', text_result[:60])
        else:
            append_check('text_ai', 'fail', '文本 AI 未返回有效结果')
    except Exception as e:
        append_check('text_ai', 'fail', '文本 AI 接口不可用', str(e))

    # 图片AI
    try:
        image_result = ai_image_generate('Generate a plain white square on a clean background.', size='512x512')
        if image_result:
            append_check('image_ai', 'pass', '图片 AI 接口可用')
        else:
            append_check('image_ai', 'fail', '图片 AI 未返回有效结果')
    except Exception as e:
        append_check('image_ai', 'fail', '图片 AI 接口不可用', str(e))

    # Amazon 账号
    try:
        from amazon.accounts import AccountManager
        mgr = AccountManager()
        acc = mgr.get_account(account_id) if account_id else mgr.get_default_account()
        if not acc:
            append_check('amazon_account', 'warn', '未配置默认亚马逊账号')
        else:
            result = mgr.test_connection(acc.get('seller_id'))
            append_check(
                'amazon_account',
                'pass' if result.get('success') else 'fail',
                result.get('message', '账号检测完成'),
                acc.get('name', acc.get('seller_id', '')),
            )
    except Exception as e:
        append_check('amazon_account', 'fail', '亚马逊账号检测失败', str(e))

    # xls 兼容写回
    xls_support = _detect_xls_preserve_support()
    if xls_support.get('supported'):
        append_check('xls_preserve', 'pass', '`.xls` 保样式写回可用', xls_support.get('app', ''))
    else:
        append_check('xls_preserve', 'warn', '`.xls` 将回退为兼容模式写回', '建议将模板升级为 .xlsx')

    # 媒体存储
    try:
        media_store = get_media_store()
        result = media_store.healthcheck()
        if not getattr(media_store, 'enabled', lambda: False)():
            append_check('media_store', 'warn', result.get('message', '媒体存储未启用'))
        else:
            append_check(
                'media_store',
                'pass' if result.get('success') else 'fail',
                result.get('message', '媒体存储检测完成'),
                f"bucket={result.get('bucket', '')} prefix={result.get('prefix', '')}".strip(),
            )
    except Exception as e:
        append_check('media_store', 'fail', '媒体存储检测失败', str(e))

    overall_status = 'pass'
    if any(check['status'] == 'fail' for check in checks):
        overall_status = 'fail'
    elif any(check['status'] == 'warn' for check in checks):
        overall_status = 'warn'

    return jsonify({
        'success': overall_status != 'fail',
        'status': overall_status,
        'checks': checks,
    })


# ===== 账号管理 API =====

@app.route('/api/accounts', methods=['GET'])
def list_accounts():
    """列出所有亚马逊账号"""
    try:
        from amazon.accounts import AccountManager
        mgr = AccountManager()
        return jsonify({
            'accounts': mgr.list_accounts(),
            'marketplaces': mgr.get_marketplace_options(),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/accounts', methods=['POST'])
def add_account():
    """添加亚马逊账号"""
    try:
        from amazon.accounts import AccountManager
        mgr = AccountManager()
        data = request.json
        success = mgr.add_account(data)
        if success:
            return jsonify({'success': True, 'message': '账号添加成功'})
        else:
            return jsonify({'error': '账号已存在'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/accounts/test', methods=['POST'])
def test_account():
    """测试账号连接"""
    try:
        from amazon.accounts import AccountManager
        mgr = AccountManager()
        seller_id = request.json.get('seller_id')
        result = mgr.test_connection(seller_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/accounts/default', methods=['POST'])
def set_default_account():
    """设置默认账号"""
    try:
        from amazon.accounts import AccountManager
        mgr = AccountManager()
        seller_id = request.json.get('seller_id')
        mgr.set_default(seller_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ===== 对比预览 API =====

@app.route('/api/comparison')
def get_comparison():
    """获取AI处理前后的对比数据"""
    filepath = request.args.get('file', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 400

    try:
        processor = ExcelProcessor()
        data = processor.read_input(filepath)
        col_map = processor.detect_columns()

        comparisons = []
        for item in data:
            comp = {
                'sku': item.get(col_map.get('sku', ''), item.get('SKU', '')),
                'original': {},
                'ai_generated': {},
                'status': item.get('submit_status', 'PENDING'),
            }

            # 原始值 — 兼容输入Excel和输出对比Excel
            title_col = col_map.get('title', '')
            comp['original']['title'] = (
                item.get(title_col, '') if title_col else ''
            ) or item.get('原始标题(中文)', '') or item.get('item_name', '')

            desc_col = col_map.get('description', '')
            comp['original']['description'] = (
                item.get(desc_col, '') if desc_col else ''
            ) or item.get('原始描述', '') or item.get('product_description', '')

            # AI值 — 兼容多种列名格式
            comp['ai_generated']['title'] = (
                item.get('AI标题', '') or
                item.get('→ AI标题(英文)', '') or
                item.get('AI标题(英文)', '')
            )
            comp['ai_generated']['description'] = (
                item.get('AI商品描述', '') or
                item.get('→ AI描述(英文)', '') or
                item.get('AI描述(英文)', '')
            )
            comp['ai_generated']['bullets'] = [
                (item.get(f'AI卖点{i}', '') or item.get(f'→ AI卖点{i}', ''))
                for i in range(1, 6)
            ]
            comp['ai_generated']['keywords'] = (
                item.get('AI搜索关键词', '') or
                item.get('AI搜索关键词', '')
            )

            # 状态
            comp['status'] = (
                item.get('submit_status', '') or
                item.get('提交状态', '') or
                'PENDING'
            )

            # 图片路径
            # 原始图片URL
            img_col = col_map.get('image_url', '')
            comp['original']['image'] = (
                item.get(img_col, '') if img_col else ''
            ) or item.get('原始主图', '') or ''

            # AI处理后的图片
            ai_img_path = (
                item.get('AI主图路径', '') or
                item.get('→ AI主图(白底)', '') or ''
            )
            if ai_img_path and os.path.exists(ai_img_path):
                # 返回相对URL供前端访问
                img_name = os.path.basename(ai_img_path)
                comp['ai_generated']['image'] = f'/api/output-image/{img_name}'
            else:
                comp['ai_generated']['image'] = ''

            comparisons.append(comp)

        return jsonify({
            'total': len(comparisons),
            'comparisons': comparisons,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ===== 字段验证 API =====

@app.route('/api/products')
def get_products():
    """读取Excel并按parent_sku分组，返回商品→SKU两层结构"""
    filepath = request.args.get('file', '')
    if not filepath or not os.path.exists(filepath):
        return jsonify({'error': '文件不存在'}), 400
    try:
        cached = read_excel_cached(filepath)
        data = cached['data']
        col_map = cached['col_map']
        headers = cached['headers']
        template_meta, template_definition = _load_template_definition_for_file(filepath)
        template_summary = template_definition_summary(template_definition)

        title_col = col_map.get('title', '')
        sku_col = col_map.get('sku', '')

        # 检测父SKU列和变体主题列
        parent_col = None
        variation_col = None
        for h in headers:
            hl = str(h).lower()
            if hl in ('parent_sku', '父sku', 'parent sku', 'parent_asin', '父asin'):
                parent_col = h
            if hl in ('variation_theme', '变体主题'):
                variation_col = h

        # 检测品牌、价格、主图列
        brand_col = col_map.get('brand', '')
        price_col = col_map.get('price', '')
        img_col = col_map.get('image_url', '')

        def cell_text(row, header_name):
            return str(row.get(header_name, '') or '').strip() if header_name else ''

        groups = {}
        for i, item in enumerate(data):
            parent_key = (str(item.get(parent_col, '') or '').strip() if parent_col else '') or f'__solo_{i}'
            sku = str(item.get(sku_col, '') or item.get('SKU', '') or f'ROW-{i+1}').strip()
            title = str(item.get(title_col, '') or item.get('item_name', '') or item.get('标题', '') or '').strip()
            brand = str(item.get(brand_col, '') or '').strip() if brand_col else ''
            price_val = item.get(price_col, '') if price_col else ''
            main_image = str(item.get(img_col, '') or '').strip() if img_col else ''
            variation_theme = str(item.get(variation_col, '') or '').strip() if variation_col else ''

            is_solo = parent_key.startswith('__solo_')

            if parent_key not in groups:
                groups[parent_key] = {
                    'type': 'standalone' if is_solo else 'parent',
                    'parent_sku': sku if is_solo else parent_key,
                    'title': title,
                    'brand': brand,
                    'variation_theme': variation_theme,
                    'skus': []
                }
            # 更新group-level品牌和变体（以第一个非空值为准）
            if brand and not groups[parent_key]['brand']:
                groups[parent_key]['brand'] = brand
            if variation_theme and not groups[parent_key]['variation_theme']:
                groups[parent_key]['variation_theme'] = variation_theme

            raw = {k: str(v) for k, v in item.items() if not k.startswith('_') and v is not None and str(v).strip()}

            # 构建AI处理后的图片URL
            ai_image_result = _build_ai_image_result_data(item)
            ai_main_preview = ai_image_result.get('ai_main_image_preview', '')
            ai_media_locator = ai_image_result.get('ai_media_locator', '')
            ai_public_image_url = ai_image_result.get('ai_public_image_url', '')
            main_image_preview_url = _resolve_current_image_preview(
                main_image,
                ai_locator=ai_media_locator,
                ai_preview=ai_main_preview,
            )

            ai_fields = ['AI标题', 'AI商品描述', 'AI搜索关键词', 'AI主图路径']
            ai_fields.extend([f'AI卖点{i}' for i in range(1, 6)])
            ai_fields.extend([f'AI副图{i}路径' for i in range(2, 10)])
            ai_status = _derive_ai_status(item, ai_fields)

            groups[parent_key]['skus'].append({
                'sku': sku,
                'title': title,
                'item_name': title,
                'brand': brand,
                'manufacturer': cell_text(item, col_map.get('manufacturer', '')),
                'model_number': cell_text(item, col_map.get('model_number', '')) or cell_text(item, col_map.get('part_number', '')),
                'product_type': cell_text(item, col_map.get('product_type', '')),
                'item_type_keyword': cell_text(item, col_map.get('item_type_keyword', '')),
                'price': str(price_val) if price_val is not None else '',
                'list_price': cell_text(item, col_map.get('list_price', '')),
                'currency': cell_text(item, col_map.get('currency', '')),
                'quantity': cell_text(item, col_map.get('quantity', '')),
                'condition_type': cell_text(item, col_map.get('condition_type', '')),
                'fulfillment_channel': cell_text(item, col_map.get('fulfillment_channel', '')),
                'main_image_url': main_image,
                'main_image_preview_url': main_image_preview_url,
                'ai_main_image': ai_image_result.get('ai_main_image', ai_main_preview),
                'ai_main_image_preview': ai_main_preview,
                'ai_image_url': ai_image_result.get('ai_image_url', ai_public_image_url or ai_main_preview),
                'ai_media_locator': ai_media_locator,
                'ai_public_image_url': ai_image_result.get('ai_public_image_url', ai_public_image_url),
                'ai_upload_status': ai_image_result.get('ai_upload_status', ''),
                'ai_upload_error': ai_image_result.get('ai_upload_error', ''),
                'bullet_point_1': cell_text(item, col_map.get('bullet_point_1', '')),
                'bullet_point_2': cell_text(item, col_map.get('bullet_point_2', '')),
                'bullet_point_3': cell_text(item, col_map.get('bullet_point_3', '')),
                'bullet_point_4': cell_text(item, col_map.get('bullet_point_4', '')),
                'bullet_point_5': cell_text(item, col_map.get('bullet_point_5', '')),
                'product_description': cell_text(item, col_map.get('description', '')),
                'generic_keywords': cell_text(item, col_map.get('keywords', '')),
                'upc': cell_text(item, col_map.get('upc', '')),
                'external_product_id_type': cell_text(item, col_map.get('external_product_id_type', '')),
                'product_identity_mode': str(item.get(col_map.get('product_identity_mode', ''), '') or item.get('product_identity_mode', '') or '').strip(),
                'material': cell_text(item, col_map.get('material', '')),
                'country_of_origin': cell_text(item, col_map.get('country_of_origin', '')),
                'included_components': cell_text(item, col_map.get('included_components', '')),
                'care_instructions': cell_text(item, col_map.get('care_instructions', '')),
                'special_feature': cell_text(item, col_map.get('special_feature', '')),
                'number_of_items': cell_text(item, col_map.get('number_of_items', '')),
                'unit_count': cell_text(item, col_map.get('unit_count', '')),
                'ai_title': str(item.get('AI标题', '') or item.get('→ AI标题(英文)', '') or '').strip(),
                'ai_bullet_1': str(item.get('AI卖点1', '') or item.get('→ AI卖点1', '') or '').strip(),
                'ai_bullet_2': str(item.get('AI卖点2', '') or item.get('→ AI卖点2', '') or '').strip(),
                'ai_bullet_3': str(item.get('AI卖点3', '') or item.get('→ AI卖点3', '') or '').strip(),
                'ai_bullet_4': str(item.get('AI卖点4', '') or item.get('→ AI卖点4', '') or '').strip(),
                'ai_bullet_5': str(item.get('AI卖点5', '') or item.get('→ AI卖点5', '') or '').strip(),
                'ai_description': str(item.get('AI商品描述', '') or item.get('→ AI描述(英文)', '') or '').strip(),
                'ai_keywords': str(item.get('AI搜索关键词', '') or '').strip(),
                'color': str(item.get(col_map.get('color', ''), '') or '').strip(),
                'size': str(item.get(col_map.get('size', ''), '') or '').strip(),
                'item_weight': cell_text(item, col_map.get('weight', '')),
                'item_weight_unit': cell_text(item, col_map.get('item_weight_unit', '')),
                'item_length': cell_text(item, col_map.get('item_length', '')),
                'item_width': cell_text(item, col_map.get('item_width', '')),
                'item_height': cell_text(item, col_map.get('item_height', '')),
                'dimension_unit': cell_text(item, col_map.get('dimension_unit', '')),
                'package_weight': cell_text(item, col_map.get('package_weight', '')),
                'package_weight_unit': cell_text(item, col_map.get('package_weight_unit', '')),
                'package_length': cell_text(item, col_map.get('package_length', '')),
                'package_width': cell_text(item, col_map.get('package_width', '')),
                'package_height': cell_text(item, col_map.get('package_height', '')),
                'ai_status': ai_status,
                'preview_status': str(item.get('preview_status', '') or item.get('预览状态', '') or '').strip(),
                'preview_message': str(item.get('preview_message', '') or item.get('预览信息', '') or '').strip(),
                'preview_time': str(item.get('preview_time', '') or item.get('预览时间', '') or '').strip(),
                'preview_account': str(item.get('preview_account', '') or item.get('预览账号', '') or '').strip(),
                'submit_status': str(item.get('submit_status', '') or item.get('提交状态', '') or 'PENDING'),
                'asin': str(item.get('asin', '') or item.get('ASIN', '') or '').strip(),
                'submission_id': str(item.get('submission_id', '') or item.get('提交ID', '') or '').strip(),
                'submit_time': str(item.get('submit_time', '') or item.get('提交时间', '') or '').strip(),
                'submit_message': str(item.get('submit_message', '') or item.get('提交信息', '') or item.get('问题详情', '') or '').strip(),
                'validation_status': str(item.get('validation_status', '') or '').strip(),
                'validation_errors': str(item.get('validation_errors', '') or '').strip(),
                'validation_warnings': str(item.get('validation_warnings', '') or '').strip(),
                'listing_check_status': str(item.get('listing_check_status', '') or item.get('缺项诊断状态', '') or '').strip(),
                'listing_check_summary': str(item.get('listing_check_summary', '') or item.get('缺项诊断摘要', '') or '').strip(),
                'listing_check_missing_fields': str(item.get('listing_check_missing_fields', '') or item.get('缺失字段清单', '') or '').strip(),
                'listing_check_issues': str(item.get('listing_check_issues', '') or item.get('缺项诊断问题', '') or '').strip(),
                'listing_check_time': str(item.get('listing_check_time', '') or item.get('缺项诊断时间', '') or '').strip(),
                'listing_check_account': str(item.get('listing_check_account', '') or item.get('缺项诊断账号', '') or '').strip(),
                'template_id': str(item.get('template_id', '') or item.get('模板ID', '') or template_meta.get('template_id', '') or '').strip(),
                'template_product_type': str(item.get('template_product_type', '') or item.get('模板产品类型', '') or template_summary.get('product_type', '') or '').strip(),
                'template_variation_mode': str(item.get('template_variation_mode', '') or item.get('模板变体模式', '') or template_summary.get('variation_mode', '') or '').strip(),
                'template_required_total': str(item.get('template_required_total', '') or item.get('模板必填总数', '') or template_summary.get('required_total', '') or '').strip(),
                'template_required_filled': str(item.get('template_required_filled', '') or item.get('模板必填已填', '') or '').strip(),
                'template_required_missing_count': str(item.get('template_required_missing_count', '') or item.get('模板必填缺失数', '') or '').strip(),
                'template_required_missing_fields': str(item.get('template_required_missing_fields', '') or item.get('模板必填缺失字段', '') or '').strip(),
                'template_recommended_missing_fields': str(item.get('template_recommended_missing_fields', '') or item.get('模板建议补充字段', '') or '').strip(),
                'template_blocking_issues': str(item.get('template_blocking_issues', '') or item.get('模板阻断问题', '') or '').strip(),
                'template_ready_to_submit': str(item.get('template_ready_to_submit', '') or item.get('模板提交就绪', '') or '').strip(),
                '_raw': raw,
                **{f'ai_image_{slot}': ai_image_result.get(f'ai_image_{slot}', '') for slot in range(2, 10)},
                **{f'ai_image_{slot}_preview': ai_image_result.get(f'ai_image_{slot}_preview', '') for slot in range(2, 10)},
                **{f'ai_media_locator_{slot}': ai_image_result.get(f'ai_media_locator_{slot}', '') for slot in range(2, 10)},
                **{f'ai_public_image_{slot}': ai_image_result.get(f'ai_public_image_{slot}', '') for slot in range(2, 10)},
                **{f'ai_image_{slot}_upload_status': ai_image_result.get(f'ai_image_{slot}_upload_status', '') for slot in range(2, 10)},
                **{f'ai_image_{slot}_upload_error': ai_image_result.get(f'ai_image_{slot}_upload_error', '') for slot in range(2, 10)},
                **{
                    f'image_{slot}_preview_url': _resolve_current_image_preview(
                        str(item.get(col_map.get(f'image_{slot}', ''), '') or item.get(f'other_image_url_{slot-1}', '') or '').strip(),
                        ai_locator=ai_image_result.get(f'ai_media_locator_{slot}', ''),
                        ai_preview=ai_image_result.get(f'ai_image_{slot}_preview', ''),
                    ) for slot in range(2, 10)
                },
            })

        return jsonify({
            'total_products': len(groups),
            'total_skus': len(data),
            'columns': headers[:30],
            'column_mapping': col_map,
            'template': template_summary,
            'products': list(groups.values()),
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/output-image/<path:filename>')
def serve_output_image(filename):
    """提供输出图片的静态文件服务"""
    if '..' in filename:
        return jsonify({'error': '非法文件路径'}), 400
    images_dir = os.path.join(config.OUTPUT_DIR, 'images')
    filepath = os.path.realpath(os.path.join(images_dir, filename))
    allowed_dir = os.path.realpath(images_dir)
    if not filepath.startswith(allowed_dir + os.sep) and filepath != allowed_dir:
        return jsonify({'error': '非法文件路径'}), 400
    if os.path.exists(filepath):
        return send_file(filepath, mimetype='image/jpeg')
    return jsonify({'error': '图片不存在'}), 404


def _apply_validation_results_to_file(input_file: str, results: list):
    """将校验状态回写到 Excel，便于刷新后继续查看。"""
    updates_by_sku = {}
    for result in results:
        sku = str(result.get('sku', '') or '').strip()
        if not sku:
            continue

        issues = result.get('issues', []) or []
        error_messages = [issue.get('message', '') for issue in issues if issue.get('level') == 'error']
        warning_messages = [issue.get('message', '') for issue in issues if issue.get('level') == 'warning']

        if error_messages:
            status = 'fail'
        elif warning_messages:
            status = 'warn'
        else:
            status = 'pass'

        updates_by_sku[sku] = {
            'validation_status': status,
            'validation_errors': '; '.join(msg for msg in error_messages if msg),
            'validation_warnings': '; '.join(msg for msg in warning_messages if msg),
        }

    if updates_by_sku:
        _persist_bulk_row_updates(input_file, updates_by_sku)


def _apply_preview_results_to_file(input_file: str, headers: list, results: list, account_name: str):
    """将 Amazon 预览结果回写到 Excel，便于后续按结果筛选与修复。"""
    preview_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    updates_by_sku = {}
    for result in results:
        sku = str(result.get('sku', '') or '').strip()
        if not sku:
            continue
        updates_by_sku[sku] = _build_preview_persist_updates(headers, result, preview_time, account_name)

    if updates_by_sku:
        _persist_bulk_row_updates(input_file, updates_by_sku)


@app.route('/api/validate', methods=['POST'])
def validate_fields():
    """上架前校验端点 — 检查必填字段、长度限制、格式等"""
    try:
        data = request.json
        input_file = data.get('input_file', '')
        selected_skus = data.get('skus', [])
        product_type_override = data.get('product_type', '')
        from amazon.mapper import FieldMapper

        mapper = FieldMapper()

        # 支持两种模式：从文件读取 或 直接传入products
        if input_file and os.path.exists(input_file):
            processor = ExcelProcessor()
            rows = processor.read_input(input_file)
            col_map = processor.detect_columns()

            products = []
            for item in rows:
                mapped = mapper.map_excel_row(item, col_map)
                if product_type_override:
                    mapped['product_type'] = product_type_override
                products.append(mapped)

            # 按SKU过滤
            if selected_skus:
                sku_set = set(selected_skus)
                products = [p for p in products if p.get('sku', '') in sku_set]
        else:
            products = data.get('products', [])

        results = []
        for product in products:
            validation = mapper.validate_required_fields(product)
            issues = []

            for message in validation['errors']:
                issues.append({'level': 'error', 'field': '', 'message': message})
            for message in validation['warnings']:
                issues.append({'level': 'warning', 'field': '', 'message': message})
            for message in validation['info']:
                issues.append({'level': 'info', 'field': '', 'message': message})

            results.append({
                'sku': product.get('sku', 'N/A'),
                'valid': validation['valid'],
                'issues': issues,
                'schema_required_missing': validation.get('schema_required_missing', []),
            })

        if input_file and os.path.exists(input_file):
            _apply_validation_results_to_file(input_file, results)

        passed = sum(1 for r in results if r['valid'])
        failed = len(results) - passed
        _record_instant_task(
            kind='validate',
            title='批量校验',
            status='completed',
            input_file=input_file,
            total=len(results),
            success=passed,
            failed=failed,
            message=f'校验完成：{passed} 通过，{failed} 失败',
        )

        return jsonify({
            'total': len(results),
            'valid': passed,
            'invalid': failed,
            'results': results,
        })
    except Exception as e:
        logger.error(f'校验失败: {e}')
        _record_instant_task(
            kind='validate',
            title='批量校验',
            status='failed',
            input_file=input_file,
            message=str(e),
            error=str(e),
        )
        return jsonify({'error': str(e)}), 500


@app.route('/api/listing-check', methods=['POST'])
def listing_check():
    """按 SKU 汇总本地校验、Amazon 预览、后台 issues 与媒体连通性。"""
    input_file = ''
    try:
        data = request.json or {}
        input_file = data.get('input_file', '') or data.get('file', '')
        selected_skus = [str(sku or '').strip() for sku in (data.get('skus', []) or []) if str(sku or '').strip()]
        account_id = str(data.get('account_id', '') or '').strip()

        if not input_file or not os.path.exists(input_file):
            return jsonify({'error': '文件不存在'}), 400
        if not selected_skus:
            return jsonify({'error': '没有选中的SKU'}), 400

        processor = ExcelProcessor()
        rows = processor.read_input(input_file)
        col_map = processor.detect_columns()
        headers = list(processor.headers)
        sku_col = col_map.get('sku', '')
        sku_set = set(selected_skus)
        matched = [
            item for item in rows
            if str(item.get(sku_col, '') or item.get('SKU', '')).strip() in sku_set
        ]
        if not matched:
            return jsonify({'error': f'未匹配到 {len(selected_skus)} 个 SKU'}), 400

        account, mapper, listings_api, context_message = _build_listing_api_context(account_id)

        results = []
        account_name = account.get('name', account.get('seller_id', '')) if account else ''
        for item in matched:
            product = mapper.map_excel_row(item, col_map)
            result = _run_listing_check_for_product(product, mapper=mapper, listings_api=listings_api)
            if context_message:
                result['account_message'] = context_message
            result['account_name'] = account_name
            results.append(result)
            if listings_api is not None:
                time.sleep(0.2)

        _apply_listing_check_results_to_file(input_file, headers, results, account_name)

        pass_count = sum(1 for item in results if item.get('status') == 'pass')
        warn_count = sum(1 for item in results if item.get('status') == 'warn')
        fail_count = sum(1 for item in results if item.get('status') == 'fail')

        _record_instant_task(
            kind='listing_check',
            title='缺项诊断',
            status='completed',
            input_file=input_file,
            total=len(results),
            success=pass_count,
            failed=fail_count,
            account=account_name,
            message=f'诊断完成：{pass_count} 通过，{warn_count} 警告，{fail_count} 失败',
            warning=context_message,
        )

        return jsonify({
            'success': True,
            'total': len(results),
            'pass': pass_count,
            'warn': warn_count,
            'fail': fail_count,
            'account': account_name,
            'account_message': context_message,
            'results': results,
        })
    except Exception as e:
        logger.error(f'缺项诊断失败: {e}')
        _record_instant_task(
            kind='listing_check',
            title='缺项诊断',
            status='failed',
            input_file=input_file,
            message=str(e),
            error=str(e),
        )
        return jsonify({'error': str(e)}), 500


def _unique_upload_path(target_dir: str, original_name: str) -> str:
    base_name = secure_filename(original_name or '') or f'upload_{uuid4().hex[:8]}.xlsx'
    stem, ext = os.path.splitext(base_name)
    ext = ext or '.xlsx'
    candidate = os.path.join(target_dir, base_name)
    while os.path.exists(candidate):
        candidate = os.path.join(target_dir, f'{stem}_{uuid4().hex[:6]}{ext}')
    return candidate


def _load_template_definition_for_file(filepath: str):
    meta = read_template_metadata(filepath)
    template_id = str(meta.get('template_id', '') or '').strip()
    definition = None
    if template_id:
        try:
            definition = load_template_definition(template_id)
        except Exception:
            definition = None
    return meta, definition


def _template_column_issues(headers: list, template_definition: dict) -> list:
    if not template_definition:
        return []
    header_set = {str(header or '').strip() for header in headers if str(header or '').strip()}
    issues = []
    for column in template_definition.get('columns', []) or []:
        key = str(column.get('key', '') or '').strip()
        if not key or key in header_set:
            continue
        level = str(column.get('level', 'optional') or 'optional').strip()
        issues.append({
            'key': key,
            'label_zh': str(column.get('label_zh', '') or key).strip(),
            'level': level,
            'message': f"模板列缺失: {str(column.get('label_zh', '') or key).strip()} ({key})",
        })
    return issues


def _build_template_row_updates(headers, result_entry: dict, template_meta: dict) -> dict:
    template_data = result_entry.get('template', {}) or {}
    required_missing = [
        f"{item.get('label_zh', item.get('key', ''))} ({item.get('key', '')})"
        for item in template_data.get('required_missing', []) or []
        if str(item.get('key', '') or '').strip()
    ]
    recommended_missing = [
        f"{item.get('label_zh', item.get('key', ''))} ({item.get('key', '')})"
        for item in template_data.get('recommended_missing', []) or []
        if str(item.get('key', '') or '').strip()
    ]
    blocking_issues = [str(msg or '').strip() for msg in template_data.get('blocking_issues', []) or [] if str(msg or '').strip()]
    updates = {
        _pick_existing_header(headers, 'template_id', '模板ID', default='template_id'): str(template_data.get('template_id', '') or template_meta.get('template_id', '') or '').strip(),
        _pick_existing_header(headers, 'template_product_type', '模板产品类型', default='template_product_type'): str(template_meta.get('product_type', '') or '').strip(),
        _pick_existing_header(headers, 'template_variation_mode', '模板变体模式', default='template_variation_mode'): str(template_meta.get('variation_mode', '') or '').strip(),
        _pick_existing_header(headers, 'template_required_total', '模板必填总数', default='template_required_total'): int(template_data.get('required_total', 0) or 0),
        _pick_existing_header(headers, 'template_required_filled', '模板必填已填', default='template_required_filled'): int(template_data.get('required_filled', 0) or 0),
        _pick_existing_header(headers, 'template_required_missing_count', '模板必填缺失数', default='template_required_missing_count'): len(required_missing),
        _pick_existing_header(headers, 'template_required_missing_fields', '模板必填缺失字段', default='template_required_missing_fields'): '; '.join(required_missing),
        _pick_existing_header(headers, 'template_recommended_missing_fields', '模板建议补充字段', default='template_recommended_missing_fields'): '; '.join(recommended_missing),
        _pick_existing_header(headers, 'template_blocking_issues', '模板阻断问题', default='template_blocking_issues'): '; '.join(blocking_issues),
        _pick_existing_header(headers, 'template_ready_to_submit', '模板提交就绪', default='template_ready_to_submit'): 'yes' if not blocking_issues else 'no',
    }
    return updates


def _apply_template_results_to_file(input_file: str, headers: list, results: list, template_meta: dict):
    updates_by_sku = {}
    for result in results:
        sku = str(result.get('sku', '') or '').strip()
        if not sku:
            continue
        row_updates = _build_template_row_updates(headers, result, template_meta)
        row_updates.update(_build_listing_check_persist_updates(headers, result, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), result.get('account_name', '') or ''))
        row_updates.update(_build_preview_persist_updates(
            headers,
            {
                'status': result.get('preview', {}).get('status', ''),
                'valid': str(result.get('preview', {}).get('status', '') or '').strip().upper() == 'VALID',
                'errors': [
                    issue.get('message', '')
                    for issue in (result.get('preview', {}).get('issues', []) or [])
                    if issue.get('level') == 'error'
                ],
                'warnings': [
                    issue.get('message', '')
                    for issue in (result.get('preview', {}).get('issues', []) or [])
                    if issue.get('level') != 'error'
                ],
            },
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            result.get('account_name', '') or '',
        ))
        updates_by_sku[sku] = row_updates
    if updates_by_sku:
        _persist_bulk_row_updates(input_file, updates_by_sku)


def _build_template_column_lookup(template_definition: dict) -> dict:
    lookup = {}
    for column in (template_definition or {}).get('columns', []) or []:
        key = str(column.get('key', '') or '').strip()
        if not key:
            continue
        lookup[key] = {
            'title': str(column.get('label_zh', '') or column.get('label_en', '') or key).strip() or key,
            'description': str(column.get('description', '') or '').strip(),
            'group': str(column.get('group', '') or 'other').strip() or 'other',
        }
    return lookup


def _merge_template_diagnostic(result_entry: dict, template_eval: dict, column_issues: list):
    template_info = dict(template_eval or {})
    blocking = list(template_info.get('blocking_issues') or [])
    if column_issues:
        blocking.extend(issue.get('message', '') for issue in column_issues if issue.get('message'))
    for item in result_entry.get('missing_fields', []) or []:
        name = str(item.get('name', '') or '').strip()
        title = str(item.get('title', '') or name).strip() or name
        if name:
            blocking.append(f"Amazon 预览缺失：{title} ({name})")
    template_info['blocking_issues'] = [msg for msg in blocking if str(msg or '').strip()]
    result_entry['template'] = template_info
    summary = summarize_template_issues(result_entry)
    if summary:
        existing = str(result_entry.get('summary_text', '') or '').strip()
        result_entry['summary_text'] = f"{summary}；{existing}" if existing else summary
    if template_info.get('blocking_issues'):
        result_entry['status'] = 'fail'
    return result_entry


def _execute_template_diagnosis(task_id: str, input_file: str, account_id: str = '', selected_skus: list = None):
    stages = ['解析上传文件', '匹配字段', '本地校验', 'Amazon 预览', '生成诊断结果']
    try:
        _set_task_stage(task_id, stages[0], 1, stages, progress=0, total=1, current_item='读取模板文件')
        processor = ExcelProcessor()
        rows = processor.read_input(input_file)
        col_map = processor.detect_columns()
        headers = list(processor.headers)
        meta, template_definition = _load_template_definition_for_file(input_file)
        column_issues = _template_column_issues(headers, template_definition)
        template_column_lookup = _build_template_column_lookup(template_definition)

        sku_col = col_map.get('sku', '')
        sku_set = {str(sku or '').strip() for sku in (selected_skus or []) if str(sku or '').strip()}
        if sku_set:
            rows = [
                item for item in rows
                if str(item.get(sku_col, '') or item.get('SKU', '')).strip() in sku_set
            ]
        if not rows:
            raise ValueError('没有可诊断的SKU')

        _set_task_stage(task_id, stages[1], 2, stages, progress=0, total=len(rows), current_item='解析模板字段')
        account, mapper, listings_api, context_message = _build_listing_api_context(account_id)
        account_name = account.get('name', account.get('seller_id', '')) if account else ''
        results = []
        for idx, row in enumerate(rows, start=1):
            sku = str(row.get(sku_col, '') or row.get('SKU', '') or f'ROW-{idx}').strip()
            current_item = f'{sku} · {idx}/{len(rows)}'
            _set_task_stage(task_id, stages[2], 3, stages, progress=idx - 1, total=len(rows), current_item=current_item)

            if template_definition and not str(row.get(col_map.get('product_type', ''), '') or row.get('product_type', '') or '').strip():
                row['product_type'] = template_definition.get('product_type', '')

            product = mapper.map_excel_row(row, col_map)
            if template_definition and not product.get('product_type'):
                product['product_type'] = template_definition.get('product_type', '')

            template_eval = evaluate_template_row(row, template_definition, col_map=col_map)

            _set_task_stage(task_id, stages[3], 4, stages, progress=idx - 1, total=len(rows), current_item=current_item)
            result = _run_listing_check_for_product(product, mapper=mapper, listings_api=listings_api)
            result['account_name'] = account_name
            if context_message:
                result['account_message'] = context_message
            for missing in result.get('missing_fields', []) or []:
                name = str(missing.get('name', '') or '').strip()
                if not name:
                    continue
                meta_info = template_column_lookup.get(name, {})
                if meta_info.get('title'):
                    missing['title'] = meta_info['title']
                if meta_info.get('description'):
                    missing['description'] = meta_info['description']
                if meta_info.get('group'):
                    missing['group'] = meta_info['group']
            _merge_template_diagnostic(result, template_eval, column_issues)
            results.append(result)

        _set_task_stage(task_id, stages[4], 5, stages, progress=len(rows), total=len(rows), current_item='写回诊断结果')
        _apply_template_results_to_file(input_file, headers, results, meta)
        preview_missing_fields = []
        for result in results:
            for item in result.get('missing_fields', []) or []:
                if str(item.get('source', '') or '').strip() == 'amazon_preview':
                    preview_missing_fields.append(item)
        if preview_missing_fields and template_definition:
            update_template_overlay(
                product_type=template_definition.get('product_type', ''),
                marketplace=template_definition.get('marketplace', DEFAULT_MARKETPLACE),
                missing_fields=preview_missing_fields,
            )

        pass_count = sum(1 for item in results if item.get('status') == 'pass')
        warn_count = sum(1 for item in results if item.get('status') == 'warn')
        fail_count = sum(1 for item in results if item.get('status') == 'fail')
        result_payload = {
            'input_file': input_file,
            'account': account_name,
            'context_message': context_message,
            'template': template_definition_summary(template_definition),
            'column_issues': column_issues,
            'pass': pass_count,
            'warn': warn_count,
            'fail': fail_count,
            'total': len(results),
            'results': results,
        }
        message = f'诊断完成：{pass_count} 通过，{warn_count} 警告，{fail_count} 失败'
        _complete_task_record(task_id, message=message, result=result_payload, total=len(rows), progress=len(rows), current_item='诊断完成')
    except Exception as exc:
        _append_task_log(task_id, f'❌ {exc}')
        _fail_task_record(task_id, str(exc))


def _execute_template_generation(task_id: str, product_type: str, variation_mode: str):
    stages = ['拉取类目定义', '生成模板']
    try:
        _set_task_stage(task_id, stages[0], 1, stages, progress=0, total=2, current_item=product_type)
        definition = ensure_template_definition(
            product_type=product_type,
            marketplace=DEFAULT_MARKETPLACE,
            variation_mode=variation_mode,
        )
        _set_task_stage(task_id, stages[1], 2, stages, progress=1, total=2, current_item='写入 Excel 模板')
        workbook_path, definition = ensure_template_workbook(definition['template_id'])
        filename = os.path.basename(workbook_path)
        result_payload = {
            'template_id': definition['template_id'],
            'filename': filename,
            'download_url': f"/api/templates/{definition['template_id']}/download",
            'product_type': definition.get('product_type', ''),
            'marketplace': definition.get('marketplace', DEFAULT_MARKETPLACE),
            'variation_mode': definition.get('variation_mode', variation_mode),
            'required_total': definition.get('required_total', 0),
            'recommended_total': definition.get('recommended_total', 0),
            'column_count': len(definition.get('columns', []) or []),
        }
        _complete_task_record(
            task_id,
            message=f"模板已生成：{product_type}",
            result=result_payload,
            progress=2,
            total=2,
            current_item='模板已生成',
            result_file=workbook_path,
        )
    except Exception as exc:
        _append_task_log(task_id, f'❌ {exc}')
        _fail_task_record(task_id, str(exc))


def _run_submit_operation(input_file: str, skus: list, preview: bool = True, account_id: str = '',
                          progress_callback=None) -> dict:
    processor = ExcelProcessor()
    all_data = processor.read_input(input_file)
    col_map = processor.detect_columns()
    headers = list(processor.headers)

    sku_col = col_map.get('sku', '')
    sku_set = set(skus or [])
    matched = [item for item in all_data if str(item.get(sku_col, '') or item.get('SKU', '')).strip() in sku_set]
    if not matched:
        raise ValueError(f'未匹配到{len(skus)}个SKU')

    def emit(stage_name: str, done: int, total: int, current_item: str):
        if progress_callback:
            progress_callback(stage_name, done, total, current_item)

    results = []

    if preview:
        emit('读取商品文件', 0, len(matched), '准备预览验证')
        from amazon.accounts import AccountManager
        from amazon.auth import AmazonAuth
        from amazon.listings import ListingsAPI
        from amazon.mapper import FieldMapper

        mgr = AccountManager()
        acc = mgr.get_account(account_id) if account_id else mgr.get_default_account()
        mapper = FieldMapper(acc.get('marketplace_id', 'ATVPDKIKX0DER')) if acc else FieldMapper()

        listings_api = None
        if acc and all([acc.get('lwa_client_id'), acc.get('lwa_client_secret'), acc.get('refresh_token')]):
            auth = AmazonAuth(
                client_id=acc['lwa_client_id'],
                client_secret=acc['lwa_client_secret'],
                refresh_token=acc['refresh_token'],
            )
            listings_api = ListingsAPI(
                auth=auth,
                seller_id=acc['seller_id'],
                marketplace_id=acc.get('marketplace_id', 'ATVPDKIKX0DER'),
            )

        total = len(matched)
        for idx, item in enumerate(matched, start=1):
            product = mapper.map_excel_row(item, col_map)
            sku = product.get('sku', '')
            emit('本地字段校验', idx - 1, total, sku)
            validation = mapper.validate_required_fields(product)

            if not validation['valid'] or listings_api is None:
                results.append({
                    'sku': sku,
                    'valid': validation['valid'],
                    'errors': validation['errors'],
                    'warnings': validation['warnings'],
                    'source': 'local',
                })
                continue

            emit('调用 Amazon 预览', idx - 1, total, sku)
            preview_result = listings_api.put_listings_item(sku, product, preview=True)
            issues = preview_result.get('issues', []) or []
            errors = list(validation['errors'])
            warnings = list(validation['warnings'])

            for issue in issues:
                message = issue.get('message', '')
                severity = str(issue.get('severity', '')).upper()
                if severity == 'ERROR':
                    errors.append(message)
                else:
                    warnings.append(message)

            preview_status = str(preview_result.get('status', '')).upper()
            is_valid = preview_status == 'VALID' and len(errors) == 0

            results.append({
                'sku': sku,
                'valid': is_valid,
                'errors': errors,
                'warnings': warnings,
                'source': 'amazon_preview',
                'status': preview_status or 'UNKNOWN',
            })
            time.sleep(1.0)

        valid_count = sum(1 for r in results if r.get('valid'))
        if input_file and os.path.exists(input_file):
            _apply_preview_results_to_file(
                input_file=input_file,
                headers=headers,
                results=results,
                account_name=acc.get('name', acc.get('seller_id', '')) if acc else '',
            )
        return {
            'success': True,
            'mode': 'preview',
            'total': len(results),
            'valid': valid_count,
            'invalid': len(results) - valid_count,
            'results': results,
            'message': f'预校验完成: {valid_count}/{len(results)} 个商品可提交',
        }

    emit('读取商品文件', 0, len(matched), '准备正式提交')
    from amazon.accounts import AccountManager
    from amazon.auth import AmazonAuth
    from amazon.listings import ListingsAPI
    from amazon.mapper import FieldMapper

    mgr = AccountManager()
    acc = mgr.get_account(account_id) if account_id else mgr.get_default_account()
    if not acc:
        raise ValueError('未配置亚马逊账号，请先在设置页面添加账号')
    mapper = FieldMapper(acc.get('marketplace_id', 'ATVPDKIKX0DER'))

    auth = AmazonAuth(
        client_id=acc['lwa_client_id'],
        client_secret=acc['lwa_client_secret'],
        refresh_token=acc['refresh_token'],
    )
    listings_api = ListingsAPI(
        auth=auth,
        seller_id=acc['seller_id'],
        marketplace_id=acc.get('marketplace_id', 'ATVPDKIKX0DER'),
    )

    total = len(matched)
    precheck_results = []
    for idx, item in enumerate(matched, start=1):
        product = mapper.map_excel_row(item, col_map)
        sku = product.get('sku', '')
        result_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        emit('提交门禁检查', idx - 1, total, sku)
        precheck_result = _run_listing_check_for_product(
            product,
            mapper=mapper,
            listings_api=listings_api,
            include_listing_lookup=False,
            include_media_probe=False,
        )
        precheck_results.append(precheck_result)
        if precheck_result.get('status') == 'fail':
            results.append({
                'sku': sku,
                'status': 'PREVIEW_BLOCKED',
                'submit_time': result_time,
                'message': precheck_result.get('summary_text', '') or '提交前诊断未通过',
                'issues': [{'severity': 'ERROR', 'message': precheck_result.get('summary_text', '') or '提交前诊断未通过'}],
                })
            continue

        emit('调用 Amazon 正式提交', idx - 1, total, sku)
        try:
            submit_result = listings_api.put_listings_item(sku, product, preview=False)
            status = submit_result.get('status', 'UNKNOWN')
            result_entry = {
                'sku': sku,
                'status': status,
                'submit_time': result_time,
                'submission_id': submit_result.get('submissionId', ''),
                'issues': submit_result.get('issues', []),
            }
            for ident in submit_result.get('identifiers', []):
                if ident.get('asin'):
                    result_entry['asin'] = ident['asin']
                    break
            results.append(result_entry)
            time.sleep(0.5)
        except Exception as submit_err:
            logger.error(f'SKU {sku} 提交异常: {submit_err}')
            results.append({
                'sku': sku,
                'status': 'ERROR',
                'submit_time': result_time,
                'message': str(submit_err),
            })

    _apply_listing_check_results_to_file(
        input_file=input_file,
        headers=headers,
        results=precheck_results,
        account_name=acc.get('name', acc.get('seller_id', '')),
    )

    accepted = sum(1 for r in results if r.get('status') == 'ACCEPTED')
    failed = sum(1 for r in results if r.get('status') in ('ERROR', 'INVALID', 'SKIPPED', 'PREVIEW_BLOCKED'))

    submission_record = {
        'timestamp': datetime.now().isoformat(),
        'account': acc.get('name', acc.get('seller_id', '')),
        'total': len(results),
        'accepted': accepted,
        'failed': failed,
        'results': results,
    }

    submissions_dir = os.path.join(config.OUTPUT_DIR, 'submissions')
    os.makedirs(submissions_dir, exist_ok=True)
    record_file = os.path.join(submissions_dir, f'submit_{time.strftime("%Y%m%d_%H%M%S")}.json')
    with open(record_file, 'w', encoding='utf-8') as f:
        json.dump(submission_record, f, ensure_ascii=False, indent=2)

    persist_warning = ''
    try:
        updates_by_sku = {}
        for result_entry in results:
            sku = str(result_entry.get('sku', '') or '').strip()
            if not sku:
                continue
            submit_time = str(result_entry.get('submit_time', '') or datetime.now().strftime('%Y-%m-%d %H:%M:%S')).strip()
            updates_by_sku[sku] = _build_submit_persist_updates(headers, result_entry, submit_time)
        if updates_by_sku:
            _persist_bulk_row_updates(input_file, updates_by_sku)
    except Exception as persist_err:
        persist_warning = f'提交结果已返回，但写回Excel失败: {persist_err}'
        logger.error(persist_warning)

    return {
        'success': True,
        'mode': 'submit',
        'total': len(results),
        'accepted': accepted,
        'failed': failed,
        'results': results,
        'record_file': record_file,
        'persist_warning': persist_warning,
        'message': f'提交完成: {accepted}个已接受, {failed}个失败',
    }


def _execute_submit_task(task_id: str, input_file: str, skus: list, preview: bool, account_id: str):
    stages = ['读取商品文件', '本地字段校验', '调用 Amazon 预览'] if preview else ['读取商品文件', '提交门禁检查', '调用 Amazon 正式提交']

    def on_progress(stage_name: str, done: int, total: int, current_item: str):
        stage_index = stages.index(stage_name) + 1 if stage_name in stages else 1
        _set_task_stage(task_id, stage_name, stage_index, stages, progress=done, total=total, current_item=current_item)

    try:
        result = _run_submit_operation(
            input_file=input_file,
            skus=skus,
            preview=preview,
            account_id=account_id,
            progress_callback=on_progress,
        )
        total = int(result.get('total', len(skus)) or len(skus) or 0)
        _complete_task_record(
            task_id,
            message=result.get('message', '执行完成'),
            result=result,
            progress=total,
            total=total,
            current_item='完成',
        )
    except Exception as exc:
        _fail_task_record(task_id, str(exc))


# ===== Schema 驱动的动态模板 API =====

@app.route('/api/templates/recommend', methods=['POST'])
def api_template_recommend():
    """根据链接/标题/关键词推荐美国站 product type。"""
    data = request.json or {}
    try:
        result = recommend_product_types(
            source_url=str(data.get('source_url', '') or data.get('url', '') or '').strip(),
            title=str(data.get('title', '') or '').strip(),
            keyword=str(data.get('keyword', '') or '').strip(),
            marketplace=DEFAULT_MARKETPLACE,
        )
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/templates/generate', methods=['POST'])
def api_templates_generate():
    """异步生成美国站官方模板。"""
    data = request.json or {}
    product_type = str(data.get('product_type', '') or '').strip()
    variation_mode = str(data.get('variation_mode', 'single') or 'single').strip().lower()
    variation_mode = 'variation' if variation_mode == 'variation' else 'single'
    if not product_type:
        return jsonify({'error': '请先选择 product_type'}), 400

    task = _start_task_record(
        kind='template_generate',
        title='生成官方模板',
        input_file='',
        status='running',
        progress=0,
        total=2,
        stage_name='等待开始',
        stage_index=0,
        stages=['拉取类目定义', '生成模板'],
        product_type=product_type,
        marketplace=DEFAULT_MARKETPLACE,
        variation_mode=variation_mode,
        result={},
    )
    thread = threading.Thread(
        target=_execute_template_generation,
        args=(task['id'], product_type, variation_mode),
        daemon=True,
    )
    thread.start()
    return jsonify({'success': True, 'task_id': task['id']})


@app.route('/api/templates/<template_id>/download')
def api_download_generated_template(template_id):
    """下载生成好的官方模板。"""
    try:
        workbook_path, definition = ensure_template_workbook(template_id)
        return send_file(
            workbook_path,
            as_attachment=True,
            download_name=os.path.basename(workbook_path),
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@app.route('/api/template-upload', methods=['POST'])
def api_template_upload():
    """上传模板结果文件并自动启动诊断。"""
    if 'file' not in request.files:
        return jsonify({'error': '请选择文件'}), 400

    upload = request.files['file']
    original_filename = upload.filename or ''
    if not original_filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': '仅支持 .xlsx 或 .xls 文件'}), 400

    os.makedirs(config.INPUT_DIR, exist_ok=True)
    filepath = _unique_upload_path(config.INPUT_DIR, original_filename)
    upload.save(filepath)

    try:
        processor = ExcelProcessor()
        rows = processor.read_input(filepath)
        col_map = processor.detect_columns()
        meta, template_definition = _load_template_definition_for_file(filepath)
        task = _start_task_record(
            kind='template_diagnose',
            title='模板自动诊断',
            input_file=filepath,
            status='running',
            progress=0,
            total=max(len(rows), 1),
            stage_name='等待开始',
            stage_index=0,
            stages=['解析上传文件', '匹配字段', '本地校验', 'Amazon 预览', '生成诊断结果'],
            result={},
        )
        thread = threading.Thread(
            target=_execute_template_diagnosis,
            args=(task['id'], filepath, str(request.form.get('account_id', '') or '').strip(), []),
            daemon=True,
        )
        thread.start()
        return jsonify({
            'success': True,
            'filename': os.path.basename(filepath),
            'filepath': filepath,
            'total_rows': len(rows),
            'column_mapping': col_map,
            'template': template_definition_summary(template_definition),
            'template_meta': meta,
            'task_id': task['id'],
        })
    except Exception as e:
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'error': f'读取模板失败: {str(e)}'}), 400


@app.route('/api/template-diagnose', methods=['POST'])
def api_template_diagnose():
    """对已上传模板文件执行自动诊断。"""
    data = request.json or {}
    input_file = str(data.get('input_file', '') or data.get('file', '') or '').strip()
    account_id = str(data.get('account_id', '') or '').strip()
    skus = [str(sku or '').strip() for sku in (data.get('skus', []) or []) if str(sku or '').strip()]
    if not input_file or not os.path.exists(input_file):
        return jsonify({'error': '输入文件不存在'}), 400

    task = _start_task_record(
        kind='template_diagnose',
        title='模板自动诊断',
        input_file=input_file,
        status='running',
        progress=0,
        total=max(len(skus), 1),
        stage_name='等待开始',
        stage_index=0,
        stages=['解析上传文件', '匹配字段', '本地校验', 'Amazon 预览', '生成诊断结果'],
        result={},
    )
    thread = threading.Thread(
        target=_execute_template_diagnosis,
        args=(task['id'], input_file, account_id, skus),
        daemon=True,
    )
    thread.start()
    return jsonify({'success': True, 'task_id': task['id']})


@app.route('/api/product-types')
def api_product_types():
    """搜索产品类型"""
    keyword = request.args.get('keyword', '').strip()
    if not keyword:
        return jsonify({'error': '请输入搜索关键词'}), 400
    try:
        from amazon.sp_client import SPClient
        client = SPClient(marketplace_id='ATVPDKIKX0DER')
        types = client.search_product_types(keyword)
        return jsonify({'productTypes': types})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/schema')
def api_schema():
    """获取产品类型的字段清单"""
    product_type = request.args.get('product_type', '').strip()
    marketplace = request.args.get('marketplace', 'US').strip()
    if not product_type:
        return jsonify({'error': '请指定产品类型'}), 400
    try:
        from amazon.schema_manager import fetch_schema, parse_schema
        raw = fetch_schema(product_type, marketplace)
        fields = parse_schema(raw)
        return jsonify({
            'product_type': product_type,
            'marketplace': marketplace,
            **fields,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-template', methods=['POST'])
def api_generate_template():
    """兼容旧接口：同步生成 Excel 模板并返回下载链接。"""
    data = request.json or {}
    product_type = data.get('product_type', '').strip()
    marketplace = DEFAULT_MARKETPLACE
    variation_mode = str(data.get('variation_mode', 'single') or 'single').strip().lower()
    variation_mode = 'variation' if variation_mode == 'variation' else 'single'
    if not product_type:
        return jsonify({'error': '请指定产品类型'}), 400
    try:
        definition = ensure_template_definition(product_type, marketplace, variation_mode)
        output_path, definition = ensure_template_workbook(definition['template_id'])
        filename = os.path.basename(output_path)

        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f"/api/templates/{definition['template_id']}/download",
            'template_id': definition['template_id'],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export', methods=['POST'])
def export_excel():
    """导出处理后的Excel — 支持SP-API格式和对比格式"""
    try:
        data = request.json
        input_file = data.get('input_file', '') or data.get('file', '')
        export_format = str(data.get('format', 'comparison') or 'comparison').strip()
        export_format = {
            'sp_api': 'sp-api',
            'sp-api': 'sp-api',
            'comparison': 'comparison',
        }.get(export_format, export_format)
        selected_skus = data.get('selected_skus', [])

        if not input_file or not os.path.exists(input_file):
            return jsonify({'error': '输入文件不存在'}), 400

        if export_format not in ('sp-api', 'comparison'):
            return jsonify({'error': f'无效的导出格式: {export_format}，支持 sp-api 或 comparison'}), 400

        # 读取数据
        processor = ExcelProcessor()
        all_data = processor.read_input(input_file)
        col_map = processor.detect_columns()

        # 按SKU过滤
        if selected_skus:
            sku_col = col_map.get('sku', '')
            sku_set = set(selected_skus)
            export_data = [item for item in all_data
                           if str(item.get(sku_col, '') or item.get('SKU', '')).strip() in sku_set]
            if not export_data:
                return jsonify({'error': '未匹配到选中的SKU'}), 400
        else:
            export_data = all_data

        timestamp = time.strftime('%Y%m%d_%H%M%S')
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)

        if export_format == 'comparison':
            # 对比格式 — 使用 write_comparison_output
            filename = f'对比导出_{timestamp}.xlsx'
            output_path = os.path.join(config.OUTPUT_DIR, filename)
            processor.write_comparison_output(export_data, output_path, col_map)

        elif export_format == 'sp-api':
            # SP-API格式 — 映射为SP-API字段结构的平面Excel
            from amazon.mapper import FieldMapper
            mapper = FieldMapper()

            sp_data = []
            sp_headers = ['sku', 'product_type', 'item_name', 'brand',
                          'bullet_point_1', 'bullet_point_2', 'bullet_point_3',
                          'bullet_point_4', 'bullet_point_5',
                          'product_description', 'generic_keywords',
                          'main_image_url', 'other_image_url_1', 'other_image_url_2',
                          'other_image_url_3', 'other_image_url_4',
                          'standard_price', 'quantity', 'condition_type',
                          'fulfillment_channel', 'upc']

            for item in export_data:
                mapped = mapper.map_excel_row(item, col_map)
                sp_row = {}
                for h in sp_headers:
                    sp_row[h] = ''  # 确保所有列存在

                sp_row['sku'] = mapped.get('sku', '')
                sp_row['product_type'] = mapped.get('product_type', '')
                sp_row['item_name'] = mapped.get('title', '')
                sp_row['brand'] = mapped.get('brand', '')
                for i in range(1, 6):
                    sp_row[f'bullet_point_{i}'] = mapped.get(f'bullet_point_{i}', '')
                sp_row['product_description'] = mapped.get('description', '')
                sp_row['generic_keywords'] = mapped.get('keywords', '')
                sp_row['main_image_url'] = mapped.get('main_image_url', '')
                for i in range(1, 5):
                    sp_row[f'other_image_url_{i}'] = mapped.get(f'other_image_{i}', '')
                sp_row['standard_price'] = mapped.get('price', '')
                sp_row['quantity'] = mapped.get('quantity', '')
                sp_row['condition_type'] = mapped.get('condition_type', '')
                sp_row['fulfillment_channel'] = mapped.get('fulfillment_channel', '')
                sp_row['upc'] = mapped.get('upc', '')

                sp_data.append(sp_row)

            # 用write_output写出SP-API格式的平面Excel
            filename = f'SP-API导出_{timestamp}.xlsx'
            output_path = os.path.join(config.OUTPUT_DIR, filename)

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "SP-API Listing Data"

            # 写表头
            for col_idx, header in enumerate(sp_headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写数据
            for row_idx, row_data in enumerate(sp_data, start=2):
                for col_idx, header in enumerate(sp_headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

            wb.save(output_path)

        logger.info(f'✅ 导出完成: {filename} ({len(export_data)} 条)')
        _record_instant_task(
            kind='export',
            title='导出SP-API模板' if export_format == 'sp-api' else '导出对比表',
            status='completed',
            input_file=input_file,
            total=len(export_data),
            result_file=output_path,
            message=f'已导出 {len(export_data)} 条商品',
        )
        return jsonify({
            'success': True,
            'filename': filename,
            'download_url': f'/api/download/{filename}',
            'total_exported': len(export_data),
        })

    except Exception as e:
        logger.error(f'导出失败: {e}')
        import traceback
        logger.error(traceback.format_exc())
        _record_instant_task(
            kind='export',
            title='导出失败',
            status='failed',
            input_file=data.get('input_file', '') or data.get('file', '') if isinstance(data, dict) else '',
            message=str(e),
            error=str(e),
        )
        return jsonify({'error': str(e)}), 500


@app.route('/api/download-template')
def download_template():
    """下载标准采集模板Excel"""
    template_name = 'amazon_listing_template.xlsx'
    template_path = os.path.join(config.OUTPUT_DIR, template_name)

    if not os.path.exists(template_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "商品采集模板"
        headers = [
            'SKU', 'item_name', 'brand_name', 'product_description',
            'bullet_point_1', 'bullet_point_2', 'bullet_point_3',
            'bullet_point_4', 'bullet_point_5',
            'generic_keywords', 'main_image_url',
            'other_image_url_1', 'other_image_url_2', 'other_image_url_3',
            'other_image_url_4',
            'standard_price', 'quantity', 'condition_type',
            'fulfillment_channel', 'product_id', 'product_id_type',
            'product_identity_mode',
            'product_type', 'parent_sku', 'parent_child',
            'variation_theme', 'color_name', 'size_name',
            'material_type', 'manufacturer', 'part_number',
            'country_of_origin', 'item_weight', 'item_length',
            'item_width', 'item_height',
            'are_batteries_included', 'batteries_required',
            'battery_type',
        ]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)
        wb.save(template_path)

    return send_file(template_path, as_attachment=True, download_name=template_name)


@app.route('/api/accounts', methods=['DELETE'])
def delete_account():
    """删除亚马逊账号"""
    try:
        from amazon.accounts import AccountManager
        mgr = AccountManager()
        seller_id = request.json.get('seller_id')
        if not seller_id:
            return jsonify({'error': '缺少seller_id'}), 400
        success = mgr.remove_account(seller_id)
        if success:
            return jsonify({'success': True, 'message': '账号已删除'})
        else:
            return jsonify({'error': '账号不存在'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/submit', methods=['POST'])
def submit_to_amazon():
    """提交商品到亚马逊 (预览验证或正式提交)"""
    try:
        data = request.json
        skus = data.get('skus', [])
        input_file = data.get('file', '')
        preview = data.get('preview', True)
        account_id = data.get('account_id', '')

        if not skus:
            return jsonify({'error': '没有选中的SKU'}), 400
        if not input_file or not os.path.exists(input_file):
            return jsonify({'error': '文件不存在'}), 400
        payload = _run_submit_operation(
            input_file=input_file,
            skus=skus,
            preview=preview,
            account_id=account_id,
        )
        if preview:
            _record_instant_task(
                kind='preview_submit',
                title='Amazon 预览验证',
                status='completed',
                input_file=input_file,
                total=payload.get('total', 0),
                success=payload.get('valid', 0),
                failed=payload.get('invalid', 0),
                account=account_id,
                message=payload.get('message', ''),
            )
        else:
            _record_instant_task(
                kind='submit',
                title='正式提交到亚马逊',
                status='completed' if not payload.get('failed') else ('failed' if payload.get('accepted', 0) == 0 else 'completed'),
                input_file=input_file,
                total=payload.get('total', 0),
                success=payload.get('accepted', 0),
                failed=payload.get('failed', 0),
                account=account_id,
                result_file=payload.get('record_file', ''),
                message=payload.get('message', ''),
                warning=payload.get('persist_warning', ''),
            )
        return jsonify(payload)

    except Exception as e:
        logger.error(f'提交失败: {e}')
        _record_instant_task(
            kind='submit' if not preview else 'preview_submit',
            title='正式提交到亚马逊' if not preview else 'Amazon 预览验证',
            status='failed',
            input_file=input_file,
            message=str(e),
            error=str(e),
        )
        return jsonify({'error': str(e)}), 500


@app.route('/api/submit-task', methods=['POST'])
def submit_to_amazon_task():
    """异步启动 Amazon 预览验证或正式提交任务。"""
    data = request.json or {}
    skus = data.get('skus', []) or []
    input_file = data.get('file', '')
    preview = bool(data.get('preview', True))
    account_id = data.get('account_id', '')

    if not skus:
        return jsonify({'error': '没有选中的SKU'}), 400
    if not input_file or not os.path.exists(input_file):
        return jsonify({'error': '文件不存在'}), 400

    title = 'Amazon 预览验证' if preview else '正式提交到亚马逊'
    stages = ['解析文件', '本地校验', 'Amazon 预览'] if preview else ['解析文件', '提交门禁检查', '正式提交']
    task = _start_task_record(
        kind='preview_submit' if preview else 'submit',
        title=title,
        input_file=input_file,
        status='running',
        progress=0,
        total=len(skus),
        stage_name=stages[0],
        stage_index=1,
        stages=stages,
        result={},
    )

    thread = threading.Thread(
        target=_execute_submit_task,
        args=(task['id'], input_file, skus, preview, account_id),
        daemon=True,
    )
    thread.start()
    return jsonify({'success': True, 'task_id': task['id']})


@app.route('/api/submission-history')
def submission_history():
    """查看提交历史记录"""
    submissions_dir = os.path.join(config.OUTPUT_DIR, 'submissions')
    if not os.path.exists(submissions_dir):
        return jsonify({'records': []})

    records = []
    for f in sorted(os.listdir(submissions_dir), reverse=True):
        if f.endswith('.json'):
            filepath = os.path.join(submissions_dir, f)
            try:
                with open(filepath, 'r', encoding='utf-8') as fh:
                    data = json.load(fh)
                records.append({
                    'file': f,
                    'timestamp': data.get('timestamp', ''),
                    'account': data.get('account', ''),
                    'total': data.get('total', 0),
                    'accepted': data.get('accepted', 0),
                    'failed': data.get('failed', 0),
                })
            except Exception:
                pass

    return jsonify({'records': records[:50]})


@app.route('/api/update-field', methods=['POST'])
def update_field():
    """更新单个商品字段 (内联编辑)"""
    try:
        data = request.json
        input_file = data.get('file', '')
        sku = data.get('sku', '')
        field = data.get('field', '')
        value = data.get('value', '')

        if not all([input_file, sku, field]):
            return jsonify({'error': '参数不完整'}), 400
        if not os.path.exists(input_file):
            return jsonify({'error': '文件不存在'}), 400

        processor = ExcelProcessor()
        processor.read_input(input_file)
        col_map = processor.detect_columns()
        header_name = _logical_field_to_excel_header(field, col_map)
        _persist_row_updates(input_file, sku, {header_name: value})
        return jsonify({'success': True, 'message': f'{sku}.{header_name} 已更新为 {value}'})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/update-product', methods=['POST'])
def update_product():
    """批量更新单个商品的多个字段。"""
    try:
        data = request.json
        input_file = data.get('file', '')
        sku = data.get('sku', '')
        updates = dict(data.get('updates') or {})

        if not input_file or not sku or not updates:
            return jsonify({'error': '参数不完整'}), 400
        if not os.path.exists(input_file):
            return jsonify({'error': '文件不存在'}), 400

        processor = ExcelProcessor()
        processor.read_input(input_file)
        col_map = processor.detect_columns()

        mapped_updates = {}
        for field, value in updates.items():
            if field is None or str(field).strip() == '':
                continue
            header_name = _logical_field_to_excel_header(str(field), col_map)
            mapped_updates[header_name] = value

        if not mapped_updates:
            return jsonify({'error': '没有可更新的字段'}), 400

        _persist_row_updates(input_file, sku, mapped_updates)
        return jsonify({'success': True, 'updated_fields': list(mapped_updates.keys())})

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ===== SP-API 品类字段注册表 =====

_field_registry = None

def _load_field_registry():
    """加载品类字段注册表"""
    global _field_registry
    if _field_registry is not None:
        return _field_registry
    registry_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                  'config', 'sp_api_fields.json')
    if os.path.exists(registry_path):
        with open(registry_path, 'r', encoding='utf-8') as f:
            _field_registry = json.load(f)
    else:
        _field_registry = {'categories': {}, 'cross_category_fields': []}
    return _field_registry


@app.route('/api/field-registry')
def get_field_registry():
    """返回完整品类字段注册表"""
    registry = _load_field_registry()
    return jsonify({'success': True, 'registry': registry})


@app.route('/api/category-fields')
def get_category_fields():
    """返回指定品类的字段列表"""
    category = request.args.get('category', '')
    registry = _load_field_registry()
    if not category:
        # 返回品类列表(不含具体字段)
        cats = {}
        for key, cat in registry.get('categories', {}).items():
            cats[key] = {'name_en': cat['name_en'], 'name_zh': cat['name_zh'], 'icon': cat.get('icon', ''),
                         'field_count': len(cat.get('fields', []))}
        return jsonify({'success': True, 'categories': cats,
                        'cross_category_fields': registry.get('cross_category_fields', [])})
    cat = registry.get('categories', {}).get(category)
    if not cat:
        return err(f'未知品类: {category}', 404)
    return jsonify({'success': True, 'category': category, 'fields': cat.get('fields', []),
                    'cross_category_fields': registry.get('cross_category_fields', [])})


@app.route('/api/selected-fields', methods=['GET', 'POST'])
def selected_fields():
    """获取/保存用户选择的品类字段配置"""
    config_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                'config', 'selected_fields.json')
    if request.method == 'GET':
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                return jsonify({'success': True, 'config': json.load(f)})
        return jsonify({'success': True, 'config': {'category': '', 'fields': []}})
    # POST - 保存
    data = request.json
    os.makedirs(os.path.dirname(config_path), exist_ok=True)
    with open(config_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return jsonify({'success': True, 'message': '品类字段配置已保存'})


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    os.makedirs(config.INPUT_DIR, exist_ok=True)
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    print("🚀 亚马逊商品处理工具 Web界面 V2")
    print(f"   访问: http://localhost:{config.WEB_PORT}")
    print("   新功能: 前后对比 | 多账号管理 | 字段验证 | 品类字段选择器")
    app.run(host='0.0.0.0', port=config.WEB_PORT, debug=config.WEB_DEBUG)
